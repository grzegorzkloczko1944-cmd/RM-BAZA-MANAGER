"""
Moduł importu BOM z Excel (LOGISTYKA_OUT.xlsx) lub CSV
Funkcje pomocnicze + konwerter CSV -> XLSX
"""
import csv
import re
import tempfile
import threading
import time
from pathlib import Path
from typing import Generator, Dict, Optional, Tuple
import openpyxl


def csv_to_xlsx(csv_path: Path, encoding: str = "auto") -> Path:
    """
    Konwertuje plik CSV na XLSX z arkuszem ZBIORCZY.
    
    Plik tymczasowy XLSX jest tworzony obok oryginału (ten sam katalog).
    
    Args:
        csv_path: Ścieżka do pliku CSV
        encoding: Kodowanie CSV ('auto' = próbuj utf-8-sig, cp1250, latin-1)
    
    Returns:
        Path do utworzonego pliku XLSX
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(f"Plik CSV nie istnieje: {csv_path}")
    
    # Wykryj kodowanie
    encodings = ["utf-8-sig", "utf-8", "cp1250", "latin-1"]
    if encoding != "auto":
        encodings = [encoding]
    
    rows = None
    used_encoding = None
    for enc in encodings:
        try:
            with open(csv_path, "r", encoding=enc, newline="") as f:
                # Wykryj separator (;  ,  \t)
                sample = f.read(4096)
                f.seek(0)
                sniffer = csv.Sniffer()
                try:
                    dialect = sniffer.sniff(sample, delimiters=";,\t")
                except csv.Error:
                    dialect = csv.excel
                    dialect.delimiter = ";"  # domyślnie średnik (PL)
                
                reader = csv.reader(f, dialect)
                rows = list(reader)
                used_encoding = enc
                break
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    if rows is None:
        raise ValueError(f"Nie udało się odczytać CSV żadnym kodowaniem: {csv_path.name}")
    
    if not rows:
        raise ValueError(f"Plik CSV jest pusty: {csv_path.name}")
    
    print(f"📄 CSV: {csv_path.name} ({len(rows)} wierszy, encoding={used_encoding})")
    
    # Utwórz XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZBIORCZY"
    
    for row in rows:
        ws.append(row)
    
    # Zapisz obok oryginału
    xlsx_path = csv_path.with_suffix(".xlsx")
    # Jeśli plik już istnieje, dodaj _csv suffix
    if xlsx_path.exists():
        xlsx_path = csv_path.with_name(csv_path.stem + "_csv.xlsx")
    
    wb.save(xlsx_path)
    wb.close()
    
    print(f"✅ CSV → XLSX: {xlsx_path.name} ({len(rows)} wierszy)")
    return xlsx_path


def norm(s) -> str:
    """Normalizacja tekstu: usuń nadmiarowe spacje, non-breaking spaces"""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\u00A0", " ")).strip()


def parse_thickness_mm(material_text: str) -> Optional[float]:
    """
    Wyciąga grubość z zapisu 'gr1,5mm' → 1.5
    
    Przykłady:
        "Blacha gr1,5mm"     → 1.5
        "STAL gr 2mm"        → 2.0
        "AL gr0,8 mm"        → 0.8
        "Blacha 3mm"         → None (brak "gr")
    
    Returns:
        float lub None jeśli nie znaleziono
    """
    if not material_text:
        return None
    
    # Regex: \bgr\s*([0-9]+(?:[\.,][0-9]+)?)\s*mm\b
    m = re.search(
        r"\bgr\s*([0-9]+(?:[\.,][0-9]+)?)\s*mm\b",
        str(material_text),
        flags=re.IGNORECASE
    )
    
    if not m:
        return None
    
    try:
        return float(m.group(1).replace(",", "."))
    except Exception:
        return None


def normalize_type_label(t: str) -> str:
    """
    Normalizuje opisy typu z Excela/DB do kanonicznych wartości w bazie:
    X, XX, Z, ZZ, STANDARD, ZNORMALIZOWANE, UNKNOWN.
    
    W LOGISTYKA_OUT.xlsx w kolumnie 'Typ' bywają etykiety opisowe (np. 'CIĘCIE (X)', 
    'CIĘCIE+GIĘCIE (XX)', 'ZŁOŻENIE (Z)', 'MODUŁ (ZZ)').
    
    UWAGA: NIE wolno wykrywać samej litery 'Z' regexem, bo psuje 'ZNORMALIZOWANE'.
    """
    tt = norm(t).upper()
    if not tt:
        return "UNKNOWN"
    
    # Najpierw pełne słowa (żeby nie psuć ZNORMALIZOWANE)
    if "ZNORMALIZOWANE" in tt:
        return "ZNORMALIZOWANE"
    
    if tt in ("STANDARD", "STAND."):
        return "STANDARD"
    
    if tt in ("UNKNOWN", "NIEZNANE"):
        return "UNKNOWN"
    
    # Kolejność ma znaczenie (XX przed X)
    if tt == "ZZ" or "(ZZ" in tt or "(ZZ)" in tt or "MODUŁ" in tt:
        return "ZZ"
    
    if tt == "Z" or "(Z" in tt or "(Z)" in tt or "ZŁOŻENIE" in tt:
        return "Z"
    
    if tt == "XX" or "(XX" in tt or "(XX)" in tt or ("GIĘCIE" in tt and "X" in tt):
        return "XX"
    
    if tt == "X" or "(X)" in tt or ("CIĘCIE" in tt and "GIĘCIE" not in tt and "XX" not in tt):
        return "X"
    
    return "UNKNOWN"


def infer_type_from_drawing_no(dn: str) -> str:
    """
    Inferencja typu z Nr rysunku.
    
    - XX: końcówka numeru → 'XX'
    - X:  końcówka numeru → 'X'
    - ZZ: końcówka numeru → 'ZZ'
    - Z:  końcówka numeru → 'Z'
    - pusty numer: 'ZNORMALIZOWANE'
    - reszta: 'STANDARD'
    """
    s = norm(dn).upper()
    if not s:
        return "ZNORMALIZOWANE"
    
    # Kolejność ma znaczenie: dłuższe sufiksy przed krótszymi
    if s.endswith("ZZ"):
        return "ZZ"
    if s.endswith("Z"):
        return "Z"
    if s.endswith("XX"):
        return "XX"
    if s.endswith("X"):
        return "X"
    
    return "STANDARD"


def iter_zbiorczy_data_rows(excel_path: Path) -> Generator[Dict[str, str], None, None]:
    """
    Czyta LOGISTYKA_OUT.xlsx → arkusz ZBIORCZY.
    
    Zwraca rekordy tylko dla wierszy danych, pomija:
      - nagłówki sekcji (ELEMENTY ...)
      - puste wiersze
      - powtórzone wiersze nagłówków kolumn
    
    Uwaga: elementy ZNORMALIZOWANE mogą mieć pusty 'Nr rysunku' – wtedy rekord nadal jest zwracany.
    
    Yields:
        Dict z kluczami: Nr rysunku, Nazwa, Opis, Ilość całkowita, Typ, Materiał, Dostawca, ...
    """
    # Szybsze podejście: czytaj wszystko do listy i natychmiast zamknij plik
    # To eliminuje problem z wolnym dostępem do komórek przy uszkodzonym formatowaniu
    
    results = []
    wb = None
    load_method = None
    
    try:
        # Próba załadowania workbooka - WIELOKROTNY FALLBACK
        print(f"📊 [iter_zbiorczy] Ładowanie {excel_path.name}...")
        
        # Próba 1: data_only=True (SZYBKIE - pełny random access do komórek)
        # UWAGA: NIE używać read_only=True! W trybie read_only ws[r] jest O(n²)!
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            load_method = "data_only=True"
            print(f"✅ [iter_zbiorczy] Załadowano: {load_method}")
        except Exception as e1:
            print(f"⚠️  [iter_zbiorczy] Próba 1 nieudana: {type(e1).__name__}")
            
            # Próba 2: domyślne parametry (ostatnia deska ratunku)
            try:
                wb = openpyxl.load_workbook(excel_path)
                load_method = "domyślne parametry"
                print(f"✅ [iter_zbiorczy] Załadowano: {load_method}")
            except Exception as e2:
                print(f"❌ [iter_zbiorczy] Wszystkie próby nieudane")
                print(f"   Błąd końcowy: {type(e2).__name__}: {str(e2)[:200]}")
                raise Exception(
                    f"Nie udało się otworzyć pliku Excel żadną metodą.\n"
                    f"Ostatni błąd: {type(e2).__name__}: {str(e2)[:200]}\n\n"
                    f"MOŻLIWE ROZWIĄZANIE:\n"
                    f"Plik może być uszkodzony lub zawierać niezgodne formatowanie.\n"
                    f"Spróbuj otworzyć plik w Microsoft Excel i zapisać ponownie."
                ) from e2
        
        if wb is None:
            raise Exception("Nie udało się otworzyć pliku Excel")
        
        # Znajdź arkusz ZBIORCZY
        if "ZBIORCZY" in wb.sheetnames:
            ws = wb["ZBIORCZY"]
        else:
            # Fallback - pierwszy arkusz
            ws = wb[wb.sheetnames[0]]
        
        print(f"📄 [iter_zbiorczy] Arkusz: {ws.title}, wierszy: {ws.max_row}")
        
        header_row = None
        colmap = {}
        
        def looks_like_header(row_vals):
            """Wykryj czy wiersz to nagłówek kolumn"""
            return (
                "Nr rysunku" in row_vals and
                "Nazwa" in row_vals and
                any(k in row_vals for k in (
                    "Ilość całkowita", "Ilość", "Ilość (BOM)", 
                    "Ilość BOM", "Ilość (zam.)", "Ilość zam."
                ))
            )
        
        max_r = ws.max_row or 0
        
        # CZYTAJ WSZYSTKO DO PAMIĘCI (szybkie)
        print(f"🔍 [iter_zbiorczy] Rozpoczynam pętlę przez {max_r} wierszy...")
        for r in range(1, max_r + 1):
            # Progress co 50 wierszy
            if r % 50 == 0:
                print(f"  📊 Wiersz: {r}/{max_r}, rekordów: {len(results)}")
            
            try:
                # Normalizuj wartości w wierszu
                row = [norm(c.value) for c in ws[r]]
                
                # Wykryj sekcję (ELEMENTY ..., PEŁNA ...)
                if row and row[0] and (
                    row[0].upper().startswith("ELEMENTY") or 
                    row[0].upper().startswith("PEŁNA")
                ):
                    # Reset nagłówka dla nowej sekcji
                    header_row = None
                    colmap = {}
                    continue
                
                # Wykryj nagłówek kolumn
                if looks_like_header(row):
                    header_row = r
                    colmap = {v: i + 1 for i, v in enumerate(row) if v}
                    
                    # Aliasy nagłówków ilości (różne wersje Excela)
                    if "Ilość" not in colmap:
                        for _k in ("Ilość (BOM)", "Ilość BOM", "Ilość całkowita"):
                            if _k in colmap:
                                colmap["Ilość"] = colmap[_k]
                                break
                    
                    # Dodatkowo: fallback dla różnych wariantów
                    if "Ilość (BOM)" not in colmap and "Ilość" in colmap:
                        colmap["Ilość (BOM)"] = colmap["Ilość"]
                    
                    # Aliasy dla kolumny modułu (Katalog w BOM, Moduł w eksporcie RM_BAZA)
                    if "Katalog" not in colmap and "Moduł" in colmap:
                        colmap["Katalog"] = colmap["Moduł"]
                    
                    # Aliasy dla kolumny Ilość (zam.) (różne warianty)
                    if "Ilość (zam.)" not in colmap:
                        for _k in ("Ilość zam.", "Ilość zamówiona", "Ilość zamówionych"):
                            if _k in colmap:
                                colmap["Ilość (zam.)"] = colmap[_k]
                                break
                    
                    continue
                
                # Jeśli nie mamy jeszcze nagłówka - pomiń
                if not header_row:
                    continue
                
                # Pomiń całkowicie puste wiersze
                if all(
                    (c is None or str(c).strip() == "") 
                    for c in [ws.cell(r, c).value for c in range(1, 10)]
                ):
                    continue
                
                # Funkcja pomocnicza - pobierz wartość z kolumny
                def get(colname):
                    c = colmap.get(colname)
                    return ws.cell(r, c).value if c else None
                
                # Buduj rekord
                rec = {
                    "Nr rysunku": norm(get("Nr rysunku")),
                    "Nazwa": norm(get("Nazwa")),
                    "Opis": norm(get("Opis")),
                    "Ilość całkowita": norm(
                        get("Ilość całkowita") if "Ilość całkowita" in colmap else get("Ilość")
                    ),
                    "Ilość (zam.)": norm(get("Ilość (zam.)")) if "Ilość (zam.)" in colmap else None,
                    "Typ": norm(get("Typ")),
                    "Materiał": norm(get("Materiał")),
                    "Dostawca": norm(get("Dostawca")),
                    "Pliki 3D": norm(get("Pliki 3D")),
                    "Katalog": norm(get("Katalog")),
                    "Status": norm(get("Status")),
                    "Uwagi": norm(get("Uwagi")),
                }
                
                # Odrzuć powtórzone nagłówki (czasem przy błędnych merge)
                if looks_like_header(list(rec.values())):
                    continue
                
                # Rekord musi mieć przynajmniej nazwę lub numer
                if not rec["Nr rysunku"] and not rec["Nazwa"]:
                    continue
                
                # Dodaj do listy wyników
                results.append(rec)
                
            except Exception as row_err:
                print(f"⚠️  [iter_zbiorczy] Błąd w wierszu {r}: {row_err}")
                continue
        
        print(f"✅ [iter_zbiorczy] Wczytano {len(results)} rekordów")
    
    finally:
        # ZAWSZE zamknij workbook
        if wb is not None:
            try:
                wb.close()
                print(f"📕 [iter_zbiorczy] Workbook zamknięty")
            except Exception as close_err:
                print(f"⚠️  [iter_zbiorczy] Błąd zamykania: {close_err}")
    
    # Zwróć jako generator (dla kompatybilności wstecznej)
    for rec in results:
        yield rec


# Cache lokalizacji folderow projektow: {(root, projekt) -> Path|None}.
# find_project_folder robi iterdir() po V:\ i V:\ZP (dysk sieciowy, setki
# folderow) - a odpowiedz praktycznie sie nie zmienia w trakcie sesji.
_project_folder_cache: Dict[Tuple[str, str], Optional[Path]] = {}
_project_folder_lock = threading.Lock()


def _reset_project_folder_cache() -> None:
    """Wyrzuc cache lokalizacji folderow projektow."""
    with _project_folder_lock:
        _project_folder_cache.clear()


def find_project_folder(v_drive_root: Path, project_name: str) -> Optional[Path]:
    """
    Znajduje folder projektu w V:\\ na podstawie nazwy projektu z RM_BAZA.

    Nazwa projektu w RM_BAZA (np. "2556 Olmaj") i nazwa folderu w V:\\
    (np. "2556 Olmaj" lub "2556 Olmaj Wciskarka") nie muszą być identyczne
    - dopasowanie po pierwszym członie (numerze/prefiksie projektu, do
    pierwszej spacji).

    Przeszukiwany jest też podkatalog V:\\ZP\\ - projekty ZP* leżą tam, a nie
    bezpośrednio w korzeniu (analogicznie do SERVER_DIR/ZP na serwerze).

    Returns:
        Path do folderu projektu, lub None jeśli nie znaleziono.
    """
    v_drive_root = Path(v_drive_root)
    if not v_drive_root.exists():
        return None

    project_name = norm(project_name)
    if not project_name:
        return None

    prefix = project_name.split(" ")[0].upper()
    if not prefix:
        return None

    cache_key = (str(v_drive_root), project_name.upper())
    with _project_folder_lock:
        if cache_key in _project_folder_cache:
            return _project_folder_cache[cache_key]

    result = _locate_project_folder(v_drive_root, project_name, prefix)
    with _project_folder_lock:
        _project_folder_cache[cache_key] = result
    return result


def _locate_project_folder(v_drive_root: Path, project_name: str, prefix: str) -> Optional[Path]:
    """Wlasciwe przeszukanie dyskow (bez cache) - patrz find_project_folder."""
    search_roots = [v_drive_root]
    zp_root = v_drive_root / "ZP"
    try:
        if zp_root.is_dir():
            search_roots.append(zp_root)
    except OSError:
        pass

    candidates = []
    for root in search_roots:
        try:
            candidates += [p for p in root.iterdir() if p.is_dir()]
        except OSError:
            continue

    if not candidates:
        return None

    # Dopasowanie dokładne po pełnej nazwie
    for p in candidates:
        if norm(p.name).upper() == project_name.upper():
            return p

    # Dopasowanie po prefiksie (pierwszy człon nazwy folderu)
    for p in candidates:
        folder_prefix = norm(p.name).split(" ")[0].upper()
        if folder_prefix == prefix:
            return p

    return None


def find_out_files(project_folder: Path) -> list:
    """Zwraca listę wszystkich plików *_OUT.xlsx w folderze projektu (nierekurencyjnie)."""
    project_folder = Path(project_folder)
    if not project_folder.exists():
        return []
    return sorted(project_folder.glob("*_OUT.xlsx"))


def find_assembly_tree_rows(out_path: Path) -> list:
    """
    Czyta arkusz "DRZEWKO TEKST" z pliku *_OUT.xlsx.

    Zwraca listę dictów: {poziom, nr_rysunku, nazwa, ilosc_lokalna,
    ilosc_calkowita, typ, sciezka} - sciezka to lista segmentów
    'Nr rysunku' od korzenia do danego elementu (włącznie).

    Zwraca [] jeśli arkusza brak lub plik nie da się otworzyć.
    """
    out_path = Path(out_path)
    rows = []
    wb = None
    try:
        wb = openpyxl.load_workbook(out_path, data_only=True)
        if "DRZEWKO TEKST" not in wb.sheetnames:
            return []
        ws = wb["DRZEWKO TEKST"]

        header = [norm(c.value) for c in ws[1]]
        colmap = {v: i for i, v in enumerate(header) if v}
        required = ("Poziom", "Nr rysunku", "Nazwa", "Ścieżka")
        if not all(k in colmap for k in required):
            return []

        for r in range(2, (ws.max_row or 1) + 1):
            vals = [c.value for c in ws[r]]
            if not vals or all(v is None for v in vals):
                continue

            def get(colname):
                idx = colmap.get(colname)
                return vals[idx] if idx is not None and idx < len(vals) else None

            nr = norm(get("Nr rysunku"))
            sciezka_raw = norm(get("Ścieżka"))
            if not nr or not sciezka_raw:
                continue

            rows.append({
                "poziom": get("Poziom"),
                "nr_rysunku": nr,
                "nazwa": norm(get("Nazwa")),
                "ilosc_lokalna": get("Ilość lokalna"),
                "ilosc_calkowita": get("Ilość całkowita"),
                "typ": norm(get("Typ")),
                "sciezka": [norm(seg) for seg in sciezka_raw.split(">")],
            })
    except Exception as e:
        print(f"⚠️  [find_assembly_tree_rows] Błąd czytania {out_path.name}: {e}")
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    return rows


# Indeks DWF folderu projektu: {(root, projekt) -> (czas_budowy, {PREFIKS: Path})}.
# Bez niego kazde zaznaczenie wiersza globowalo folder projektu po sieci (~34 ms).
# TTL jest krotki, wiec nowo dograny rysunek i tak pojawi sie po chwili - to
# zachowuje dotychczasowa obietnice "szuka na biezaco", placac za nia raz na
# PROJECT_INDEX_TTL zamiast przy kazdym klikniecu.
PROJECT_INDEX_TTL = 60.0  # sekundy
_project_dwf_index: Dict[Tuple[str, str], Tuple[float, Dict[str, Path]]] = {}
_project_index_lock = threading.Lock()


def _reset_project_dwf_index() -> None:
    """Wyrzuc indeksy folderow projektow - nastepne zapytanie przeskanuje od nowa."""
    with _project_index_lock:
        _project_dwf_index.clear()


def _build_project_dwf_index(project_folder: Path) -> Dict[str, Path]:
    """Mapa {PREFIKS numeru rysunku -> Path} dla DWF w folderze projektu.

    DWF leza bezposrednio w folderze projektu i w gałęziach "Rysunki ..." -
    reszta (Templates, Design Data, OldVersions, ...) to pliki robocze Inventora.

    W galezie "Rysunki ..." schodzimy REKURENCYJNIE: czesc projektow zagniezdza
    rysunki dodatkowo ("Rysunki/Rysunki 100/..."), a plaskie przeszukanie
    jednego poziomu gubilo takie pliki (np. 2610 Sigma - 75 rysunkow widocznych
    jako 8). Kolejnosc: plytsze wygrywa, wiec plik lezacy wyzej ma pierwszenstwo.
    """
    index: Dict[str, Path] = {}

    def _add_files(directory: Path) -> None:
        try:
            entries = list(directory.glob("*.dwf"))
        except OSError:
            return
        for p in entries:
            prefix = norm(p.stem).split(" ")[0].upper()
            if prefix and prefix not in index:
                index[prefix] = p

    # 1) rysunki lezace bezposrednio w folderze projektu
    _add_files(project_folder)

    # 2) galezie "Rysunki ..." - przechodzone wszerz, zeby plytsze wygralo
    try:
        queue = [
            p for p in project_folder.iterdir()
            if p.is_dir() and norm(p.name).upper().startswith("RYSUNKI")
        ]
    except OSError:
        return index

    while queue:
        current = queue.pop(0)
        _add_files(current)
        try:
            queue += [
                p for p in current.iterdir()
                if p.is_dir() and p.name.lower() not in LIBRARY_EXCLUDED_DIRS
            ]
        except OSError:
            continue

    return index


def find_dwf_for_drawing(v_drive_root, project_name: str, drawing_no: str) -> Optional[Path]:
    """
    Szuka pliku .dwf dla danego numeru rysunku w folderze projektu na V:\\.

    Pliki .dwf są nazwane "<Nr rysunku> <Nazwa>.dwf" (np.
    "DCR-100.01X Płyta główna.dwf") - dopasowanie po prefiksie nazwy pliku
    (do pierwszej spacji), bez rozróżniania wielkości liter.

    DWF leżą tylko bezpośrednio w folderze projektu i w podfolderach
    "Rysunki ..." - inne podfoldery (Templates, Design Data, ContentCenter,
    Workspace, Libraries, Importowane komponenty, oldversions, ...) to pliki
    robocze/biblioteczne Inventora i są pomijane.

    Lista plików jest indeksowana i odświeżana co PROJECT_INDEX_TTL sekund -
    nowy rysunek dograny na V:\\ pojawi się po tym czasie (albo od razu po
    _reset_project_dwf_index()).

    Returns:
        Path do pliku .dwf, lub None jeśli nie znaleziono (brak folderu V:\\,
        brak folderu projektu, lub brak pasującego pliku).
    """
    drawing_no_n = norm(drawing_no)
    if not drawing_no_n:
        return None

    cache_key = (str(v_drive_root), norm(project_name).upper())
    now = time.monotonic()

    with _project_index_lock:
        cached = _project_dwf_index.get(cache_key)
        if cached is not None and (now - cached[0]) < PROJECT_INDEX_TTL:
            index = cached[1]
        else:
            index = None

    if index is None:
        project_folder = find_project_folder(Path(v_drive_root), project_name)
        if not project_folder:
            return None
        index = _build_project_dwf_index(project_folder)
        with _project_index_lock:
            _project_dwf_index[cache_key] = (now, index)

    return index.get(drawing_no_n.upper())


# Katalogi biblioteki pomijane przy szukaniu DWF - stare wersje, pliki robocze
# i systemowe Inventora. Ta sama lista co w skanowaniu biblioteki w GUI.
LIBRARY_EXCLUDED_DIRS = {
    "oldversions",
    "design data",
    "importowane komponenty",
    "templates",
    "$recycle.bin",
    "system volume information",
    "@recycle",
    ".git",
    ".svn",
    "node_modules",
}

DEFAULT_LIBRARY_ROOT = "B:/"

# Cache indeksu biblioteki: {prefiks numeru rysunku -> Path do .dwf}. Biblioteka
# to ~1600 plikow DWF w glebokim drzewie na dysku sieciowym - pelny przemarsz
# przy kazdym zaznaczeniu wiersza dawalby sekundy zwloki. Indeks budowany raz na
# sesje; unieważnia go _reset_library_dwf_index() (przycisk odswiezania w GUI).
_library_dwf_index: Optional[Dict[str, Path]] = None
_library_index_lock = threading.Lock()
_library_index_thread: Optional[threading.Thread] = None


def _reset_library_dwf_index() -> None:
    """Wyrzuc zbudowany indeks - kolejne zapytanie przeskanuje biblioteke na nowo."""
    global _library_dwf_index
    with _library_index_lock:
        _library_dwf_index = None


def prewarm_library_dwf_index(library_root=None) -> None:
    """Zbuduj indeks biblioteki w tle (nieblokujaco).

    Pelny przemarsz B:\\ to ~7s na dysku sieciowym. Wywolane przy starcie GUI
    sprawia, ze pierwsze zaznaczenie wiersza nie zamraza okna - do czasu
    gotowosci indeksu miniatury pozycji bibliotecznych sa po prostu pomijane.
    """
    global _library_index_thread

    with _library_index_lock:
        if _library_dwf_index is not None:
            return
        if _library_index_thread is not None and _library_index_thread.is_alive():
            return

        def _worker():
            global _library_dwf_index
            root = Path(library_root) if library_root else Path(DEFAULT_LIBRARY_ROOT)
            try:
                if not root.exists():
                    return
                built = _build_library_dwf_index(root)
            except Exception:
                return
            with _library_index_lock:
                _library_dwf_index = built

        _library_index_thread = threading.Thread(
            target=_worker, daemon=True, name="library-dwf-index"
        )
        _library_index_thread.start()


def _build_library_dwf_index(library_root: Path) -> Dict[str, Path]:
    """Zbuduj mape {PREFIKS -> Path} dla wszystkich .dwf w bibliotece.

    Przy duplikatach numeru rysunku wygrywa plik nowszy (mtime) - w bibliotece
    ten sam detal bywa w kilku podkatalogach tematycznych.
    """
    index: Dict[str, Path] = {}
    mtimes: Dict[str, float] = {}

    stack = [library_root]
    while stack:
        current = stack.pop()
        try:
            entries = list(current.iterdir())
        except OSError:
            continue

        for entry in entries:
            try:
                if entry.is_dir():
                    if entry.name.lower() not in LIBRARY_EXCLUDED_DIRS:
                        stack.append(entry)
                    continue
                if entry.suffix.lower() != ".dwf":
                    continue
            except OSError:
                continue

            prefix = norm(entry.stem).split(" ")[0].upper()
            if not prefix:
                continue
            try:
                mtime = entry.stat().st_mtime
            except OSError:
                mtime = 0.0
            if prefix not in index or mtime > mtimes.get(prefix, 0.0):
                index[prefix] = entry
                mtimes[prefix] = mtime

    return index


def find_dwf_in_library(drawing_no: str, library_root=None) -> Optional[Path]:
    """Szuka pliku .dwf dla numeru rysunku w bibliotece (B:\\).

    Pozycje biblioteczne (dwf_biblioteka=1) nie leza w folderze projektu na
    V:\\, tylko w bibliotece - stad osobne wyszukiwanie. Konwencja nazw ta sama
    co w projektach: "<Nr rysunku> <Nazwa>.dwf", dopasowanie po prefiksie do
    pierwszej spacji, bez rozroznienia wielkosci liter.

    Returns:
        Path do pliku .dwf, albo None gdy brak biblioteki lub pasujacego pliku.
    """
    drawing_no_n = norm(drawing_no)
    if not drawing_no_n:
        return None

    index = _library_dwf_index
    if index is None:
        # Indeks jeszcze niegotowy - zamow budowe w tle i odpusc ten strzal.
        # Blokowanie GUI na ~7s przy zaznaczeniu wiersza byloby gorsze niz
        # brak miniatury przez pierwsze kilka sekund pracy.
        prewarm_library_dwf_index(library_root)
        return None

    # Bez stat()/is_file() na trafieniu - to strzal po sieci (~7ms) przy kazdym
    # zaznaczeniu wiersza. Nieaktualny wpis (plik przeniesiony po zbudowaniu
    # indeksu) i tak konczy sie lagodnie: dwf_thumb zwroci brak podgladu.
    return index.get(drawing_no_n.upper())


def find_assembly_for_drawing(v_drive_root, project_name: str, drawing_no: str) -> Optional[Dict]:
    """
    Szuka na bieżąco (bez zapisu do bazy), do jakiego złożenia należy dany
    detal, przeszukując pliki *_OUT.xlsx w folderze projektu na V:\\.

    Returns dict:
        {
            'out_file': Path,          # plik OUT w którym znaleziono detal
            'row': {...},               # wpis z find_assembly_tree_rows dla tego detalu
            'all_rows': [...],          # wszystkie wiersze drzewka z tego pliku OUT (do budowy pełnego drzewa)
            'parent_drawing_no': str,   # bezpośredni rodzic (przedostatni segment ścieżki), None jeśli root
        }
    lub None, jeśli nie znaleziono (brak folderu V:\\, brak plików OUT, lub
    detal nie występuje w żadnym drzewku).
    """
    drawing_no_n = norm(drawing_no)
    if not drawing_no_n:
        return None

    project_folder = find_project_folder(Path(v_drive_root), project_name)
    if not project_folder:
        return None

    for out_path in find_out_files(project_folder):
        rows = find_assembly_tree_rows(out_path)
        if not rows:
            continue

        for row in rows:
            if row["nr_rysunku"].upper() == drawing_no_n.upper():
                sciezka = row["sciezka"]
                parent = sciezka[-2] if len(sciezka) >= 2 else None
                return {
                    "out_file": out_path,
                    "row": row,
                    "all_rows": rows,
                    "parent_drawing_no": parent,
                }

    return None


def excel_import_material_thickness(
    con,  # sqlite3.Connection
    project_id: int,
    excel_path: Path
) -> Tuple[int, int]:
    """
    Importuje Materiał (tekst) i grubość z LOGISTYKA_OUT.xlsx (arkusz ZBIORCZY).
    
    Zasady:
      - materiał: ustawiamy mat_effective_text tylko jeśli pusty w DB i wartość w Excelu niepusta
      - grubość: ustawiamy thickness_mm + thickness_src='CSV' tylko jeśli thickness_src != 'USER'
                 i da się wyciągnąć z materiału (grX,XXmm)
      - Dopasowanie:
          A) jeśli wiersz ma 'Nr rysunku' → po Nr rysunku (po norm)
          B) jeśli 'Nr rysunku' puste (ZNORMALIZOWANE) → po Nazwa (+ opcjonalnie Opis jeśli istnieje w DB)
    
    Args:
        con: Połączenie SQLite z bazą projektu
        project_id: ID projektu (w v10 może być zawsze z project_con, więc opcjonalne)
        excel_path: Ścieżka do LOGISTYKA_OUT.xlsx
    
    Returns:
        (upd_mat, upd_thk) - liczba zaktualizowanych rekordów
    """
    # Mapy z Excela
    by_dn = {}           # drawing_no_norm -> material_text
    by_name_desc = {}    # (name_norm, desc_norm) -> material_text (dla wierszy bez Nr rysunku)
    by_name = {}         # fallback: name_norm -> material_text
    
    # KROK 1: Buduj mapy z Excela
    for rec in iter_zbiorczy_data_rows(excel_path):
        dn = norm(rec.get("Nr rysunku"))
        name = norm(rec.get("Nazwa"))
        desc = norm(rec.get("Opis"))
        mt = norm(rec.get("Materiał"))
        
        if not mt:
            continue
        
        if dn:
            # Ma numer rysunku
            by_dn.setdefault(dn, mt)
        elif name:
            # Brak numeru (ZNORMALIZOWANE) - użyj nazwa+opis
            by_name_desc.setdefault((name, desc), mt)
            # Fallback: jeśli w Excelu brak opisu albo w DB opis będzie NULL
            by_name.setdefault(name, mt)
    
    if not by_dn and not by_name_desc and not by_name:
        # Brak danych do importu
        print("⚠️  Brak danych z materiałem w Excelu!")
        return (0, 0)
    
    print(f"📊 Mapy z Excela:")
    print(f"   by_dn (po nr rysunku): {len(by_dn)} wpisów")
    print(f"   by_name_desc (po nazwa+opis): {len(by_name_desc)} wpisów")
    print(f"   by_name (po nazwie): {len(by_name)} wpisów")
    
    upd_mat = 0
    upd_thk = 0
    matched = 0  # Ile dopasowań
    
    # KROK 2: Pobierz items z DB
    # V10: używamy work_* i src_* (COALESCE)
    rows = con.execute(
        """
        SELECT id,
               COALESCE(NULLIF(work_drawing_no, ''), src_drawing_no) AS drawing_no,
               COALESCE(NULLIF(work_name, ''), src_name) AS name,
               COALESCE(NULLIF(work_desc, ''), src_desc) AS descr,
               mat_effective_text,
               thickness_src
        FROM items
        """,
    ).fetchall()
    
    # KROK 3: Dopasuj i aktualizuj
    for item_id, drawing_no, name, descr, mat_effective, thickness_src in rows:
        key_dn = norm(drawing_no)
        key_name = norm(name)
        
        # Znajdź materiał z Excela
        mt = None
        if key_dn and key_dn in by_dn:
            # Dopasowanie po nr rysunku
            mt = by_dn[key_dn]
        elif (not key_dn) and key_name:
            # Dopasowanie po nazwa+opis (ZNORMALIZOWANE)
            key_desc = norm(descr)
            mt = by_name_desc.get((key_name, key_desc))
            if mt is None:
                # Fallback: po samej nazwie
                mt = by_name.get(key_name)
        
        if not mt:
            # Brak dopasowania
            continue
        
        matched += 1
        
        # AKTUALIZACJA MATERIAŁU
        # V10: Nadpisz mat_effective_text TYLKO jeśli puste
        if not mat_effective or mat_effective.strip() == "":
            con.execute(
                "UPDATE items SET mat_effective_text = ?, updated_at = datetime('now') WHERE id = ?",
                (mt, item_id)
            )
            upd_mat += 1
        
        # AKTUALIZACJA GRUBOŚCI
        # Wyciągnij grubość z materiału (regex "grX,XXmm")
        th = parse_thickness_mm(mt)
        if th is not None and str(thickness_src or "").upper() != "USER":
            con.execute(
                "UPDATE items SET thickness_mm = ?, thickness_src = 'CSV', updated_at = datetime('now') WHERE id = ?",
                (th, item_id)
            )
            upd_thk += 1
    
    print(f"📊 Podsumowanie importu:")
    print(f"   Items w DB: {len(rows)}")
    print(f"   Dopasowań: {matched}")
    print(f"   Zaktualizowano materiał: {upd_mat}")
    print(f"   Zaktualizowano grubość: {upd_thk}")
    
    return upd_mat, upd_thk
