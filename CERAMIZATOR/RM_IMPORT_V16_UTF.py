
def norm_key(s: object) -> str:
    """Klucz normalizowany do porównań (alias dla normalize_ws_for_match)."""
    try:
        return normalize_ws_for_match(s)
    except Exception:
        return str(s).strip() if s is not None else ""

def first_token(s: object) -> str:
    k = norm_key(s)
    return k.split(" ", 1)[0] if k else ""

def build_canonical_filename(nr: object, nazwa: object, ext: str) -> str:
    nr0 = first_token(nr)
    name0 = str(nazwa).strip() if nazwa is not None else ""
    if name0:
        return f"{nr0} {name0}{ext}"
    return f"{nr0}{ext}"


def display_missing_filename(nr: object, nazwa: object, ext: str, candidates=None) -> str:
    """
    Prefer canonical '<Nr> <Nazwa>.<ext>' if Nazwa is present.
    If Nazwa missing, but we have candidate paths (project/library), use the best candidate filename.
    """
    name = str(nazwa).strip() if nazwa is not None else ""
    if name:
        return build_canonical_filename(nr, name, ext)
    if candidates:
        try:
            # candidates may be list of Path-like
            return candidates[0].name
        except Exception:
            pass
    # fallback: just number
    return build_canonical_filename(nr, "", ext)


#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
LOGISTYKA_AUTOMAT v1_5 (CSV Edition + BŁĘDY + DRZEWKO + ZBIORCZY SUMA)

Nowe rzeczy:
- Arkusz BŁĘDY:
    * dodatkowy typ błędu:
      "Brak DWF dla numeru z BOM" – dla numerów RMPAK z BOM,
      dla których w projekcie nie ma pliku DWF
- Arkusze PEŁNA TABELA (BOM) + ELEMENTY MODUŁY (ZZ) + STANDARD + X + XX + ZNORMALIZOWANE:
    * nowa kolumna "Ilość całkowita"
    * Ilość (lokalna) jak w BOM
    * Ilość całkowita:
        - dla RMPAK = suma na projekt z DRZEWKA
        - dla znormalizowanych = Ilość lokalna z BOM
- Arkusz ZBIORCZY:
    * MODUŁY / STANDARD / X / XX – ilości całkowite z DRZEWKA
    * ELEMENTY ZNORMALIZOWANE – kopiowane z BOM (jak w arkuszu ELEMENTY ZNORMALIZOWANE)
"""

import sys
import re
import json
from pathlib import Path
from typing import Dict, List, Tuple, Any
import threading

# GUI
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Plik konfiguracyjny dla GUI (ostatnia ścieżka)
GUI_CONFIG_FILE = Path.home() / ".logistyka_automat_config.json"

import pandas as pd
import math
import subprocess
import tempfile
import hashlib
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class UserCancelled(Exception):
    """Wyjątek używany do przerwania przetwarzania przez użytkownika."""
    pass


# ---------------------------
# Numer RMPAK (Twoje założenie):
# PREFIX: 2–4 znaki alfanum (litery/cyfry), potem '-', potem ###.##, opcjonalny sufiks: X / XX / Z / ZZ
# Przykłady prefixów: NN, 11, 1N, N1, NNN, N1N, NNNN, 1NNN, 1111, NS, AB12
# ---------------------------


# ---------------------------
# Numer RMPAK (Twoje założenie):
# PREFIX: 2–4 znaki alfanum (litery/cyfry), potem '-', potem ###.##, opcjonalny sufiks: X / XX / Z / ZZ
# Przykłady prefixów: NN, 11, 1N, N1, NNN, N1N, NNNN, 1NNN, 1111, NS, AB12
# ---------------------------
RMPAK_RE = re.compile(r"^(?P<prefix>[A-Za-z0-9]{2,4})-(?P<num>\d{3}\.\d{2})(?P<suf>(XX|X|ZZ|Z|Y)?)$", re.IGNORECASE)

def is_rmpak_number(raw: object) -> bool:
    k = normalize_drawing_for_match(str(raw) if raw is not None else "")
    return bool(RMPAK_RE.match(k.split(" ", 1)[0]))

def split_rmpak(raw: object):
    k = normalize_drawing_for_match(str(raw) if raw is not None else "")
    k = k.split(" ", 1)[0]
    m = RMPAK_RE.match(k)
    if not m:
        return "", "", ""
    return (m.group("prefix").upper(), m.group("num"), (m.group("suf") or "").upper())

    k = norm_key(raw)
    m = RMPAK_RE.match(k)
    if not m:
        return "", "", ""
    return (m.group("prefix").upper(), m.group("num"), (m.group("suf") or "").upper())


import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --------------------------------
# KONFIGURACJA PODSTAWOWA
# --------------------------------

# OUTPUT_EXCEL_NAME - dynamicznie generowana na podstawie ROOT BOM

SHEET_FULL = "PEŁNA TABELA (BOM)"
SHEET_MODULES = "ELEMENTY MODUŁY (ZZ)"
SHEET_STANDARD = "ELEMENTY STANDARD"
SHEET_X = "ELEMENTY DO CIĘCIA (X)"
SHEET_XX = "ELEMENTY DO CIĘCIA GIĘCIA (XX)"
SHEET_NORM = "ELEMENTY ZNORMALIZOWANE"
SHEET_COMBINED = "ZBIORCZY"
SHEET_ERRORS = "BŁĘDY"

SHEET_TREE_ASCII = "DRZEWKO ASCII"
SHEET_TREE_TEXT = "DRZEWKO TEKST"

# >>> kolumny docelowe, z nową "Ilość całkowita" <<<
TARGET_COLUMNS = [
    "Poz.",
    "Nr rysunku",
    "Nazwa",
    "Opis",
    "Ilość",
    "Ilość całkowita",
    "Materiał",
    "Dostawca",
    "Pliki 3D",
    "Katalog",
]

# kolumna techniczna – nazwa źródła BOM (plik CSV)
SOURCE_COL = "__Źródło__"

# RMPAK core: 2–4 znaki alfanumeryczne + - + ###.##
RE_RMPAK_CORE = re.compile(r"^[A-Za-z0-9]{2,4}-\d{3}\.\d{2}")
# baza numeru RMPAK (bez sufiksów X/XX/Z/ZZ)
RE_RMPAK_BASE = re.compile(r"^([A-Za-z0-9]{2,4}-\d{3}\.\d{2})")
# katalog z numeru RMPAK (### z części po myślniku)
RE_RMPAK_KAT = re.compile(r"^[A-Za-z0-9]{2,4}-(\d{3})\.\d{2}")
# trzycyfrowy numer katalogu (np. 100, 200, 900)
RE_KATALOG = re.compile(r"\b(\d{3})\b")


# --------------------------------
# FUNKCJE POMOCNICZE
# --------------------------------

def log(msg: str) -> None:
    print(msg)
    # Jeśli GUI jest aktywne, wyślij też tam
    if hasattr(log, 'gui_callback') and log.gui_callback:
        try:
            log.gui_callback(msg)
        except:
            pass

# Callback dla GUI (zostanie ustawiony przez aplikację)
log.gui_callback = None



def normalize_ws_for_match(s: str) -> str:
    """
    Normalizacja klucza do dopasowań (TYLKO logika, NIGDY zapis do Excela).

    Cel: ujednolicić "ten sam numer" zapisany inaczej w CSV / nazwach plików:
    - trailing/leading whitespace (spacje, taby, CR/LF)
    - twarda spacja NBSP
    - niewidzialne znaki zero-width/BOM
    - różne typy myślników/dashy (zamiana na zwykły '-')
    - kolaps wielokrotnych spacji

    UWAGA: nie wolno używać tego do modyfikacji wartości, które zapisujesz do LOGISTYKA_OUT.xlsx.
    """
    if s is None:
        return ""

    s = str(s)

    # Unicode normalization (stabilizuje formy znaków)
    try:
        import unicodedata
        s = unicodedata.normalize("NFKC", s)
    except Exception:
        pass

    # NBSP -> zwykła spacja
    s = s.replace("\u00A0", " ")

    # Usuń niewidzialne znaki
    for ch in ("\u200B", "\u200C", "\u200D", "\uFEFF"):
        s = s.replace(ch, "")

    # Ujednolić różne myślniki na '-'
    for h in ("\u2010", "\u2011", "\u2012", "\u2013", "\u2014", "\u2212", "\u00AD"):
        s = s.replace(h, "-")

    # Tab/CR/LF -> spacja
    s = s.replace("\t", " ").replace("\r", " ").replace("\n", " ")

    # Kolaps whitespace + strip
    s = re.sub(r"\s+", " ", s).strip()

    # Usuń spacje wokół myślnika, jeśli się wkleiły (np. "2556 - 100.07XX")
    s = re.sub(r"\s*-\s*", "-", s)

    return s

def normalize_drawing_for_match(num: str) -> str:
    """Czyści Nr rysunku do porównań plikowych i błędów."""
    return normalize_ws_for_match(num)

def canonical_name_from_bom(nr_raw: str, nazwa_raw: object, ext: str) -> str:
    """Buduje kanoniczną nazwę pliku: '<NR> <NAZWA><ext>' (jeśli NAZWA pusta -> '<NR><ext>').
    Uwaga: NR jest brany jako pierwszy token (po normalizacji logicznej), bez zmian widocznych w Excelu.
    """
    nr_key = normalize_drawing_for_match(nr_raw)
    if not nr_key:
        nr_key = normalize_ws_for_match(nr_raw) if 'normalize_ws_for_match' in globals() else str(nr_raw).strip()
        nr_key = (nr_key.split(' ', 1)[0] if nr_key else '')
    name = '' if nazwa_raw is None else str(nazwa_raw).strip()
    if name:
        return f"{nr_key} {name}{ext}"
    return f"{nr_key}{ext}"



def rmpak_base(num: str) -> str:
    """Zwraca bazę numeru RMPAK (bez sufiksów X/XX/Z/ZZ). Dla niermpak zwraca ""."""
    if not isinstance(num, str):
        return ""
    n = normalize_ws_for_match(num)
    if not n:
        return ""
    m = RE_RMPAK_BASE.match(n)
    return m.group(1) if m else ""


def rmpak_suffix(num: str) -> str:
    """Zwraca sufiks (X/XX/Z/ZZ) po bazie RMPAK. Dla braku sufiksu -> ""."""
    n = normalize_ws_for_match(num)
    base = rmpak_base(n)
    if not num:
        return ""
    suf = n[len(base):].strip()
    # porządkujemy do znanych sufiksów
    suf_u = suf.upper()
    if suf_u in ("X", "XX", "Z", "ZZ"):
        return suf_u
    return suf_u


def katalog_from_rmpak_number(num: str) -> str:
    """Wyciąga trzycyfrowy katalog z numeru RMPAK (np. 2556-720.04ZZ -> '720')."""
    n = normalize_ws_for_match(num)
    m = RE_RMPAK_KAT.match(n)
    return m.group(1) if m else ""

def extract_drawing_number_from_filename(path: Path) -> str:
    """Numer rysunku = pierwszy token przed spacją w nazwie pliku (bez rozszerzenia)."""
    stem = path.stem
    parts = stem.split(" ", 1)
    if not parts:
        return stem
    return parts[0].strip()


def extract_number_from_source_name(source_name: str) -> str:
    """Z nazwy BOM (np. '2556-000.00ZZ Nalewarka 10N') wyciąga numer rysunku."""
    parts = str(source_name).split(" ", 1)
    if not parts:
        return str(source_name).strip()
    return parts[0].strip()


def extract_material_from_dxf_filename(path: Path) -> str:
    """
    Z nazwy DXF wyciąga materiał w formacie dokładnie jak w pliku DXF, np.:
      "304 gr1,5mm", "PA6 gr2mm", "Stal czarna gr10mm"
    Uwaga: NIE wolno rozcinać po przecinku w grubości (np. "1,5mm").

    Reguła:
    - jeśli w nazwie występuje separator ", " (przecinek + spacja), bierzemy fragment po OSTATNIM ", "
    - fallback: wyciągamy regexem końcówkę zawierającą "gr...mm"
    - jeśli nie znajdziemy -> pusty string
    """
    stem = path.stem

    # Najbezpieczniej rozdzielać po ", " (nie łamie "1,5mm")
    if ", " in stem:
        tail = stem.rsplit(", ", 1)[-1].strip()
        if re.search(r"\bgr\d+(?:,\d+)?mm\b", tail, flags=re.IGNORECASE):
            return tail

    # Fallback: znajdź fragment zakończony "gr...mm" (bez normalizacji formatu)
    m = re.search(r"(?i)\b(.+?\bgr\d+(?:,\d+)?mm)\b", stem)
    return m.group(1).strip() if m else ""



def extract_katalog_from_folder(folder_name: str) -> str:
    """
    Z nazwy katalogu wyciąga trzycyfrowy numer katalogu.
    Jeśli brak takiej sekwencji -> pusty string.
    """
    m = RE_KATALOG.search(folder_name)
    if m:
        return m.group(1)
    return ""


def katalog_for_path(path: Path) -> str:
    """Katalog z folderu danego pliku (np. 'Rysunki 900 2556' -> '900')."""
    return extract_katalog_from_folder(path.parent.name)


def format_file_list(paths: List[Path] | List[str] | None) -> str:
    """Zwraca wieloliniową listę plików (pełne ścieżki) do kolumny 'Lista plików'."""
    if not paths:
        return ""
    out: List[str] = []
    for p in paths:
        try:
            s = str(p)
        except Exception:
            s = ""
        s = s.strip()
        if s:
            out.append(s)
    return "\n".join(out)

def classify_type(num: str) -> str:
    """
    Klasyfikacja typu elementu na potrzeby DRZEWKA/ZBIORCZY.
    """
    if not isinstance(num, str):
        return ""
    n = num.strip()
    if not is_rmpak_number(n):
        return "ZNORMALIZOWANE"
    if n.endswith("ZZ"):
        return "MODUŁ (ZZ)"
    if n.endswith("XX"):
        return "CIĘCIE+GIĘCIE (XX)"
    if n.endswith("X") and not n.endswith("XX"):
        return "CIĘCIE (X)"
    if n.endswith("Z") and not n.endswith("ZZ"):
        return "ZŁOŻENIE (Z)"
    return "STANDARD"


def parse_qty(q: str) -> float:
    """Parsuje ilość typu '2', '2,5', '2.5'. Zwraca float lub None."""
    if q is None:
        return None
    q = str(q).strip()
    if not q:
        return None
    # Inventor/CSV czasem daje spacje w liczbach typu "7, 5" – usuń wszystkie białe znaki.
    q = re.sub(r"\s+", "", q)
    q = q.replace(",", ".")
    try:
        return float(q)
    except ValueError:
        return None


def format_qty(x: float) -> str:
    """Formatowanie ilości całkowitej z powrotem do stringa (z przecinkiem)."""
    if x is None:
        return ""
    if float(x).is_integer():
        return str(int(round(x)))
    s = f"{x}"
    return s.replace(".", ",")


# --------------------------------
# SKANOWANIE PROJEKTU (pomija OldVersions)
# --------------------------------

def scan_project_files(project_dir: Path, cancel_check=None) -> Dict[str, object]:
    """
    Rekurencyjnie skanuje katalog projektu i podkatalogi.
    Pomija wszystkie ścieżki, które w którymkolwiek komponencie mają nazwę 'OldVersions'.
    """
    csv_files: List[Path] = []
    idw_files: List[Path] = []
    idw_map: Dict[str, List[Path]] = {}
    dxf_map: Dict[str, List[Path]] = {}
    stl_map: Dict[str, List[Path]] = {}
    stp_map: Dict[str, List[Path]] = {}
    dwf_map: Dict[str, List[Path]] = {}

    exts_dxf = {".dxf"}
    exts_stl = {".stl"}
    exts_stp = {".stp", ".step"}
    exts_dwf = {".dwf"}
    exts_csv = {".csv"}
    exts_idw = {".idw"}

    for path in project_dir.rglob("*"):
        if cancel_check and cancel_check():
            raise UserCancelled("Skanowanie zostało przerwane przez użytkownika.")
        if not path.is_file():
            continue

        if _is_ignored_folder(path):
            continue

        ext = path.suffix.lower()

        if ext in exts_csv:
            csv_files.append(path)
            continue
        if ext in exts_idw:
            idw_files.append(path)
            num_raw = extract_drawing_number_from_filename(path)
            num = normalize_drawing_for_match(num_raw)
            idw_map.setdefault(num, []).append(path)
            continue

        num_raw = extract_drawing_number_from_filename(path)
        num = normalize_drawing_for_match(num_raw)

        if ext in exts_dxf:
            dxf_map.setdefault(num, []).append(path)
        elif ext in exts_stl:
            stl_map.setdefault(num, []).append(path)
        elif ext in exts_stp:
            stp_map.setdefault(num, []).append(path)
        elif ext in exts_dwf:
            dwf_map.setdefault(num, []).append(path)

    log(f"Znaleziono CSV: {len(csv_files)}")
    log(f"Znaleziono IDW: {len(idw_files)}")
    log(f"Znaleziono DXF: {sum(len(v) for v in dxf_map.values())}")
    log(f"Znaleziono STL: {sum(len(v) for v in stl_map.values())}")
    log(f"Znaleziono STP/STEP: {sum(len(v) for v in stp_map.values())}")
    log(f"Znaleziono DWF: {sum(len(v) for v in dwf_map.values())}")

    return {
        "csv_files": csv_files,
        "idw_files": idw_files,
        "idw_map": idw_map,
        "dxf_map": dxf_map,
        "stl_map": stl_map,
        "stp_map": stp_map,
        "dwf_map": dwf_map,
    }


# --------------------------------
# WCZYTANIE BOM Z CSV
# --------------------------------
# --------------------------------
# BIBLIOTEKA DWF – DOKŁADKA (fallback + raport + linki)
# --------------------------------

LIBRARY_DWF_ROOT = Path(r"C:\biblioteka")

# mapy tylko do raportowania/formatowania (nie wpływają na logikę BOM)
DWF_MAP_PROJECT_FOR_REPORT: Dict[str, List[Path]] = {}
DWF_MAP_LIBRARY_FOR_REPORT: Dict[str, List[Path]] = {}

def _is_in_oldversions(path: Path) -> bool:
    try:
        return any(p.lower() == "oldversions" for p in path.parts)
    except Exception:
        return False



def _is_ignored_folder(path: Path) -> bool:
    try:
        return any(p.lower() in ("oldversions", "templates") for p in path.parts)
    except Exception:
        return False

def scan_library_dwf(library_dir: Path) -> Dict[str, List[Path]]:
    """
    Skanuje bibliotekę (rekurencyjnie) TYLKO pod kątem plików .dwf.
    Zwraca mapę: {nr_rysunku_normalized -> [ścieżki_dwf]}
    Pomija ścieżki zawierające 'OldVersions' (tak jak projekt).
    """
    dwf_map_lib: Dict[str, List[Path]] = {}
    if not library_dir or not library_dir.exists():
        return dwf_map_lib

    for path in library_dir.rglob("*.dwf"):
        if not path.is_file():
            continue
        if _is_ignored_folder(path):
            continue
        num_raw = extract_drawing_number_from_filename(path)
        num = normalize_drawing_for_match(num_raw)
        if not num:
            continue
        dwf_map_lib.setdefault(num, []).append(path)

    return dwf_map_lib

def merge_dwf_maps(dwf_project: Dict[str, List[Path]],
                   dwf_library: Dict[str, List[Path]]) -> Dict[str, List[Path]]:
    """
    Łączy DWF z projektu i z biblioteki (fallback).
    Nie usuwa istniejących ścieżek, tylko dopina dodatkowe.
    """
    merged: Dict[str, List[Path]] = {}
    for k, paths in (dwf_project or {}).items():
        merged.setdefault(k, []).extend(paths)
    for k, paths in (dwf_library or {}).items():
        merged.setdefault(k, []).extend(paths)

    # unikalność ścieżek per klucz (bez zmiany kolejności)
    for k, paths in list(merged.items()):
        seen = set()
        uniq: List[Path] = []
        for p in paths:
            sp = str(p)
            if sp not in seen:
                seen.add(sp)
                uniq.append(p)
        merged[k] = uniq

    return merged

def _mtime_safe(p: Path) -> float:
    try:
        return p.stat().st_mtime
    except Exception:
        return 0.0

def pick_newest_path(paths: List[Path]) -> Path | None:
    """Wybiera najnowszą wersję (po dacie modyfikacji)."""
    if not paths:
        return None
    return max(paths, key=_mtime_safe)

def sort_paths_newest_first(paths: List[Path]) -> List[Path]:
    return sorted(paths or [], key=_mtime_safe, reverse=True)

def _sanitize_for_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s or "")
    s = s.strip("._-")
    return s or "link"

def ensure_explorer_select_lnk(target_file: Path) -> Path | None:
    """Tworzy skrót .lnk, który uruchamia: explorer.exe /select,"<target_file>".
    Zwraca ścieżkę do .lnk lub None gdy nie udało się utworzyć.
    """
    try:
        target_file = target_file.resolve()
    except Exception:
        target_file = Path(str(target_file))

    try:
        link_dir = Path(tempfile.gettempdir()) / "LOGISTYKA_AUTOMAT_LINKS"
        link_dir.mkdir(parents=True, exist_ok=True)

        stem = _sanitize_for_filename(target_file.stem)[:40]
        h = hashlib.md5(str(target_file).encode("utf-8", "ignore")).hexdigest()[:10]
        lnk_path = link_dir / f"{stem}_{h}.lnk"

        if lnk_path.exists():
            return lnk_path

        # PowerShell COM: WScript.Shell CreateShortcut
        lnk_str = str(lnk_path)
        file_str = str(target_file)
        workdir = str(target_file.parent)

        def ps_escape(s: str) -> str:
            # escape for single-quoted PS string
            return s.replace("'", "''")

        ps = (
            "$w=New-Object -ComObject WScript.Shell;"
            f"$s=$w.CreateShortcut('{ps_escape(lnk_str)}');"
            "$s.TargetPath='explorer.exe';"
            f"$s.Arguments='/select,\"{ps_escape(file_str)}\"';"
            f"$s.WorkingDirectory='{ps_escape(workdir)}';"
            "$s.WindowStyle=1;"
            "$s.Save();"
        )

        # Na Windows: ukryj okno powershell'a
        startupinfo = None
        if sys.platform == "win32":
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE

        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            capture_output=True,
            text=True,
            check=False,
            startupinfo=startupinfo,
        )

        if lnk_path.exists():
            return lnk_path
    except Exception:
        return None

    return None

def hyperlink_to_select_file(target_file: Path) -> str | None:
    """Zwraca hyperlink (URI) do .lnk, który otworzy folder z zaznaczonym plikiem.
    Jeśli nie uda się utworzyć .lnk, zwraca URI do folderu.
    """
    newest = target_file
    lnk = ensure_explorer_select_lnk(newest)
    try:
        if lnk and lnk.exists():
            return lnk.resolve().as_uri()
    except Exception:
        pass
    try:
        return newest.parent.resolve().as_uri()
    except Exception:
        return None

def apply_library_links_to_ws(ws) -> None:
    """
    Jeśli w danym arkuszu kolumna 'Nr rysunku' ma DWF tylko w bibliotece (brak w projekcie),
    to:
      - ustawia hyperlink w komórce numeru do NAJNOWSZEGO pliku DWF z biblioteki,
      - zmienia kolor czcionki i podkreślenie (żeby było widać, że to link z biblioteki).

    Ważne:
      - Nie dodaje kolumn i nie zmienia wartości komórek (Excel pozostaje "święty").
      - Działa też dla arkuszy z wieloma sekcjami (np. ZBIORCZY),
        bo wykrywa pozycję kolumny 'Nr rysunku' dynamicznie na podstawie nagłówka.
    """
    if not DWF_MAP_LIBRARY_FOR_REPORT:
        return

    proj_map = DWF_MAP_PROJECT_FOR_REPORT or {}
    lib_map = DWF_MAP_LIBRARY_FOR_REPORT or {}

    current_col_nr = None

    # Helper: czy to jest wiersz nagłówków kolumn (zawiera 'Nr rysunku')
    def _find_nr_col_in_row(r: int) -> int | None:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip() == "Nr rysunku":
                return c
        return None

    for r in range(1, ws.max_row + 1):
        # aktualizuj kolumnę Nr rysunku jeśli trafimy na nagłówek
        nr_col = _find_nr_col_in_row(r)
        if nr_col:
            current_col_nr = nr_col
            continue  # wiersz nagłówka pomijamy

        if not current_col_nr:
            continue

        cell = ws.cell(row=r, column=current_col_nr)
        val = cell.value
        if not val:
            continue

        s = str(val).strip()
        if not s or s == "Nr rysunku":
            continue

        key = normalize_drawing_for_match(s)
        if not key:
            continue

        in_proj = key in proj_map
        in_lib = key in lib_map and bool(lib_map.get(key))
        if in_proj or not in_lib:
            continue

        newest = pick_newest_path(lib_map.get(key, []))
        if not newest:
            continue

        try:
            hl = hyperlink_to_select_file(newest)
            if hl:
                cell.hyperlink = hl
        except Exception:
            pass

        # Zielony = "z biblioteki"
        try:
            f = copy(cell.font) if cell.font else Font()
            f.color = "0000FF"
            f.underline = "single"
            cell.font = f
        except Exception:
            pass


def apply_dwf_hyperlinks_to_zbiorczy(ws) -> None:
    """
    ZBIORCZY: w kolumnie 'Nr rysunku' dodaj hyperlink do DWF:
      - jeśli DWF jest w PROJEKCIE -> link do najnowszego DWF w projekcie,
      - jeśli w projekcie brak, ale jest w BIBLIOTECE -> link do najnowszego DWF w bibliotece + kolor (biblioteka).

    To jest niezależne od 'apply_library_links_to_ws' (tam jest tylko biblioteka).
    """
    proj_map = DWF_MAP_PROJECT_FOR_REPORT or {}
    lib_map = DWF_MAP_LIBRARY_FOR_REPORT or {}

    current_col_nr = None

    def _find_nr_col_in_row(r: int) -> int | None:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip() == "Nr rysunku":
                return c
        return None

    for r in range(1, ws.max_row + 1):
        nr_col = _find_nr_col_in_row(r)
        if nr_col:
            current_col_nr = nr_col
            continue

        if not current_col_nr:
            continue

        cell = ws.cell(row=r, column=current_col_nr)
        val = cell.value
        if not val:
            continue

        s = str(val).strip()
        if not s or s == "Nr rysunku":
            continue

        key = normalize_drawing_for_match(s)
        if not key:
            continue

        target = None
        from_library = False

        if key in proj_map and proj_map.get(key):
            target = pick_newest_path(proj_map.get(key, []))
            from_library = False
        elif key in lib_map and lib_map.get(key):
            target = pick_newest_path(lib_map.get(key, []))
            from_library = True

        if not target:
            continue

        try:
            hl = hyperlink_to_select_file(target)
            if hl:
                cell.hyperlink = hl
        except Exception:
            pass

        try:
            f = copy(cell.font) if cell.font else Font()
            f.underline = "single"
            if from_library:
                # niebieski = link z biblioteki (łatwo odróżnić)
                f.color = "0000FF"
            else:
                # standardowy link (Excel zwykle sam to koloruje, ale zostawiamy podkreślenie)
                if not getattr(f, "color", None):
                    pass
            cell.font = f
        except Exception:
            pass


def read_bom_csv(csv_path: Path) -> pd.DataFrame:
    """
    Czyta BOM z CSV z Inventora:

        Poz.;Nr rysunku;Nazwa;Opis;Ilość;Materiał;Dostawca
    """
    log(f"  Czytam CSV: {csv_path}")
    encodings_to_try = [
        "utf-8-sig",
        "utf-8",
        "cp1250",
        "iso-8859-2",
    ]

    df = None
    last_err = None
    for enc in encodings_to_try:
        try:
            df = pd.read_csv(
                csv_path,
                sep=";",
                dtype=str,
                encoding=enc,
                engine="python",
                quotechar='"',
            )
            log(f"    Użyto kodowania: {enc}")
            break
        except UnicodeDecodeError as e:
            last_err = e
            continue

    if df is None:
        raise last_err
    df = df.fillna("")

    expected_cols = ["Poz.", "Nr rysunku", "Nazwa", "Opis", "Ilość", "Materiał", "Dostawca"]
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        raise ValueError(f"CSV {csv_path} nie ma oczekiwanych kolumn: {missing}")

    df = df[expected_cols]
    return df


# --------------------------------
# NADPISYWANIE MATERIAŁU Z DXF
# --------------------------------

def update_material_from_dxf(df_bom: pd.DataFrame,
                             dxf_map: Dict[str, List[Path]],
                             dwf_map_project: Dict[str, List[Path]] | None = None,
                             dwf_map_library: Dict[str, List[Path]] | None = None) -> pd.DataFrame:
    """
    Nadpisuje kolumnę "Materiał" z nazw plików DXF (format: "... grX,XXmm").

    Zasada (minimalnie inwazyjna):
      1) Jeżeli numer ma DWF TYLKO w bibliotece (brak w projekcie, jest w bibliotece),
         to MATERIAŁ bierzemy WYŁĄCZNIE z DXF z tego samego katalogu biblioteki co DWF.
         (DXF z projektu ignorujemy w tym przypadku.)
      2) W pozostałych przypadkach: DXF z projektu > CSV (dotychczasowa logika).

    Jeżeli w katalogu biblioteki (dla bibliotecznego DWF) nie ma DXF – nie nadpisujemy materiału z BOM/CSV.
    """
    if "Nr rysunku" not in df_bom.columns or "Materiał" not in df_bom.columns:
        return df_bom

    dwf_map_project = dwf_map_project or {}
    dwf_map_library = dwf_map_library or {}

    materials: List[str] = df_bom["Materiał"].astype(str).tolist()
    numbers: List[str] = df_bom["Nr rysunku"].astype(str).tolist()

    def _is_library_dwf(key: str) -> bool:
        has_proj = bool(key in dwf_map_project and dwf_map_project.get(key))
        has_lib = bool(key in dwf_map_library and dwf_map_library.get(key))
        return (not has_proj) and has_lib

    def _pick_library_dxf_for_key(key: str) -> Path | None:
        """
        Szuka DXF tylko w katalogu biblioteki odpowiadającego DWF (ten sam folder co znaleziony DWF).
        Dobór pliku: najnowszy DXF w tym folderze, którego numer po normalizacji == key.
        """
        try:
            dwf_path = pick_newest_path(dwf_map_library.get(key, []))
        except Exception:
            dwf_path = None
        if not dwf_path:
            return None

        folder = dwf_path.parent
        if not folder or not folder.exists():
            return None

        candidates: List[Path] = []
        try:
            for dxf_path in folder.glob("*.dxf"):
                num_raw = extract_drawing_number_from_filename(dxf_path)
                k = normalize_drawing_for_match(num_raw)
                if k == key:
                    candidates.append(dxf_path)
        except Exception:
            candidates = []

        if not candidates:
            return None

        try:
            return pick_newest_path(candidates)
        except Exception:
            return candidates[0]

    for i, num in enumerate(numbers):
        key = normalize_drawing_for_match(num)
        if not key:
            continue

        # 1) DWF tylko w bibliotece -> materiał tylko z DXF w tym samym katalogu biblioteki
        if _is_library_dwf(key):
            dxf_lib = _pick_library_dxf_for_key(key)
            if dxf_lib:
                mat = extract_material_from_dxf_filename(dxf_lib)
                if mat:
                    materials[i] = mat
            # jeśli brak DXF w bibliotece -> nie nadpisujemy (zostaje z CSV/BOM)
            continue

        # 2) Standard: DXF z projektu > CSV/BOM
        if key not in dxf_map:
            continue

        dxf_list = dxf_map.get(key, [])
        if not dxf_list:
            continue

        dxf_path = dxf_list[0]
        mat = extract_material_from_dxf_filename(dxf_path)
        if mat:
            materials[i] = mat

    df_bom = df_bom.copy()
    df_bom["Materiał"] = materials
    return df_bom

    materials: List[str] = df_bom["Materiał"].astype(str).tolist()
    numbers: List[str] = df_bom["Nr rysunku"].astype(str).tolist()

    for i, num in enumerate(numbers):
        key = normalize_drawing_for_match(num)
        if not key:
            continue
        if key not in dxf_map:
            continue

        dxf_list = dxf_map[key]
        if not dxf_list:
            continue

        dxf_path = dxf_list[0]
        mat = extract_material_from_dxf_filename(dxf_path)
        if mat:
            materials[i] = mat

    df_bom = df_bom.copy()
    df_bom["Materiał"] = materials
    return df_bom


# --------------------------------
# PLIKI 3D (STL / STP)
# --------------------------------

def assign_3d_files(df_bom: pd.DataFrame,
                    stl_map: Dict[str, List[Path]],
                    stp_map: Dict[str, List[Path]]) -> pd.DataFrame:
    """
    Uzupełnia kolumnę "Pliki 3D":
        STL       jeśli jest .stl
        STP       jeśli jest .stp/.step
        STP, STL  jeśli oba
        ""        jeśli brak modelu 3D
    """
    numbers: List[str] = df_bom["Nr rysunku"].astype(str).tolist()
    pliki_3d: List[str] = []

    for num in numbers:
        key = normalize_drawing_for_match(num)
        has_stl = bool(key and key in stl_map and stl_map[key])
        has_stp = bool(key and key in stp_map and stp_map[key])

        if has_stl and has_stp:
            pliki_3d.append("STP, STL")
        elif has_stp:
            pliki_3d.append("STP")
        elif has_stl:
            pliki_3d.append("STL")
        else:
            pliki_3d.append("")

    df_bom = df_bom.copy()
    df_bom["Pliki 3D"] = pliki_3d
    return df_bom


# --------------------------------
# KATALOG (dla każdego rysunku)
# --------------------------------

def assign_katalog(df_bom: pd.DataFrame,
                   csv_path: Path,
                   dxf_map: Dict[str, List[Path]],
                   stl_map: Dict[str, List[Path]],
                   stp_map: Dict[str, List[Path]],
                   dwf_map: Dict[str, List[Path]]) -> pd.DataFrame:
    """
    Dla KAŻDEGO wiersza BOM:
      1) próbujemy znaleźć jego DXF/STP/STL/DWF,
         jeśli plik jest w podkatalogu z numerem (np. 400/600/800),
         bierzemy ten numer,
      2) jeśli nie ma żadnego pliku albo brak numeru -> używamy numeru z katalogu BOM.
    """
    base_folder_name = csv_path.parent.name
    base_katalog = extract_katalog_from_folder(base_folder_name)

    numbers: List[str] = df_bom["Nr rysunku"].astype(str).tolist()
    katalogs: List[str] = []

    for num in numbers:
        key = normalize_drawing_for_match(num)
        # usuń "nan" i puste po normalizacji
        if key.lower() == "nan":
            key = ""
        kat = ""

        if key:
            candidate_paths: List[Path] = []
            if key in dxf_map:
                candidate_paths.extend(dxf_map[key])
            if key in stp_map:
                candidate_paths.extend(stp_map[key])
            if key in stl_map:
                candidate_paths.extend(stl_map[key])
            if key in dwf_map:
                candidate_paths.extend(dwf_map[key])

            for p in candidate_paths:
                kat_candidate = katalog_for_path(p)
                if kat_candidate:
                    kat = kat_candidate
                    break

        if not kat:
            kat = base_katalog

        # fallback 2: jeśli dalej pusto, a to numer RMPAK – wyciągnij katalog z numeru
        if not kat:
            kat = katalog_from_rmpak_number(key) if is_rmpak_number(key) else ""

        katalogs.append(kat)

    df_bom = df_bom.copy()
    df_bom["Katalog"] = katalogs
    return df_bom


# --------------------------------
# KLASYFIKACJA NA SEKCJE (BOM)
# --------------------------------

def split_sections(df_full: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Dzieli BOM na:
        FULL      - pełna tabela
        MODULES   - RMPAK z sufiksem ZZ
        STANDARD  - RMPAK bez X/XX/ZZ
        X         - RMPAK z X na końcu (bez XX)
        XX        - RMPAK z XX na końcu
        NORM      - brak numeru RMPAK
    """
    df_full = df_full.copy()

    if "Nr rysunku" not in df_full.columns:
        empty = df_full.iloc[0:0].copy()
        return {
            "FULL": df_full,
            "MODULES": empty,
            "STANDARD": empty,
            "X": empty,
            "XX": empty,
            "NORM": df_full.copy(),
        }

    nr_series = df_full["Nr rysunku"].astype(str).fillna("").map(normalize_ws_for_match)

    mask_rmpak = nr_series.apply(is_rmpak_number)
    mask_xx = mask_rmpak & nr_series.str.endswith("XX")
    mask_x = mask_rmpak & nr_series.str.endswith("X") & (~mask_xx)
    mask_modules = mask_rmpak & nr_series.str.endswith("ZZ")
    # STANDARD = wszystkie RMPAK, które NIE są X/XX/ZZ
    mask_standard = mask_rmpak & (~mask_x) & (~mask_xx) & (~mask_modules)
    mask_norm = ~mask_rmpak

    df_modules = df_full[mask_modules].copy()
    df_xx = df_full[mask_xx].copy()
    df_x = df_full[mask_x].copy()
    df_standard = df_full[mask_standard].copy()
    df_norm = df_full[mask_norm].copy()

    return {
        "FULL": df_full,
        "MODULES": df_modules,
        "STANDARD": df_standard,
        "X": df_x,
        "XX": df_xx,
        "NORM": df_norm,
    }


# --------------------------------
# SZEROKOŚCI KOLUMN (10 kolumn)
# --------------------------------

def _set_column_widths(ws) -> None:
    col_widths = {
        1: 6,   # Poz.
        2: 18,  # Nr rysunku
        3: 28,  # Nazwa
        4: 45,  # Opis
        5: 7,   # Ilość
        6: 12,  # Ilość całkowita
        7: 18,  # Materiał
        8: 18,  # Dostawca
        9: 12,  # Pliki 3D
        10: 12, # Katalog
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


# --------------------------------
# BŁĘDY – ANALIZA PLIKÓW
# --------------------------------

def find_file_errors(files: Dict[str, object], active_nums: set[str] | None = None, name_by_num: Dict[str, str] | None = None) -> List[List[str]]:
    """
    BŁĘDY plikowe (dla elementów aktywnych w DRZEWKU, jeśli active_nums != None):
      - IDW bez DWF
      - DWF-X / DWF-XX bez DXF
      - DXF bez DWF
      - STP/STEP bez DWF
      - STL bez DWF
      - Brak STP/STEP gdy istnieje IDW, DWF lub DXF-XX
      - Duplikaty plików dla tego samego numeru (DXF/DWF/IDW/STP/STL)
    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []

    idw_files: List[Path] = files.get("idw_files", [])
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    dwf_map_library: Dict[str, List[Path]] = files.get("dwf_map_library", {})
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})


    if name_by_num is None:
        name_by_num = {}

    def is_active(num: str) -> bool:
        if active_nums is None:
            return True
        return normalize_drawing_for_match(num) in active_nums

    def has_dwf(num: str) -> bool:
        n = normalize_drawing_for_match(num)
        # DWF w projekcie LUB w bibliotece -> nie zgłaszaj "brak DWF"
        return n in dwf_map or (n in dwf_map_library and bool(dwf_map_library.get(n)))

    # IDW bez DWF
    for path in idw_files:
        num_raw = extract_drawing_number_from_filename(path)
        num = normalize_drawing_for_match(num_raw)
        if not num:
            continue
        if not is_rmpak_number(num):
            continue
        if not is_active(num):
            continue
        if not has_dwf(num):
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku IDW
            nazwa = name_by_num.get(num, "")
            if not nazwa:
                # Wyciągnij nazwę z pliku IDW
                stem = path.stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            missing = display_missing_filename(num, nazwa, ".dwf")
            full_name = f"{num} {nazwa}" if nazwa else num
            errors.append([
                "PLIKI – brak pliku DWF dla rysunku IDW",
                missing,
                f"Dla elementu '{full_name}' istnieje rysunek Inventor (plik IDW: {path.name}), ale brakuje pliku DWF.",
                format_file_list([path]),
                katalog_for_path(path),
            ])

    # DWF-X / DWF-XX bez DXF
    for num, dwfs in dwf_map.items():
        if not is_active(num):
            continue
        for path in dwfs:
            name = path.name
            kat = katalog_for_path(path)
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku DWF
            nazwa = name_by_num.get(num, "")
            if not nazwa:
                # Wyciągnij nazwę z pliku DWF
                stem = path.stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            full_name = f"{num} {nazwa}" if nazwa else num
            # FIX: sprawdzaj sufiks NUMERU, nie całej nazwy pliku (unikaj fałszywych wykryć jak "PAX.dwf")
            sfx = rmpak_suffix(num)
            if sfx == "XX" and normalize_drawing_for_match(num) not in dxf_map:
                errors.append([
                    "PLIKI – brak pliku DXF dla detalu XX (cięcie + gięcie)",
                    display_missing_filename(num, nazwa, ".dxf"),
                    f"Element '{full_name}' ma sufiks XX, co oznacza detal wymagający cięcia laserowego i gięcia blachy. Istnieje plik DWF ({path.name}), ale brakuje niezbędnego pliku DXF z rozwiniętym wykrojem do programowania maszyny CNC.",
                    format_file_list([path]),
                    kat,
                ])
            elif sfx == "X" and normalize_drawing_for_match(num) not in dxf_map:
                errors.append([
                    "PLIKI – brak pliku DXF dla detalu X (cięcie)",
                    display_missing_filename(num, nazwa, ".dxf"),
                    f"Element '{full_name}' ma sufiks X, co oznacza detal do cięcia laserowego. Istnieje plik DWF ({path.name}), ale brakuje niezbędnego pliku DXF z wykrojem do programowania maszyny CNC.",
                    format_file_list([path]),
                    kat,
                ])

    # DXF bez DWF
    for num, dxfs in dxf_map.items():
        if not is_active(num):
            continue
        if not has_dwf(num):
            # jeden wpis na numer (czytelniej)
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku DXF
            nazwa = name_by_num.get(num, "")
            if not nazwa and dxfs:
                # Wyciągnij nazwę z pliku DXF (bez materiału po przecinku)
                stem = dxfs[0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    name_part = stem[first_space_idx + 1:].strip()
                    # Usuń materiał jeśli jest (po przecinku)
                    if ', ' in name_part:
                        nazwa = name_part.split(', ', 1)[0].strip()
                    else:
                        nazwa = name_part
            full_name = f"{num} {nazwa}" if nazwa else num
            errors.append([
                "PLIKI – brak pliku DWF przy istniejącym DXF",
                display_missing_filename(num, nazwa, ".dwf", candidates=dxfs),
                f"Dla elementu '{full_name}' istnieje plik DXF ({len(dxfs)} plik(i/ów)), ale brakuje pliku DWF.",
                format_file_list(dxfs),
                katalog_for_path(dxfs[0]) if dxfs else "",
            ])

    # STP/STEP bez DWF (tylko dla numerów RMPAK)
    for num, stps in stp_map.items():
        if not is_active(num):
            continue
        # Pomiń modele 3D nie pasujące do schematu RMPAK (biblioteki, komponenty zewnętrzne)
        if not is_rmpak_number(num):
            continue
        if not has_dwf(num):
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku STP
            nazwa = name_by_num.get(num, "")
            if not nazwa and stps:
                stem = stps[0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            full_name = f"{num} {nazwa}" if nazwa else num
            errors.append([
                "PLIKI – brak pliku DWF przy istniejącym STP/STEP",
                display_missing_filename(num, nazwa, ".dwf", candidates=stps),
                f"Dla elementu '{full_name}' istnieje model 3D w formacie STP/STEP ({len(stps)} plik(i/ów)), ale brakuje pliku DWF z rysunkiem 2D.",
                format_file_list(stps),
                katalog_for_path(stps[0]) if stps else "",
            ])

    # STL bez DWF (tylko dla numerów RMPAK)
    for num, stls in stl_map.items():
        if not is_active(num):
            continue
        # Pomiń modele 3D nie pasujące do schematu RMPAK (biblioteki, komponenty zewnętrzne)
        if not is_rmpak_number(num):
            continue
        if not has_dwf(num):
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku STL
            nazwa = name_by_num.get(num, "")
            if not nazwa and stls:
                stem = stls[0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            full_name = f"{num} {nazwa}" if nazwa else num
            errors.append([
                "PLIKI – brak pliku DWF przy istniejącym STL",
                display_missing_filename(num, nazwa, ".dwf", candidates=stls),
                f"Dla elementu '{full_name}' istnieje plik STL ({len(stls)} plik(i/ów)) używany do druku 3D lub wizualizacji, ale brakuje pliku DWF z rysunkiem technicznym.",
                format_file_list(stls),
                katalog_for_path(stls[0]) if stls else "",
            ])

    # STP/STEP brakuje gdy istnieje IDW, DWF lub DXF z sufiksem XX
    # Sprawdzamy TYLKO dla numerów z sufiksem XX
    nums_without_stp_xx = set()
    
    # Numery z IDW-XX
    for path in idw_files:
        num = normalize_drawing_for_match(extract_drawing_number_from_filename(path))
        if num and is_active(num) and num.endswith("XX") and num not in stp_map:
            nums_without_stp_xx.add(num)
    
    # Numery z DWF-XX
    for num in dwf_map.keys():
        if is_active(num) and num.endswith("XX") and num not in stp_map:
            nums_without_stp_xx.add(num)
    
    # Numery z DXF-XX
    for num in dxf_map.keys():
        if is_active(num) and num.endswith("XX") and num not in stp_map:
            nums_without_stp_xx.add(num)
    
    # Zgłoś błąd dla każdego numeru XX bez STP
    for num in nums_without_stp_xx:
        # Pobierz nazwę z name_by_num, jeśli nie ma - spróbuj z plików
        nazwa = name_by_num.get(num, "")
        if not nazwa:
            # Spróbuj z DWF
            if num in dwf_map and dwf_map[num]:
                stem = dwf_map[num][0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            # Jeśli nie ma w DWF, spróbuj IDW
            if not nazwa:
                for path in idw_files:
                    if normalize_drawing_for_match(extract_drawing_number_from_filename(path)) == num:
                        stem = path.stem
                        first_space_idx = stem.find(' ')
                        if first_space_idx > 0:
                            nazwa = stem[first_space_idx + 1:].strip()
                            break
        full_name = f"{num} {nazwa}" if nazwa else num
        
        # Sprawdź jakie pliki istnieją
        has_idw_xx = any(normalize_drawing_for_match(extract_drawing_number_from_filename(p)) == num for p in idw_files)
        has_dwf_xx = num in dwf_map
        has_dxf_xx = num in dxf_map
        
        source_info = []
        if has_idw_xx:
            source_info.append("IDW-XX")
        if has_dwf_xx:
            source_info.append("DWF-XX")
        if has_dxf_xx:
            source_info.append("DXF-XX")
        
        source_desc = " i ".join(source_info) if source_info else "rysunki XX"
        
        errors.append([
            "PLIKI – brak pliku STP/STEP",
            display_missing_filename(num, nazwa, ".stp"),
            f"Dla elementu '{full_name}' z sufiksem XX istnieje {source_desc}, ale brakuje modelu 3D w formacie STP/STEP.",
            format_file_list([p for p in idw_files if normalize_drawing_for_match(extract_drawing_number_from_filename(p)) == num] + 
                           dxf_map.get(num, []) + dwf_map.get(num, [])),
            katalog_from_rmpak_number(num),
        ])

    # Duplikaty (ten sam numer -> >1 plik danego typu)
    def _dup(map_obj: Dict[str, List[Path]], ext_label: str, only_rmpak: bool = False) -> None:
        for num, paths in map_obj.items():
            if not paths or len(paths) <= 1:
                continue
            # Dla plików 3D (STP/STL) pomijamy niepasujące do RMPAK (kompon. biblioteczne)
            if only_rmpak and not is_rmpak_number(num):
                continue
            # Duplikaty zgłaszamy ZAWSZE – to błąd porządkowy niezależny od DRZEWKA/active_nums.
            # Pobierz nazwę z name_by_num, jeśli nie ma - wyciągnij z nazwy pliku
            nazwa = name_by_num.get(num, "") if name_by_num else ""
            if not nazwa and paths:
                stem = paths[0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    name_part = stem[first_space_idx + 1:].strip()
                    # Dla DXF usuń materiał jeśli jest (po przecinku)
                    if 'DXF' in ext_label and ', ' in name_part:
                        nazwa = name_part.split(', ', 1)[0].strip()
                    else:
                        nazwa = name_part
            full_name = f"{num} {nazwa}" if nazwa else num
            errors.append([
                f"PLIKI – duplikaty plików {ext_label}",
                (paths[0].name if paths else normalize_drawing_for_match(num)),
                f"Dla elementu '{full_name}' wykryto {len(paths)} duplikatów plików {ext_label}.\n\nMogą to być:\n1) Kopie zapasowe w różnych lokalizacjach\n2) Różne wersje tego samego elementu\n3) Błędnie nazwane pliki innych elementów\n\nZweryfikuj wszystkie pliki i pozostaw tylko jedną aktualną wersję.",
                format_file_list(paths),
                katalog_for_path(paths[0]),
            ])

    _dup(dxf_map, "DXF")
    _dup(dwf_map, "DWF")
    # IDW nie ma mapy w scan_project_files? jest idw_map, ale tu mamy listę. Zrobimy duplikaty na podstawie stem.
    idw_by_num: Dict[str, List[Path]] = {}
    for p in idw_files:
        n = normalize_drawing_for_match(extract_drawing_number_from_filename(p))
        if not n:
            continue
        idw_by_num.setdefault(n, []).append(p)
    _dup(idw_by_num, "IDW")
    _dup(stp_map, "STP/STEP", only_rmpak=True)  # Tylko RMPAK (pomijamy biblioteki)
    _dup(stl_map, "STL", only_rmpak=True)       # Tylko RMPAK (pomijamy biblioteki)

    return errors


def find_filename_mismatch_errors(files: Dict[str, object], name_by_num: Dict[str, str] | None = None, df_full = None) -> List[List[str]]:
    """
    Sprawdza czy nazwy plików zgadzają się z nazwami z BOM (CSV).
    Każdy plik powinien mieć nazwę: '<Numer> <Nazwa z BOM>[, dodatkowe info].ext'
    
    Przykład prawidłowy:
        BOM: '2608-200.03XX Wons suwaka'
        Pliki: '2608-200.03XX Wons suwaka.idw', '2608-200.03XX Wons suwaka, 304 gr2mm.dxf'
    
    Przykład BŁĘDNY:
        BOM: '2608-200.03XX Wons suwaka'
        Plik: '2608-200.03XX B.stp' ❌ (nazwa 'B' zamiast 'Wons suwaka')
    
    Zwraca rekordy: [Typ błędu, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []
    
    if not name_by_num:
        return errors
    
    idw_files: List[Path] = files.get("idw_files", [])
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})
    
    def extract_name_from_file(file_path: Path, num: str) -> str:
        """Wyciąga nazwę z pliku (część po numerze rysunku).
        
        UWAGA: Przecinek jest ucinany TYLKO dla DXF (materiał/grubość),
        dla innych typów plików przecinek jest częścią nazwy (np. TC 50,5).
        """
        stem = file_path.stem
        # Usuń numer rysunku z początku
        if stem.startswith(num):
            rest = stem[len(num):].strip()
            # Usuń przecinek i wszystko po nim TYLKO dla DXF (materiał, grubość)
            # Dla DWF/PDF/STP/STL/IDW przecinek jest częścią nazwy!
            if "," in rest and file_path.suffix.lower() == ".dxf":
                rest = rest.split(",", 1)[0].strip()
            return rest
        return stem
    
    # Zbierz wszystkie pliki per numer
    all_files_by_num: Dict[str, List[Tuple[Path, str]]] = {}  # num -> [(path, typ)]
    
    for path in idw_files:
        num = normalize_drawing_for_match(extract_drawing_number_from_filename(path))
        if num:
            all_files_by_num.setdefault(num, []).append((path, "IDW"))
    
    for num, paths in dxf_map.items():
        for path in paths:
            all_files_by_num.setdefault(num, []).append((path, "DXF"))
    
    for num, paths in dwf_map.items():
        for path in paths:
            all_files_by_num.setdefault(num, []).append((path, "DWF"))
    
    for num, paths in stp_map.items():
        for path in paths:
            all_files_by_num.setdefault(num, []).append((path, "STP"))
    
    for num, paths in stl_map.items():
        for path in paths:
            all_files_by_num.setdefault(num, []).append((path, "STL"))
    
    # Sprawdź każdy numer
    for num, file_list in all_files_by_num.items():
        expected_name = name_by_num.get(num, "")
        if not expected_name:
            continue  # Brak nazwy w BOM - nie możemy walidować
        
        mismatched_files: List[Tuple[Path, str, str]] = []  # (path, typ, znaleziona_nazwa)
        
        for path, typ in file_list:
            actual_name = extract_name_from_file(path, num)
            
            # Porównaj nazwy (case-insensitive, ignoruj różnice w spacjach)
            expected_normalized = normalize_ws_for_match(expected_name)
            actual_normalized = normalize_ws_for_match(actual_name)
            
            if expected_normalized != actual_normalized:
                mismatched_files.append((path, typ, actual_name))
        
        # Zgłoś błąd jeśli znaleziono niezgodności
        if mismatched_files:
            full_name = f"{num} {expected_name}" if expected_name else num
            types_str = ", ".join(sorted(set(t for _, t, _ in mismatched_files)))
            
            # Znajdź BOM z którego pochodzi ten numer
            source_bom = ""
            try:
                if df_full is not None and SOURCE_COL in df_full.columns:
                    import pandas as pd
                    nr_normalized = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
                    matching_rows = df_full[nr_normalized == num]
                    if not matching_rows.empty:
                        source_bom = str(matching_rows.iloc[0][SOURCE_COL])
            except Exception:
                source_bom = ""
            
            desc_lines = [
                f"Dla elementu '{full_name}' wykryto pliki z nieprawidłową nazwą ({len(mismatched_files)} plik(i/ów)).",
                "",
                f"BOM źródłowy: {source_bom}" if source_bom else "BOM źródłowy: (nieznany)",
                f"Oczekiwana nazwa z BOM: '{expected_name}'",
                "",
                "Pliki z błędną nazwą:",
            ]
            
            for file_path, file_type, found_name in mismatched_files:
                desc_lines.append(f"  • {file_type}: '{found_name}' (plik: {file_path.name})")
            
            desc_lines.extend([
                "",
                "Prawdopodobnie:",
                "1) Eksporter (makro) użył błędnej nazwy z iProperties modelu 3D zamiast z BOM",
                "2) Nazwa w iProperties (Title/Description) modelu IPT jest nieaktualna",
                "",
                "Popraw nazwy plików lub zaktualizuj iProperties w modelu IPT."
            ])
            
            errors.append([
                f"PLIKI – niezgodność nazw (BOM vs pliki {types_str})",
                mismatched_files[0][0].name,
                "\n".join(desc_lines),
                format_file_list([p for p, _, _ in mismatched_files]),
                katalog_for_path(mismatched_files[0][0]),
            ])
    
    return errors


def find_bom_missing_dxf_errors(df_full: "pd.DataFrame", dxf_map: Dict[str, List[Path]],
                                dwf_map_project: Dict[str, List[Path]] | None = None,
                                dwf_map_library: Dict[str, List[Path]] | None = None) -> List[List[str]]:
    """
    BŁĘDY/WARN: elementy z sufiksem X/XX w BOM powinny mieć DXF (dla materiału grubości),
    ale DXF może nie istnieć albo nie dać się dopasować.
    Uwaga: NIE zmieniamy danych w Excelu – sprawdzamy wyłącznie klucz logiczny.

    Jeśli DWF elementu nie istnieje w projekcie, ale istnieje w BIBLIOTECE,
    to element jest "biblioteczny" – brak DXF w projekcie jest naturalny i NIE zgłaszamy błędu.

    Zwraca rekordy: [Typ błędu, Nazwa pliku, Opis, Katalog]
    """
    errors: List[List[str]] = []

    if df_full is None or df_full.empty:
        return errors

    # szybki lookup
    dxf_keys = set(dxf_map.keys())
    _proj = dwf_map_project or {}
    _lib = dwf_map_library or {}

    # iteruj po wierszach BOM
    for _, row in df_full.iterrows():
        nr_raw = row.get("Nr rysunku", "")
        nr_str = "" if nr_raw is None else str(nr_raw)

        nr_clean = normalize_drawing_for_match(nr_str)
        if not nr_clean:
            continue

        # tylko numery RMPAK z X/XX (cięcie / cięcie+gięcie)
        if not is_rmpak_number(nr_clean):
            continue
        if not (nr_clean.endswith("X") or nr_clean.endswith("XX")):
            continue

        key = normalize_drawing_for_match(nr_clean)
        if key not in dxf_keys:
            # Jeśli DWF nie ma w projekcie, ale jest w bibliotece → element biblioteczny,
            # brak DXF w projekcie jest naturalny → nie zgłaszaj błędu.
            if (key not in _proj or not _proj[key]) and (key in _lib and _lib[key]):
                continue

            src = str(row.get(SOURCE_COL, "")).strip()
            nazwa = str(row.get("Nazwa", "")).strip()

            katalog = str(row.get("Katalog", "")).strip()
            if not katalog:
                katalog = katalog_from_rmpak_number(nr_clean)

            # Nazwa pliku: brakujący DXF (kanonicznie z BOM)
            name_for_err = canonical_name_from_bom(nr_str, nazwa, '.dxf')
            full_name = f"{key} {nazwa}" if nazwa else key
            suffix_desc = "cięcia laserowego i gięcia" if key.endswith("XX") else "cięcia laserowego"

            desc = f"Pozycja '{full_name}' występuje w zestawieniu BOM ({src if src else 'BOM nieznany'}) z sufiksem {'XX' if key.endswith('XX') else 'X'}, co oznacza detal do {suffix_desc}. Brakuje pliku DXF."
            errors.append([
                "BOM – brak wymaganego pliku DXF dla elementu X/XX",
                name_for_err,
                desc,
                format_file_list(dxf_map.get(key, [])),
                katalog,
            ])

    return errors



def find_suffix_consistency_errors(files: Dict[str, object]) -> List[List[str]]:
    """BŁĄD: dla tej samej bazy numeru (np. 2556-100.00 / NS-200.20) w projekcie
    występują różne sufiksy (X/XX/Z/ZZ/brak). To zwykle oznacza bałagan wersji / dublowanie.

    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []

    csv_files: List[Path] = files.get("csv_files", [])
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    idw_map: Dict[str, List[Path]] = files.get("idw_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})

    def split_suffix(num: str) -> Tuple[str, str]:
        n = normalize_drawing_for_match(num)
        for suf in ("XXZ", "XZ", "XX", "ZZ", "YZ", "Z", "X", "Y"):
            if n.endswith(suf):
                return n[:-len(suf)], suf
        return n, "(brak)"

    by_base: Dict[str, Dict[str, List[Path]]] = {}

    # CSV
    for p in csv_files:
        num = extract_drawing_number_from_filename(p)
        base, suf = split_suffix(num)
        by_base.setdefault(base, {}).setdefault(suf, []).append(p)

    # Pozostałe mapy (już po normalizacji klucza)
    def _add_map(mp: Dict[str, List[Path]]):
        for num, paths in mp.items():
            base, suf = split_suffix(num)
            by_base.setdefault(base, {}).setdefault(suf, []).extend(paths)

    _add_map(dwf_map)
    _add_map(dxf_map)
    _add_map(idw_map)
    _add_map(stp_map)
    _add_map(stl_map)

    for base, suf_map in by_base.items():
        sufs = [s for s, ps in suf_map.items() if ps]
        if len(set(sufs)) <= 1:
            continue

        # lista plików (unikalne)
        all_paths: List[Path] = []
        for s, ps in suf_map.items():
            all_paths.extend(ps)
        seen = set()
        uniq: List[Path] = []
        for p in all_paths:
            sp = str(p)
            if sp not in seen:
                seen.add(sp)
                uniq.append(p)

        katalog = katalog_for_path(uniq[0]) if uniq else katalog_from_rmpak_number(base)
        sufiksy_sorted = sorted(set(sufs))
        opis = f"Dla numeru bazowego '{base}' wykryto różne warianty sufiksowe: {', '.join(sufiksy_sorted)}.\n\nZnaczenie sufiksów:\n- X = cięcie laserowe\n- XX = cięcie laserowe + gięcie\n- Z/ZZ = element znormalizowany\n- (brak) = element podstawowy\n\nPrzyczyny:\n1) Stare i nowe wersje w jednym projekcie\n2) Błędne duplikowanie plików\n3) Niekonsekwentne nazewnictwo\n\nUstal która wersja jest aktualna i usuń pozostałe lub przenieś do archiwum."

        errors.append([
            "PLIKI – konflikt sufiksów dla tego samego numeru bazowego",
            (uniq[0].name if uniq else base),
            opis,
            format_file_list(uniq),
            katalog,
        ])

    return errors


def find_csv_duplicate_drawing_number_errors(files: Dict[str, object]) -> List[List[str]]:
    """BŁĄD KRYTYCZNY: dwa lub więcej plików CSV w projekcie mają ten sam
    numer rysunku w nazwie (pierwszy token przed spacją), np.:
      - "2621-000.00ZZ Ceramizator_N.csv"
      - "2621-000.00ZZ 2621 Ceramizator_N.csv"

    Każdy z nich jest traktowany jako osobny BOM (klucz __Źródło__ = stem
    pliku) i oba kwalifikują się jako root → pozycje (np. moduł "Kabina")
    są zliczane podwójnie. To psuje wykrywanie ROOT i sumowanie ilości.

    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []
    csv_files: List[Path] = files.get("csv_files", [])  # type: ignore[assignment]

    by_num: Dict[str, List[Path]] = {}
    for p in csv_files:
        num_raw = extract_drawing_number_from_filename(p)
        num = normalize_drawing_for_match(num_raw)
        if not num:
            continue
        by_num.setdefault(num, []).append(p)

    for num, paths in by_num.items():
        if len(paths) <= 1:
            continue

        uniq: List[Path] = []
        seen: set = set()
        for p in paths:
            sp = str(p)
            if sp not in seen:
                seen.add(sp)
                uniq.append(p)
        if len(uniq) <= 1:
            continue

        katalog = katalog_for_path(uniq[0]) if uniq else katalog_from_rmpak_number(num)
        names_list = "\n- " + "\n- ".join(p.name for p in uniq)
        opis = (
            f"W projekcie znaleziono {len(uniq)} plików CSV z TYM SAMYM numerem rysunku '{num}' "
            f"w nazwie (pierwszy token przed spacją):"
            f"{names_list}\n\n"
            "KONSEKWENCJE:\n"
            "- każdy plik jest traktowany jako osobny BOM (klucz źródła = pełna nazwa pliku),\n"
            "- jeśli oba są rootami (lub wskazują ten sam moduł), pozycje (np. 'Kabina') "
            "zostaną zsumowane podwójnie,\n"
            "- wykrywanie ROOT projektu daje błędne wyniki.\n\n"
            "DZIAŁANIE: usuń nieaktualne CSV (lub przenieś do 'OldVersions') tak, "
            "by w projekcie pozostał DOKŁADNIE JEDEN plik CSV o tym numerze rysunku."
        )

        errors.append([
            "PLIKI – zdublowany numer rysunku w nazwach CSV",
            uniq[0].name,
            opis,
            format_file_list(uniq),
            katalog,
        ])

    return errors


def find_bom_missing_files_errors(df_full: pd.DataFrame,
                                  files: Dict[str, object],
                                  active_nums: set[str] | None = None) -> List[List[str]]:
    """
    Błąd: Brak DWF (BOM -> DWF, DWF = prawda)

    Dla każdej pozycji w BOM (CSV):
      - Nr rysunku wygląda na RMPAK (is_rmpak_number == True)
      - po normalizacji (białe znaki / warianty myślnika) brak pliku DWF o takim numerze

    Dodatkowo (czytelniej):
      - jeśli BOM używa numeru bez sufiksu Z/ZZ, a w projekcie istnieje wariant Z/ZZ,
        zgłaszamy błąd: "BOM – brak sufiksu Z/ZZ" zamiast "Brak DWF".

    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []

    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    dwf_map_library: Dict[str, List[Path]] = files.get("dwf_map_library", {})
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})
    idw_map: Dict[str, List[Path]] = files.get("idw_map", {})

    # Numery z DWF w projekcie LUB w bibliotece – dla tych nie zgłaszamy "brak DWF"
    dwf_numbers = set(dwf_map.keys()) | {k for k, v in dwf_map_library.items() if v}

    def is_active(num: str) -> bool:
        if active_nums is None:
            return True
        n = normalize_drawing_for_match(num)
        return (n in active_nums)

    def _collect_related(num_key: str) -> List[Path]:
        out: List[Path] = []
        for mp in (dxf_map, stp_map, stl_map, idw_map, dwf_map):
            out.extend(mp.get(num_key, []))
        # unikalne, zachowując kolejność
        seen = set()
        uniq = []
        for p in out:
            sp = str(p)
            if sp not in seen:
                seen.add(sp)
                uniq.append(p)
        return uniq

    if SOURCE_COL not in df_full.columns:
        grouped = df_full.groupby("Nr rysunku", dropna=False)
        for num, group in grouped:
            num_raw = str(num)
            num_key = normalize_drawing_for_match(num_raw)
            if not num_key:
                continue
            if not is_rmpak_number(num_key):
                continue
            if not is_active(num_key):
                continue
            if num_key in dwf_numbers:
                continue

            katalog_vals = group.get("Katalog", pd.Series(dtype=str)).astype(str).tolist()
            katalog = next((normalize_ws_for_match(v) for v in katalog_vals if normalize_ws_for_match(v)), "")

            related = _collect_related(num_key)

            # brak sufiksu Z/ZZ?
            cand_paths: List[Path] = []
            for suf in ("Z", "ZZ"):
                cand_key = normalize_drawing_for_match(num_key + suf)
                cand_paths.extend(_collect_related(cand_key))
            cand_paths = [p for p in cand_paths if p]  # filtr

            if cand_paths:
                errors.append([
                    "BOM – nieprawidłowy numer (brak sufiksu Z lub ZZ)",
                    num_key,
                    f"W BOM użyto '{num_key}' (bez sufiksu), ale w projekcie istnieją pliki '{num_key}Z' lub '{num_key}ZZ'.\n\nSufiks Z/ZZ = element znormalizowany (z biblioteki standardów).\n\nPrzyczyny:\n1) Błędne wprowadzenie numeru - dodaj sufiks Z lub ZZ\n2) Nieaktualna wersja BOM - projekt używa już wariantu znormalizowanego\n\nZaktualizuj numer w BOM.",
                    format_file_list(cand_paths),
                    katalog,
                ])
            else:
                errors.append([
                    "BOM – brak pliku DWF dla pozycji z zestawienia",
                    expected,
                    f"Pozycja '{num_key}' występuje w zestawieniu BOM ({bom_name_str}), ale w projekcie brakuje pliku DWF.\n\nSprawdź czy:\n1) Plik został wygenerowany i umieszczony we właściwym katalogu\n2) Nazwa pliku jest zgodna z oczekiwaną: {expected}\n3) Element nie został omyłkowo usunięty z projektu",
                    format_file_list(related),
                    katalog,
                ])
        return errors

    for bom_name, g in df_full.groupby(SOURCE_COL, sort=False):
        bom_name_str = str(bom_name)
        katalog_vals = g.get("Katalog", pd.Series(dtype=str)).astype(str).tolist()
        katalog = next((normalize_ws_for_match(v) for v in katalog_vals if normalize_ws_for_match(v)), "")
        if not katalog:
            katalog = katalog_from_rmpak_number(extract_number_from_source_name(bom_name_str))
        bom_num_clean = normalize_drawing_for_match(extract_number_from_source_name(bom_name_str))

        for _, __row in g.iterrows():
            num_raw = str(__row.get("Nr rysunku", ""))
            nazwa_raw = __row.get("Nazwa", "")
            num_key = normalize_drawing_for_match(num_raw)
            if not num_key:
                continue
            # Uwaga: Braki DWF/DXF/STP/STL z BOM sprawdzamy także dla numerów niestandardowych (np. NS-...)
            expected = canonical_name_from_bom(num_raw, nazwa_raw, ".dwf")
            expected_no_name = canonical_name_from_bom(num_raw, "", ".dwf")
            is_self_ref = bool(bom_num_clean) and (num_key == bom_num_clean)

            # 1) Jeśli jest jakiś DWF pod numer — sprawdź nazwę tylko gdy BOM ma Nazwę.
            #    Gdy BOM nie ma Nazwy, akceptujemy dowolny DWF z tym numerem (np. z dopiskiem nazwy w pliku).
            if num_key in dwf_numbers:
                dwf_paths = dwf_map.get(num_key, [])
                bom_name = "" if nazwa_raw is None else str(nazwa_raw).strip()
                if not bom_name:
                    continue

                expected_stem_norm = normalize_ws_for_match(Path(expected).stem)
                expected_no_name_stem_norm = normalize_ws_for_match(Path(expected_no_name).stem)
                def _matches_expected(p: Path) -> bool:
                    try:
                        if p.name.lower() == expected.lower():
                            return True
                        stem_norm = normalize_ws_for_match(p.stem)
                        if stem_norm == expected_stem_norm:
                            return True
                        # Dla samoodwołania w BOM: akceptuj też plik bez nazwy (tylko numer)
                        if is_self_ref and stem_norm == expected_no_name_stem_norm:
                            return True
                        return False
                    except Exception:
                        return False

                if dwf_paths and (not any(_matches_expected(p) for p in dwf_paths)):
                    related = _collect_related(num_key)
                    nazwa_str = "" if nazwa_raw is None else str(nazwa_raw).strip()
                    full_name = f"{num_key} {nazwa_str}" if nazwa_str else num_key
                    actual_files = ", ".join([p.name for p in dwf_paths[:3]]) + ("..." if len(dwf_paths) > 3 else "")
                    errors.append([
                        "BOM/PLIKI – niezgodność nazwy pliku DWF z BOM",
                        dwf_paths[0].name,
                        f"Dla pozycji '{full_name}' z BOM ({bom_name_str}) nazwa pliku DWF nie zgadza się z oczekiwaną.\n\nOczekiwana: {expected}\nZnaleziona: {actual_files}\n\nPrzyczyny:\n1) Plik został skopiowany/przemianowany ręcznie\n2) Nazwa w pliku różni się od nazwy w BOM\n3) Nieaktualna wersja pliku\n\nZmień nazwę pliku na oczekiwaną lub zaktualizuj BOM.",
                        format_file_list(related),
                        katalog,
                    ])
                continue

            # 2) Brak jakiegokolwiek DWF pod numer → klasyczny błąd Brak DWF (nazwa pliku kanoniczna z BOM)
            related = _collect_related(num_key)

            # brak sufiksu Z/ZZ?
            cand_paths: List[Path] = []
            for suf in ("Z", "ZZ"):
                cand_key = normalize_drawing_for_match(num_key + suf)
                cand_paths.extend(_collect_related(cand_key))
            # unikalne
            seen=set(); cand=[]
            for p in cand_paths:
                sp=str(p)
                if sp not in seen:
                    seen.add(sp); cand.append(p)

            if cand:
                nazwa_str = "" if nazwa_raw is None else str(nazwa_raw).strip()
                full_name = f"{num_key} {nazwa_str}" if nazwa_str else num_key
                errors.append([
                    "BOM – nieprawidłowy numer (brak sufiksu Z lub ZZ)",
                    expected,
                    f"W BOM ({bom_name_str}) pozycja '{full_name}' użyta bez sufiksu, ale w projekcie istnieją pliki '{num_key}Z' lub '{num_key}ZZ'.\n\nSufiks Z/ZZ = element znormalizowany (z biblioteki standardów).\n\nPrzyczyny:\n1) Błędne wprowadzenie - dodaj sufiks Z lub ZZ\n2) Nieaktualna wersja BOM - projekt używa już wariantu znormalizowanego\n\nZaktualizuj numer w BOM.",
                    format_file_list(cand),
                    katalog,
                ])
            else:
                nazwa_str = "" if nazwa_raw is None else str(nazwa_raw).strip()
                full_name = f"{num_key} {nazwa_str}" if nazwa_str else num_key
                errors.append([
                    "BOM – brak pliku DWF dla pozycji z zestawienia",
                    expected,
                    f"Pozycja '{full_name}' występuje w zestawieniu BOM ({bom_name_str}), ale w projekcie brakuje pliku DWF.\n\nSprawdź czy:\n1) Plik został wygenerowany i umieszczony we właściwym katalogu\n2) Nazwa pliku jest zgodna z oczekiwaną: {expected}\n3) Element nie został omyłkowo usunięty z projektu",
                    format_file_list(related),
                    katalog,
                ])

    return errors




# --------------------------------
# BŁĘDY – ANALIZA BOM
# --------------------------------

def find_bom_duplicate_in_single_bom_errors(df_full: pd.DataFrame, csv_path_by_source: Dict[str, Path] | None = None) -> List[List[str]]:
    """
    Błąd: "BOM – detal wielorolkowy (duplikat w 1 BOM)"

    Dotyczy TYLKO pozycji, które mają Nr rysunku będący numerem RMPAK.
    Nie dotyczy elementów handlowych (puste / nan / brak numeru RMPAK).
    """
    errors: List[List[str]] = []
    if SOURCE_COL not in df_full.columns:
        return errors

    # przygotuj kolumnę techniczną do porównań
    df = df_full.copy()
    df["__NR_KEY__"] = df["Nr rysunku"].map(normalize_drawing_for_match)

    # tylko RMPAK
    df = df[df["__NR_KEY__"].map(is_rmpak_number)]

    if df.empty:
        return errors

    # indeks wystąpień w projekcie: nr -> list[(BOM, Poz, Ilość, Nazwa)]
    occ: Dict[str, List[Tuple[str, str, str, str]]] = {}
    for _, r in df.iterrows():
        k = r.get("__NR_KEY__", "")
        if not k:
            continue
        occ.setdefault(k, []).append((
            str(r.get(SOURCE_COL, "")),
            str(r.get("Poz.", "")),
            str(r.get("Ilość", "")),
            str(r.get("Nazwa", "")),
        ))

    # duplikaty w obrębie pojedynczego BOM
    for bom_name, df_bom in df.groupby(SOURCE_COL, dropna=False):
        bom_name_str = str(bom_name).strip()
        counts = df_bom.groupby("__NR_KEY__").size()
        dup_keys = [k for k, v in counts.items() if v >= 2]

        for key in dup_keys:
            rows = df_bom[df_bom["__NR_KEY__"] == key]
            katalog = ""
            if "Katalog" in rows.columns:
                katalog = next((normalize_ws_for_match(v) for v in rows["Katalog"].astype(str).tolist()
                                if normalize_ws_for_match(v)), "")

            local_lines = []
            for _, r in rows.iterrows():
                local_lines.append(f"Poz {r.get('Poz.', '')}: {r.get('Nazwa', '')} | Ilość={r.get('Ilość', '')}")

            all_occ = occ.get(key, [])
            bom_list = sorted({b for (b, _, _, _) in all_occ})
            global_lines = [f"{b} → Poz {poz} | Ilość={ilosc} | {nazwa}" for (b, poz, ilosc, nazwa) in all_occ]

            # Pobierz nazwę dla pierwszego wystąpienia
            first_nazwa = ""
            for _, r in rows.iterrows():
                nazwa_val = r.get('Nazwa', '')
                if nazwa_val and str(nazwa_val).strip():
                    first_nazwa = str(nazwa_val).strip()
                    break
            full_name = f"{key} {first_nazwa}" if first_nazwa else key
            
            opis = (
                f"KRYTYCZNY: Element '{full_name}' występuje WIELOKROTNIE w tym samym BOM.\n\n"
                f"BOM: {bom_name_str}\n\n"
                f"Duplikaty:\n- " + "\n- ".join(local_lines) +
                f"\n\nPrzyczyny:\n"
                f"1) Błąd w zestawieniu - element dodany kilka razy\n"
                f"2) Element w różnych podzespołach (OK w strukturze, NIE w płaskim BOM)\n"
                f"3) Różne warianty (powinny mieć różne numery)\n\n"
                f"Rozwiązanie: Scal wystąpienia w jedną pozycję ze zsumowaną ilością lub rozdziel na osobne numery.\n\n"
                f"Element także w BOM: {', '.join(bom_list)}\n\n"
                f"Wszystkie wystąpienia:\n- " + "\n- ".join(global_lines)
            )

            errors.append([
                "BOM – duplikat pozycji w zestawieniu (ten sam element wiele razy)",
                (csv_path_by_source.get(bom_name_str).name if (csv_path_by_source and csv_path_by_source.get(bom_name_str)) else f"{bom_name_str}.csv"),
                opis,
                format_file_list([csv_path_by_source.get(bom_name_str)]) if (csv_path_by_source and csv_path_by_source.get(bom_name_str)) else "",
                katalog,
            ])

    return errors


def find_file_number_mismatch_errors(files: Dict[str, object], df_full: pd.DataFrame, name_by_num: Dict[str, str] | None = None) -> List[List[str]]:
    """
    Błąd: "PLIKI – niezgodność numeru pliku z BOM"
    
    Sprawdza czy numery wyciągnięte z nazw plików rzeczywiście istnieją w BOM.
    Wykrywa przypadki, gdy:
    1) Plik ma numer nieistniejący w BOM
    2) Nazwa pliku pasuje do innej pozycji BOM (ta sama nazwa, ale inny numer)
    
    Przykład błędu:
        BOM: "444-555.66 Nazwa Elementu"
        Plik: "111-222.33 Nazwa Elementu.dwf"
        → BŁĄD: numer w pliku (111-222.33) nie pasuje do numeru w BOM (444-555.66)
    """
    errors: List[List[str]] = []
    
    if df_full is None or df_full.empty:
        return errors
    
    # Zbierz wszystkie numery z BOM
    bom_numbers = set()
    if "Nr rysunku" in df_full.columns:
        for nr in df_full["Nr rysunku"].astype(str):
            num_key = normalize_drawing_for_match(nr)
            if num_key:
                bom_numbers.add(num_key)
    
    # Zbierz mapę: nazwa -> numery w BOM (znormalizowane nazwy)
    name_to_bom_numbers: Dict[str, List[str]] = {}
    if "Nr rysunku" in df_full.columns and "Nazwa" in df_full.columns:
        for _, row in df_full.iterrows():
            nr = str(row.get("Nr rysunku", "") or "")
            num_key = normalize_drawing_for_match(nr)
            if not num_key:
                continue
            nazwa = str(row.get("Nazwa", "") or "").strip()
            nazwa_norm = normalize_ws_for_match(nazwa)
            if nazwa_norm:
                name_to_bom_numbers.setdefault(nazwa_norm, []).append(num_key)
    
    # Zbierz wszystkie pliki per numer
    idw_files: List[Path] = files.get("idw_files", [])
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})

    # Numery podzespołów (assembly), dla których istnieje własny sub-BOM (CSV).
    # Taki rysunek/plik to legalny rysunek złożenia – nawet jeśli sam numer nie
    # występuje w żadnym wczytanym df_full (np. brak BOM-a rodzica w bieżącym
    # zakresie), nie jest to "błędny numer". Zapobiega false-positive np. dla
    # "2621-300.10ZZ Kostka nalewaka.idw" gdy w sub-BOM 2621-300.10ZZ jest pozycja
    # "2621-300.06 Kostka nalewaka" o tej samej nazwie.
    assembly_csv_numbers: set = set()
    for csv_path in files.get("csv_files", []) or []:
        try:
            csv_num = extract_drawing_number_from_filename(csv_path)
            csv_num_key = normalize_drawing_for_match(csv_num)
            if csv_num_key:
                assembly_csv_numbers.add(csv_num_key)
        except Exception:
            pass

    # Sprawdź pliki IDW
    for path in idw_files:
        num_from_file = extract_drawing_number_from_filename(path)
        num_key = normalize_drawing_for_match(num_from_file)
        
        if not num_key:
            continue
        
        # Sprawdź czy numer z pliku istnieje w BOM (lub jest numerem złożenia z własnym sub-BOM)
        if num_key not in bom_numbers and num_key not in assembly_csv_numbers:
            # Wyciągnij nazwę z pliku
            stem = path.stem
            if stem.startswith(num_from_file):
                name_from_file = stem[len(num_from_file):].strip()
                name_from_file_norm = normalize_ws_for_match(name_from_file)
                
                # Sprawdź czy ta nazwa występuje w BOM pod innym numerem
                matching_bom_numbers = name_to_bom_numbers.get(name_from_file_norm, [])
                if matching_bom_numbers:
                    # Znaleziono tę samą nazwę w BOM, ale z innym numerem!
                    expected_num = matching_bom_numbers[0]
                    
                    # Znajdź źródło BOM dla oczekiwanego numeru
                    source_bom = ""
                    try:
                        if SOURCE_COL in df_full.columns:
                            nr_normalized = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
                            matching_rows = df_full[nr_normalized == expected_num]
                            if not matching_rows.empty:
                                source_bom = str(matching_rows.iloc[0][SOURCE_COL])
                    except Exception:
                        pass
                    
                    # Pobierz katalog
                    katalog = ""
                    try:
                        if SOURCE_COL in df_full.columns:
                            nr_normalized = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
                            matching_rows = df_full[nr_normalized == expected_num]
                            if not matching_rows.empty and "Katalog" in matching_rows.columns:
                                katalog_val = matching_rows.iloc[0].get("Katalog", "")
                                katalog = normalize_ws_for_match(str(katalog_val))
                    except Exception:
                        pass
                    
                    desc = (
                        f"KRYTYCZNY: Plik '{path.name}' ma BŁĘDNY NUMER w nazwie!\n\n"
                        f"Numer w pliku: '{num_from_file}'\n"
                        f"Nazwa w pliku: '{name_from_file}'\n\n"
                        f"BOM źródłowy: {source_bom if source_bom else '(nieznany)'}\n"
                        f"Oczekiwany numer z BOM: '{expected_num}'\n\n"
                        f"Element o tej nazwie ('{name_from_file}') występuje w BOM\n"
                        f"pod numerem '{expected_num}', a nie '{num_from_file}'.\n\n"
                        f"Prawdopodobnie:\n"
                        f"1) Plik został skopiowany i zmieniono tylko nazwę, ale nie numer\n"
                        f"2) Makro eksportujące użyło błędnego numeru z modelu 3D\n"
                        f"3) Numer w iProperties modelu IPT jest nieaktualny\n\n"
                        f"Poprawna nazwa pliku powinna być: '{expected_num} {name_from_file}.idw'\n\n"
                        f"UWAGA: Ten plik nie zostanie przypisany do żadnej pozycji BOM!"
                    )
                    
                    errors.append([
                        "PLIKI – niezgodność numeru pliku z numerem w BOM",
                        path.name,
                        desc,
                        format_file_list([path]),
                        katalog,
                    ])
    
    # Sprawdź pliki DXF/DWF/STP/STL
    for file_map, ext_name in [
        (dxf_map, "DXF"),
        (dwf_map, "DWF"),
        (stp_map, "STP"),
        (stl_map, "STL"),
    ]:
        for num_key, paths in file_map.items():
            # Sprawdź czy numer z pliku istnieje w BOM (lub jest numerem złożenia z własnym sub-BOM)
            if num_key not in bom_numbers and num_key not in assembly_csv_numbers:
                for path in paths:
                    num_from_file = extract_drawing_number_from_filename(path)
                    
                    # Wyciągnij nazwę z pliku
                    stem = path.stem
                    if stem.startswith(num_from_file):
                        name_from_file = stem[len(num_from_file):].strip()
                        
                        # Dla DXF usuń część z materiałem (po przecinku)
                        if ext_name == "DXF" and ", " in name_from_file:
                            name_from_file = name_from_file.split(", ", 1)[0].strip()
                        
                        name_from_file_norm = normalize_ws_for_match(name_from_file)
                        
                        # Sprawdź czy ta nazwa występuje w BOM pod innym numerem
                        matching_bom_numbers = name_to_bom_numbers.get(name_from_file_norm, [])
                        if matching_bom_numbers:
                            # Znaleziono tę samą nazwę w BOM, ale z innym numerem!
                            expected_num = matching_bom_numbers[0]
                            
                            # Znajdź źródło BOM dla oczekiwanego numeru
                            source_bom = ""
                            try:
                                if SOURCE_COL in df_full.columns:
                                    nr_normalized = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
                                    matching_rows = df_full[nr_normalized == expected_num]
                                    if not matching_rows.empty:
                                        source_bom = str(matching_rows.iloc[0][SOURCE_COL])
                            except Exception:
                                pass
                            
                            # Pobierz katalog
                            katalog = ""
                            try:
                                if SOURCE_COL in df_full.columns:
                                    nr_normalized = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
                                    matching_rows = df_full[nr_normalized == expected_num]
                                    if not matching_rows.empty and "Katalog" in matching_rows.columns:
                                        katalog_val = matching_rows.iloc[0].get("Katalog", "")
                                        katalog = normalize_ws_for_match(str(katalog_val))
                            except Exception:
                                pass
                            
                            desc = (
                                f"KRYTYCZNY: Plik '{path.name}' ma BŁĘDNY NUMER w nazwie!\n\n"
                                f"Numer w pliku: '{num_from_file}'\n"
                                f"Nazwa w pliku: '{name_from_file}'\n\n"
                                f"BOM źródłowy: {source_bom if source_bom else '(nieznany)'}\n"
                                f"Oczekiwany numer z BOM: '{expected_num}'\n\n"
                                f"Element o tej nazwie ('{name_from_file}') występuje w BOM\n"
                                f"pod numerem '{expected_num}', a nie '{num_from_file}'.\n\n"
                                f"Prawdopodobnie:\n"
                                f"1) Plik został skopiowany i zmieniono tylko nazwę, ale nie numer\n"
                                f"2) Makro eksportujące użyło błędnego numeru z modelu 3D\n"
                                f"3) Numer w iProperties modelu IPT jest nieaktualny\n\n"
                                f"Poprawna nazwa pliku powinna być: '{expected_num} {name_from_file}.{ext_name.lower()}'\n\n"
                                f"UWAGA: Ten plik nie zostanie przypisany do żadnej pozycji BOM!"
                            )
                            
                            errors.append([
                                f"PLIKI – niezgodność numeru pliku z numerem w BOM ({ext_name})",
                                path.name,
                                desc,
                                format_file_list([path]),
                                katalog,
                            ])
                            break  # Jeden błąd per numer wystarczy
    
    return errors


def find_bom_name_mismatch_errors(df_full: pd.DataFrame, csv_path_by_source: Dict[str, Path] | None = None) -> List[List[str]]:
    """
    Błąd: "BOM – różne nazwy dla tego samego Nr rysunku" (TYLKO RMPAK)

    Ten sam Nr rysunku (po normalizacji białych znaków) występuje w projekcie
    z różnymi wartościami kolumny "Nazwa".
    """
    errors: List[List[str]] = []

    df = df_full.copy()
    df["__NR_KEY__"] = df["Nr rysunku"].map(normalize_drawing_for_match)
    df = df[df["__NR_KEY__"].map(is_rmpak_number)]
    if df.empty:
        return errors

    # group by nr_key
    for key, g in df.groupby("__NR_KEY__"):
        names = sorted({normalize_ws_for_match(x) for x in g["Nazwa"].astype(str).tolist() if normalize_ws_for_match(x)})
        if len(names) <= 1:
            continue

        # katalog: pierwszy sensowny
        katalog = ""
        if "Katalog" in g.columns:
            katalog = next((normalize_ws_for_match(v) for v in g["Katalog"].astype(str).tolist()
                            if normalize_ws_for_match(v)), "")

        # lista wystąpień
        occ_lines = []
        for _, r in g.iterrows():
            bom = str(r.get(SOURCE_COL, ""))
            occ_lines.append(f"{bom} → Poz {r.get('Poz.', '')} | Ilość={r.get('Ilość', '')} | {r.get('Nazwa', '')}")

        opis = (
            f"NIEKONSEKWENCJA: Element '{key}' występuje w różnych BOM z różnymi nazwami.\n\n"
            f"Wykryte nazwy:\n" + "\n".join([f"  • {n}" for n in names]) + "\n\n"
            f"Ten sam numer powinien mieć jednolitą nazwę w całym projekcie.\n\n"
            f"Rozwiązanie: Ustal JEDNĄ prawidłową nazwę i zaktualizuj wszystkie BOM-y.\n\n"
            f"Wszystkie wystąpienia:\n- " + "\n- ".join(occ_lines)
        )

        # Lista plików: CSV-e, w których wystąpił konflikt
        csv_paths: List[Path] = []
        if csv_path_by_source:
            for bom_name in sorted({str(x) for x in g[SOURCE_COL].astype(str).tolist() if str(x)}):
                p = csv_path_by_source.get(bom_name)
                if p:
                    csv_paths.append(p)

        errors.append([
            "BOM – niespójne nazwy dla tego samego numeru rysunku",
            key,
            opis,
            format_file_list(csv_paths),
            katalog,
        ])

    return errors


def find_bom_spacing_errors(files: Dict[str, object]) -> List[List[str]]:
    """
    Błąd: "FORMAT – Podwójna spacja w nazwie pliku"

    Skanuje fizyczne pliki projektu (IDW, DWF, DXF, STP/STEP, STL)
    i szuka nazw z podwójną spacją, np.:
        '2540-100.04XX  Taca góra.dwf'
    zamiast poprawnego:
        '2540-100.04XX Taca góra.dwf'

    Taki błąd pochodzi z iProperties w Inventorze (trailing space w Nr rysunku
    lub leading space w Nazwie/Title).
    """
    errors: List[List[str]] = []

    # Zbierz wszystkie pliki z map
    all_paths: List[Path] = []
    for map_key in ("idw_files",):
        paths = files.get(map_key, [])
        if isinstance(paths, list):
            all_paths.extend(paths)
    for map_key in ("dwf_map", "dxf_map", "stp_map", "stl_map", "idw_map"):
        mp = files.get(map_key, {})
        if isinstance(mp, dict):
            for paths_list in mp.values():
                all_paths.extend(paths_list)

    # Deduplikuj (idw_files i idw_map mogą się pokrywać)
    seen: set = set()
    unique_paths: List[Path] = []
    for p in all_paths:
        sp = str(p)
        if sp not in seen:
            seen.add(sp)
            unique_paths.append(p)

    for path in unique_paths:
        stem = path.stem  # np. '2540-100.04XX  Taca góra'
        if "  " not in stem:
            continue

        # Poprawna nazwa = kolaps podwójnych spacji
        fixed_stem = re.sub(r"  +", " ", stem)
        ext = path.suffix  # np. '.dwf'

        opis = (
            f"Plik '{stem}{ext}' zawiera podwójną spację w nazwie.\n\n"
            f"Aktualna nazwa:  '{stem}{ext}'\n"
            f"Oczekiwana nazwa: '{fixed_stem}{ext}'\n\n"
            f"Przyczyny:\n"
            f"1) Spacja na końcu numeru rysunku w iProperties (Part Number)\n"
            f"2) Spacja na początku nazwy/tytułu w iProperties (Title/Description)\n"
            f"3) Błąd w eksporterze (makro) generującym pliki z modelu\n\n"
            f"Otwórz plik IDW i popraw spacje w iProperties, następnie ponownie wyeksportuj."
        )

        errors.append([
            "FORMAT – Podwójna spacja w nazwie pliku",
            f"{stem}{ext}",
            opis,
            format_file_list([path]),
            katalog_for_path(path),
        ])

    return errors


def find_missing_csv_for_assemblies_errors(files: Dict[str, object], csv_files: List[Path], name_by_num: Dict[str, str] | None = None) -> List[List[str]]:
    """
    Błąd: "PLIKI – brak pliku CSV (BOM) dla złożenia/modułu Z/ZZ"
    
    Sprawdza czy dla plików IDW/DWF z sufiksem Z lub ZZ istnieje odpowiadający plik CSV (BOM).
    
    Sufiks Z = ZŁOŻENIE (Z) - wymaga zestawienia BOM
    Sufiks ZZ = MODUŁ (ZZ) - wymaga zestawienia BOM z podzespołami
    
    Args:
        files: Zmapowane pliki z projektu
        csv_files: Lista plików CSV (BOM)
        name_by_num: Mapa numer -> nazwa dla wyświetlania pełnych nazw
    
    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []
    
    if name_by_num is None:
        name_by_num = {}
    
    # Zbierz numery z plików CSV (BOM)
    csv_numbers = set()
    for csv_path in csv_files:
        num_raw = extract_drawing_number_from_filename(csv_path)
        num = normalize_drawing_for_match(num_raw)
        if num and is_rmpak_number(num):
            csv_numbers.add(num)
    
    # Sprawdź pliki DWF
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    for num, paths in dwf_map.items():
        if not num:
            continue
        # Tylko numery z sufiksem Z lub ZZ
        if not (num.endswith("Z") or num.endswith("ZZ")):
            continue
        # Pomijamy jeśli to tylko pojedyncze "Z" będące częścią numeru (np. "NS-123.45Z" gdzie Z nie jest sufiksem)
        if not is_rmpak_number(num):
            continue
        
        # Sprawdź czy istnieje CSV dla tego numeru
        if num not in csv_numbers:
            katalog = katalog_for_path(paths[0]) if paths else katalog_from_rmpak_number(num)
            
            # Określ typ złożenia
            assembly_type = "MODUŁ (ZZ)" if num.endswith("ZZ") else "ZŁOŻENIE (Z)"
            
            # Pobierz nazwę elementu z name_by_num, jeśli nie ma - wyciągnij z pliku DWF
            nazwa = name_by_num.get(num, "")
            if not nazwa and paths:
                stem = paths[0].stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            full_name = f"{num} {nazwa}" if nazwa else num
            
            # Oczekiwana nazwa CSV z pełną nazwą
            expected_csv = f"{full_name}.csv"
            
            opis = (
                f"Dla elementu '{full_name}' (typ: {assembly_type}) istnieje plik DWF ({paths[0].name}), "
                f"ale brakuje odpowiadającego pliku CSV z zestawieniem BOM.\n\n"
                f"Złożenia i moduły (sufiks Z/ZZ) WYMAGAJĄ własnego zestawienia BOM zawierającego:\n"
                f"- Listę komponentów / podzespołów\n"
                f"- Ilości każdego elementu\n"
                f"- Nazwy i specyfikacje\n\n"
                f"Oczekiwany plik CSV: {expected_csv}\n\n"
                f"Rozwiązanie:\n"
                f"1) Wygeneruj BOM z Inventora (File → Export → BOM)\n"
                f"2) Zapisz jako CSV w odpowiednim katalogu ({katalog})\n"
                f"3) Upewnij się że nazwa pliku zawiera pełny numer z sufiksem: {num}"
            )
            
            errors.append([
                f"PLIKI – brak pliku CSV (BOM) dla {assembly_type}",
                expected_csv,
                opis,
                format_file_list(paths),
                katalog,
            ])
    
    # Sprawdź pliki IDW (mogą być bez DWF)
    idw_files: List[Path] = files.get("idw_files", [])
    checked_nums = set(dwf_map.keys())  # Już sprawdzone w DWF
    
    for path in idw_files:
        num = normalize_drawing_for_match(extract_drawing_number_from_filename(path))
        if not num:
            continue
        # Tylko numery z sufiksem Z lub ZZ, które nie były sprawdzone w DWF
        if num in checked_nums:
            continue
        if not (num.endswith("Z") or num.endswith("ZZ")):
            continue
        if not is_rmpak_number(num):
            continue
        
        # Sprawdź czy istnieje CSV dla tego numeru
        if num not in csv_numbers:
            katalog = katalog_for_path(path)
            
            # Określ typ złożenia
            assembly_type = "MODUŁ (ZZ)" if num.endswith("ZZ") else "ZŁOŻENIE (Z)"
            
            # Pobierz nazwę elementu z name_by_num, jeśli nie ma - wyciągnij z pliku IDW
            nazwa = name_by_num.get(num, "")
            if not nazwa:
                stem = path.stem
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    nazwa = stem[first_space_idx + 1:].strip()
            full_name = f"{num} {nazwa}" if nazwa else num
            
            # Oczekiwana nazwa CSV z pełną nazwą
            expected_csv = f"{full_name}.csv"
            
            opis = (
                f"Dla elementu '{full_name}' (typ: {assembly_type}) istnieje plik IDW ({path.name}), "
                f"ale brakuje odpowiadającego pliku CSV z zestawieniem BOM.\n\n"
                f"Złożenia i moduły (sufiks Z/ZZ) WYMAGAJĄ własnego zestawienia BOM zawierającego:\n"
                f"- Listę komponentów / podzespołów\n"
                f"- Ilości każdego elementu\n"
                f"- Nazwy i specyfikacje\n\n"
                f"Oczekiwany plik CSV: {expected_csv}\n\n"
                f"Rozwiązanie:\n"
                f"1) Wygeneruj BOM z Inventora (File → Export → BOM)\n"
                f"2) Zapisz jako CSV w odpowiednim katalogu ({katalog})\n"
                f"3) Upewnij się że nazwa pliku zawiera pełny numer z sufiksem: {num}"
            )
            
            errors.append([
                f"PLIKI – brak pliku CSV (BOM) dla {assembly_type}",
                expected_csv,
                opis,
                format_file_list([path]),
                katalog,
            ])
    
    return errors


def find_orphaned_files_errors(files: Dict[str, object], df_full: pd.DataFrame, exclude_numbers: set[str] | None = None, name_by_num: Dict[str, str] | None = None) -> List[List[str]]:
    """
    BŁĄD: "PLIKI – nieprzypisane do żadnego BOM"
    
    Pliki w projekcie które:
    - Mają prawidłowy numer RMPAK (is_rmpak_number == True)
    - NIE występują w żadnym BOM-ie (CSV)
    - NIE zostały już zgłoszone w błędach o wyższym priorytecie (exclude_numbers)
    
    To jest "zamiatanie końcowe" - wykrywa pliki które mogły być:
    - Zapomniane przy tworzeniu BOM-u
    - Stare wersje nie usunięte z projektu
    - Błędnie nazwane (powinny być w BOM ale przez literówkę nie pasują)
    
    Args:
        files: Zmapowane pliki z projektu
        df_full: Pełny DataFrame BOM
        name_by_num: Mapa numer -> nazwa dla wyświetlania pełnych nazw
    
    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []
    
    if exclude_numbers is None:
        exclude_numbers = set()
    
    if name_by_num is None:
        name_by_num = {}
    
    # Zbierz wszystkie numery z BOM
    bom_numbers = set()
    if df_full is not None and not df_full.empty and "Nr rysunku" in df_full.columns:
        for num in df_full["Nr rysunku"].astype(str).tolist():
            key = normalize_drawing_for_match(num)
            if key and is_rmpak_number(key):
                bom_numbers.add(key)
    
    # Zbierz wszystkie numery z plików w projekcie
    file_numbers: Dict[str, List[Path]] = {}  # {numer: [pliki]}
    file_names: Dict[str, str] = {}  # {numer: nazwa z pliku}
    
    idw_files: List[Path] = files.get("idw_files", [])
    dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
    dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
    stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
    stl_map: Dict[str, List[Path]] = files.get("stl_map", {})
    
    # IDW - wyciągnij nazwy z plików
    for path in idw_files:
        num_raw = extract_drawing_number_from_filename(path)
        num = normalize_drawing_for_match(num_raw)
        if num and is_rmpak_number(num):
            file_numbers.setdefault(num, []).append(path)
            # Wyciągnij nazwę z nazwy pliku
            if num not in file_names:
                stem = path.stem
                # Znajdź pierwszą spację i weź wszystko po niej jako nazwę
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    name_from_file = stem[first_space_idx + 1:].strip()
                    if name_from_file:
                        file_names[num] = name_from_file
    
    # DWF - też wyciągnij nazwy
    for num, paths in dwf_map.items():
        if num and is_rmpak_number(num):
            file_numbers.setdefault(num, []).extend(paths)
            # Wyciągnij nazwę z pierwszego pliku DWF jeśli nie mamy jeszcze nazwy
            if num not in file_names and paths:
                stem = paths[0].stem
                # Znajdź pierwszą spację i weź wszystko po niej jako nazwę
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    name_from_file = stem[first_space_idx + 1:].strip()
                    if name_from_file:
                        file_names[num] = name_from_file
    
    # Pozostałe mapy (bez wyciągania nazw - DXF ma materiał w nazwie)
    for map_obj in [dxf_map, stp_map, stl_map]:
        for num, paths in map_obj.items():
            if num and is_rmpak_number(num):
                file_numbers.setdefault(num, []).extend(paths)
    
    # Znajdź numery które są w plikach ALE:
    # - NIE MA ich w BOM
    # - NIE zostały już zgłoszone w innych błędach (wyższy priorytet)
    orphaned_numbers_raw = sorted(set(file_numbers.keys()) - bom_numbers - exclude_numbers)
    
    # Wyklucz główne zespoły TOP-LEVEL (np. 2608-000.00ZZ, NS-000.00Z)
    # Są to zespoły główne projektu które nie występują w nadrzędnych BOM-ach, ale to jest normalne
    orphaned_numbers = []
    for num in orphaned_numbers_raw:
        # Sprawdź czy to główny zespół (kończy się na .00Z lub .00ZZ)
        num_upper = num.upper()
        if num_upper.endswith(".00ZZ") or num_upper.endswith(".00Z"):
            # To jest główny zespół TOP-LEVEL - pomiń (nie jest osierocony)
            continue
        orphaned_numbers.append(num)

    if not orphaned_numbers:
        return errors

    # Zbierz wszystkie pliki do JEDNEGO wpisu
    all_paths: List[Path] = []
    seen_paths = set()
    num_summary_lines: List[str] = []

    for num in orphaned_numbers:
        paths = file_numbers.get(num, [])
        if not paths:
            continue

        # Unikalne ścieżki globalnie
        for p in paths:
            sp = str(p)
            if sp not in seen_paths:
                seen_paths.add(sp)
                all_paths.append(p)

        # Wyświetl numer z nazwą (jeśli dostępna)
        # Najpierw spróbuj z nazwy pliku, potem z BOM
        nazwa = file_names.get(num) or name_by_num.get(num, "")
        full_name = f"{num} {nazwa}" if nazwa else num
        num_summary_lines.append(f"• {full_name} ({len(paths)} plik(i/ów))")

    if not all_paths:
        return errors

    opis = (
        f"W projekcie znaleziono pliki, które NIE występują w żadnym BOM-ie (CSV).\n\n"
        f"Numery rysunków:\n" + "\n".join(num_summary_lines) + "\n\n"
        f"Możliwe przyczyny:\n"
        f"1) Element zapomniany\n"
        f"2) Stara wersja pliku nie usunięta z projektu\n"
        f"3) Literówka w numerze\n"
        f"4) Element niepotrzebnie skopiowany"
    )

    errors.append([
        "PLIKI – nieprzypisane do żadnego BOM",
        "(zbiorczo)",
        opis,
        format_file_list(all_paths),
        "",
    ])
    
    return errors


def find_orphaned_bom_errors(files: Dict[str, object], df_full: pd.DataFrame, exclude_numbers: set[str] | None = None, name_by_num: Dict[str, str] | None = None) -> List[List[str]]:
    """
    BŁĄD: "PLIKI – Osierocony BOM"
    
    Wykrywa pliki CSV (BOM) dla zespołów które:
    - Mają prawidłowy numer RMPAK (is_rmpak_number == True)
    - NIE występują w żadnym nadrzędnym BOM-ie (CSV)
    - NIE zostały już zgłoszone w błędach o wyższym priorytecie (exclude_numbers)
    - NIE są głównymi zespołami TOP-LEVEL (np. .00Z, .00ZZ)
    
    Osierocony BOM to sytuacja, gdzie:
    - Istnieje plik CSV z zestawieniem dla podzespołu (np. 2539-100.19Z)
    - Ten podzespół NIE występuje w żadnym nadrzędnym zespole
    - Może być zapomniany, stara wersja, albo literówka w numerze
    
    Args:
        files: Zmapowane pliki z projektu
        df_full: Pełny DataFrame BOM
        exclude_numbers: Numery które już zostały zgłoszone w innych błędach (pomijamy je)
        name_by_num: Mapa numer -> nazwa dla wyświetlania pełnych nazw
    
    Format wiersza: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    errors: List[List[str]] = []
    
    if exclude_numbers is None:
        exclude_numbers = set()
    
    if name_by_num is None:
        name_by_num = {}
    
    # Zbierz wszystkie numery z BOM
    bom_numbers = set()
    if df_full is not None and not df_full.empty and "Nr rysunku" in df_full.columns:
        for num in df_full["Nr rysunku"].astype(str).tolist():
            key = normalize_drawing_for_match(num)
            if key and is_rmpak_number(key):
                bom_numbers.add(key)
    
    # Zbierz numery z plików CSV
    csv_files: List[Path] = files.get("csv_files", [])
    csv_numbers: Dict[str, List[Path]] = {}  # {numer: [pliki CSV]}
    csv_names: Dict[str, str] = {}  # {numer: nazwa z pliku CSV}
    
    for path in csv_files:
        num_raw = extract_drawing_number_from_filename(path)
        num = normalize_drawing_for_match(num_raw)
        if num and is_rmpak_number(num):
            csv_numbers.setdefault(num, []).append(path)
            # Wyciągnij nazwę z nazwy pliku CSV (np. "2556-666.02ZZ Bypass nalewarki.csv")
            if num not in csv_names:
                stem = path.stem  # np. "2556-666.02ZZ Bypass nalewarki"
                # Znajdź pierwszą spację i weź wszystko po niej jako nazwę
                first_space_idx = stem.find(' ')
                if first_space_idx > 0:
                    name_from_file = stem[first_space_idx + 1:].strip()
                    if name_from_file:
                        csv_names[num] = name_from_file
    
    # Znajdź numery CSV które:
    # - NIE MA ich w BOM
    # - NIE zostały już zgłoszone w innych błędach
    orphaned_csv_raw = sorted(set(csv_numbers.keys()) - bom_numbers - exclude_numbers)
    
    # Wyklucz główne zespoły TOP-LEVEL (np. 2608-000.00ZZ, NS-000.00Z)
    # Są to zespoły główne projektu które nie występują w nadrzędnych BOM-ach, ale to jest normalne
    orphaned_csv = []
    for num in orphaned_csv_raw:
        # Sprawdź czy to główny zespół (kończy się na .00Z lub .00ZZ)
        num_upper = num.upper()
        if num_upper.endswith(".00ZZ") or num_upper.endswith(".00Z"):
            # To jest główny zespół TOP-LEVEL - pomiń (nie jest osierocony)
            continue
        orphaned_csv.append(num)

    if not orphaned_csv:
        return errors

    # Zbierz wszystkie pliki CSV do JEDNEGO wpisu
    all_csv_paths: List[Path] = []
    seen_paths = set()
    num_summary_lines: List[str] = []

    for num in orphaned_csv:
        paths = csv_numbers.get(num, [])
        if not paths:
            continue

        # Unikalne ścieżki globalnie
        for p in paths:
            sp = str(p)
            if sp not in seen_paths:
                seen_paths.add(sp)
                all_csv_paths.append(p)

        # Wyświetl numer z nazwą (jeśli dostępna)
        # Najpierw spróbuj z nazwy pliku CSV, potem z BOM
        nazwa = csv_names.get(num) or name_by_num.get(num, "")
        full_name = f"{num} {nazwa}" if nazwa else num
        num_summary_lines.append(f"• {full_name} ({len(paths)} plik(i/ów))")

    if not all_csv_paths:
        return errors

    opis = (
        f"Znaleziono pliki CSV (BOM), które NIE występują w żadnym nadrzędnym BOM-ie.\n\n"
        f"Osierocony BOM to plik zestawienia dla podzespołu, który nie jest używany w projekcie.\n\n"
        f"Numery BOM:\n" + "\n".join(num_summary_lines) + "\n\n"
        f"Możliwe przyczyny:\n"
        f"1) Podzespół został usunięty z głównego zespołu, ale CSV pozostał\n"
        f"2) Literówka w numerze (podzespół jest w BOM pod inną nazwą)\n"
        f"3) Stara wersja zestawienia nie usunięta z projektu\n"
        f"4) Podzespół zapomniany - powinien być dodany do nadrzędnego zespołu\n\n"
        f"Rozwiązanie:\n"
        f"- Jeśli podzespół jest używany: dodaj go do nadrzędnego BOM-u lub popraw numer\n"
        f"- Jeśli podzespół nie jest potrzebny: usuń plik CSV i powiązane pliki (IDW, DWF, itp.)"
    )

    errors.append([
        "PLIKI – Osierocony BOM",
        "(zbiorczo)",
        opis,
        format_file_list(all_csv_paths),
        "",
    ])
    
    return errors


def write_errors_sheet(ws, errors: List[List[str]]) -> None:
    # Format raportu BŁĘDY:
    # Nr | Typ błędu | Nazwa pliku | Opis | Lista plików | Katalog
    headers = ["Nr", "Typ błędu", "Nazwa pliku", "Opis", "Lista plików", "Katalog", "Biblioteka DWF"]
    bold_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Szerokości kolumn (czytelne w Excelu; podgląd WWW i tak potrafi to „rozjechać”)
    col_widths = {
        "A": 5,
        "B": 38,   # Typ błędu
        "C": 36,   # Nazwa pliku
        "D": 72,   # Opis
        "E": 72,   # Lista plików
        "F": 10,   # Katalog
        "G": 85,   # Biblioteka DWF
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    if errors:
        for idx, row_vals in enumerate(errors, start=1):
            # oczekiwany format: [typ, nazwa, opis, lista_plikow, katalog]
            typ = row_vals[0] if len(row_vals) > 0 else ""
            name = row_vals[1] if len(row_vals) > 1 else ""
            opis = row_vals[2] if len(row_vals) > 2 else ""
            lista = row_vals[3] if len(row_vals) > 3 else ""
            katalog = row_vals[4] if len(row_vals) > 4 else ""

            # --- DOKŁADKA: w BŁĘDY w kolumnie "Nazwa pliku" pokazujemy OBIEKT błędu (brakujący plik),
            # a nie kontekst BOM. Nie zmienia to logiki wykrywania błędów – wyłącznie prezentację raportu.
            display_name = name
            try:
                if not str(display_name).strip():
                    typ_l = (typ or "").lower()
                    if ("brak dwf" in typ_l) or ("brak dxf" in typ_l):
                        mm_any = re.search(r"(?i)brak\s+pliku\s+([^\s]+)\.(dwf|dxf|stp|step|stl)", str(opis))
                        if mm_any:
                            display_name = f"{mm_any.group(1)}.{mm_any.group(2)}"
            except Exception:
                display_name = name

            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=typ)
            ws.cell(row=idx + 1, column=3, value=display_name)
            ws.cell(row=idx + 1, column=4, value=opis)
            ws.cell(row=idx + 1, column=5, value=lista)
            ws.cell(row=idx + 1, column=6, value=katalog)
            # --- DOKŁADKA: Biblioteka DWF (TYLKO informacyjnie; nie wpływa na logikę błędów) ---
            # Zasada: jeśli w PROJEKCIE nie ma DWF dla numeru, ale jest w BIBLIOTECE,
            # to w tej kolumnie pokazujemy ścieżki (newest-first).
            lib_info = ""
            lib_folder_uri = None
            try:
                proj_map = DWF_MAP_PROJECT_FOR_REPORT or {}
                lib_map = DWF_MAP_LIBRARY_FOR_REPORT or {}

                # wyciągamy potencjalny numer rysunku z:
                # 1) Nazwa pliku (pierwszy token)
                # 2) Opis (wzorzec "brak pliku XXX.dwf")
                candidates = []

                if name:
                    token = str(name).strip().split(" ", 1)[0]
                    token_n = normalize_drawing_for_match(token)
                    if token_n:
                        candidates.append(token_n)

                if opis:
                    mm = re.search(r"(?i)brak\s+pliku\s+([^\s]+)\.dwf", str(opis))
                    if mm:
                        cand = normalize_drawing_for_match(mm.group(1))
                        if cand:
                            candidates.append(cand)

                # unikalne, w kolejności
                seen = set()
                uniq = []
                for c in candidates:
                    if c not in seen:
                        seen.add(c)
                        uniq.append(c)

                # wybieramy pierwszy kandydat, który faktycznie ma DWF w bibliotece
                chosen = None
                for c in uniq:
                    if (c not in proj_map) and (c in lib_map) and lib_map.get(c):
                        chosen = c
                        break

                if chosen:
                    paths = sort_paths_newest_first(lib_map.get(chosen, []))
                    if paths:
                        lib_info = "DWF znaleziony w bibliotece (najnowszy na górze):\n" + format_file_list(paths)
                        try:
                            lib_folder_uri = hyperlink_to_select_file(paths[0])
                        except Exception:
                            lib_folder_uri = None
            except Exception:
                lib_info = ""

            ws.cell(row=idx + 1, column=7, value=lib_info)
            # Dokładka: jeśli DWF jest w bibliotece (a nie ma w projekcie) – nazwę pliku linkujemy do folderu z biblioteki i malujemy na niebiesko
            if lib_folder_uri:
                try:
                    name_cell = ws.cell(row=idx + 1, column=3)
                    name_cell.hyperlink = lib_folder_uri
                    f = copy(name_cell.font) if name_cell.font else Font()
                    f.color = "0000FF"
                    f.underline = "single"
                    name_cell.font = f
                except Exception:
                    pass

    # Format: pełny opis/typ błędu widoczny w komórce (wrap + auto wysokość wiersza)
    # Uwaga: nie zmienia danych, tylko format wyświetlania.
    wrap_top = Alignment(wrap_text=True, vertical="top")

    # Wrap dla: Typ błędu, Opis, Lista plików
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=2).alignment = wrap_top
        ws.cell(row=r, column=4).alignment = wrap_top
        ws.cell(row=r, column=5).alignment = wrap_top
        ws.cell(row=r, column=7).alignment = wrap_top

    def _est_lines_for_cell(value: object, col_width: float) -> int:
        if value is None:
            return 1
        txt = str(value)
        per_line = max(10, int(col_width) if col_width else 10)

        total = 0
        for part in txt.splitlines() if "\n" in txt else [txt]:
            part = str(part)
            # ile linii zajmie jeden fragment
            if not part:
                total += 1
            else:
                total += max(1, (len(part) // per_line) + (1 if (len(part) % per_line) else 0))
        return max(1, total)

    # Auto-wysokość: bierzemy maksimum z (Typ błędu, Opis, Lista plików)
    width_b = float(col_widths["B"])
    width_d = float(col_widths["D"])
    width_e = float(col_widths["E"])
    width_g = float(col_widths["G"])

    for r in range(2, max_row + 1):
        cb = ws.cell(row=r, column=2)
        cd = ws.cell(row=r, column=4)
        ce = ws.cell(row=r, column=5)

        lines_b = _est_lines_for_cell(cb.value, width_b)
        lines_d = _est_lines_for_cell(cd.value, width_d)
        lines_e = _est_lines_for_cell(ce.value, width_e)
        cg = ws.cell(row=r, column=7)
        lines_g = _est_lines_for_cell(cg.value, width_g)
        lines = max(lines_b, lines_d, lines_e, lines_g)

        ws.row_dimensions[r].height = max(18, min(420, 15 * lines + 6))


# --------------------------------
# ZAPIS ARKUSZY TYPOWYCH
# --------------------------------

# --------------------------------

def write_sectioned_sheet(ws, df: pd.DataFrame, all_sources: List[str], bom_total_by_source: Dict[str, float] = None, orphan_sources: set = None) -> None:
    font_header = Font(bold=True, size=12)
    font_colheader = Font(bold=True)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    current_row = 1
    all_sources = [s for s in all_sources if s is not None]

    df_by_source: Dict[str, pd.DataFrame] = {}
    if not df.empty and SOURCE_COL in df.columns:
        for src, g in df.groupby(SOURCE_COL, sort=False):
            df_by_source[str(src)] = g

    for source_name in all_sources:
        g = df_by_source.get(source_name)
        if g is None or g.empty:
            continue

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(TARGET_COLUMNS),
        )

        # --- Nagłówek BOM: "{BOM} - ilość całkowita Xszt" (ilość BOM-u w projekcie z DRZEWKA) ---
        qty = 0.0
        try:
            if bom_total_by_source:
                qty = float(bom_total_by_source.get(str(source_name), 0.0) or 0.0)
        except Exception:
            qty = 0.0

        qty_str = format_qty(qty)
        title = f"{str(source_name)} - ilość całkowita {qty_str}szt"

        # Ostrzeżenie dla OSIEROCONY BOM (ilość całkowita 0)
        is_orphan = False
        try:
            if orphan_sources and str(source_name) in orphan_sources:
                is_orphan = True
        except Exception:
            is_orphan = False

        if is_orphan or (qty_str == "0"):
            title = f"{str(source_name)} - ilość całkowita 0szt !!! OSIEROCONY BOM !!!"
            red_fill = PatternFill(fill_type="solid", fgColor="FF0000")
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = red_fill
        else:
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = font_header
            cell.fill = yellow_fill

        current_row += 2

        for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=col_name)
            c.font = font_colheader
        current_row += 1

        for _, row in g.iterrows():
            for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
                ws.cell(row=current_row, column=col_idx, value=row.get(col_name, ""))
            current_row += 1

        current_row += 1

    _set_column_widths(ws)


def write_bom_with_sections(
    ws,
    df_full: pd.DataFrame,
    bom_total_by_source: Dict[str, float] | None = None,
    orphan_sources: set[str] | None = None,
) -> None:
    """
    Arkusz PEŁNA TABELA (BOM) – pełny BOM + podtabele (MODUŁY/STANDARD/X/XX/NORM)
    z nową kolumną "Ilość całkowita".
    """
    if df_full.empty:
        return

    font_bom_header = Font(bold=True, size=12)
    font_section_header = Font(bold=True)
    font_colheader = Font(bold=True)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    current_row = 1
    groups = df_full.groupby(SOURCE_COL, sort=False)

    for source_name, df_src in groups:
        if source_name is None:
            source_name = ""

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(TARGET_COLUMNS),
        )
        # --- Nagłówek BOM: "{BOM} - ilość całkowita Xszt" (X = ilość BOM-u w projekcie z DRZEWKA) ---
        qty = 0.0
        try:
            if bom_total_by_source:
                qty = float(bom_total_by_source.get(str(source_name), 0.0) or 0.0)
        except Exception:
            qty = 0.0

        qty_str = format_qty(qty)
        title = f"{str(source_name)} - ilość całkowita {qty_str}szt"

        # OSIEROCONY BOM: czerwony pasek + dopisek
        is_orphan = False
        try:
            if orphan_sources and str(source_name) in orphan_sources:
                is_orphan = True
        except Exception:
            is_orphan = False

        if is_orphan or (qty_str == "0"):
            title = f"{str(source_name)} - ilość całkowita 0szt !!! OSIEROCONY BOM !!!"
            red_fill = PatternFill(fill_type="solid", fgColor="FF0000")
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = red_fill
        else:
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = font_bom_header
            cell.fill = yellow_fill
        current_row += 2

        for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=col_name)
            c.font = font_colheader
        current_row += 1

        df_all = df_src[TARGET_COLUMNS]
        for _, row in df_all.iterrows():
            for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
                ws.cell(row=current_row, column=col_idx, value=row.get(col_name, ""))
            current_row += 1

        current_row += 1

        sections_src = split_sections(df_src)
        section_defs = [
            ("MODULES", "ELEMENTY MODUŁY (ZZ)"),
            ("STANDARD", "ELEMENTY STANDARD"),
            ("X", "ELEMENTY DO CIĘCIA (X)"),
            ("XX", "ELEMENTY DO CIĘCIA I GIĘCIA (XX)"),
            ("NORM", "ELEMENTY ZNORMALIZOWANE"),
        ]

        for key, title in section_defs:
            df_sec = sections_src[key]
            df_sec_vals = df_sec[TARGET_COLUMNS]
            if df_sec_vals.empty:
                continue

            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(TARGET_COLUMNS),
            )
            c = ws.cell(row=current_row, column=1, value=title)
            c.font = font_section_header
            current_row += 1

            for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
                col_cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                col_cell.font = font_colheader
            current_row += 1

            for _, row in df_sec_vals.iterrows():
                for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
                    ws.cell(row=current_row, column=col_idx, value=row.get(col_name, ""))
                current_row += 1

            current_row += 1

        current_row += 1

    _set_column_widths(ws)


# --------------------------------
# BUDOWANIE HIERARCHII (DRZEWKO)
# --------------------------------

def build_hierarchy(df_full: pd.DataFrame) -> Dict[str, Any]:
    """
    Buduje strukturę:
      - edges: relacje rodzic -> [dziecko, ilość]
      - labels: etykiety "Nr Nazwa"
      - types: typy (MODUŁ/STD/X/XX/Z/ZNORMALIZOWANE)
      - roots: złożenia nadrzędne

    Uwaga:
    - w DRZEWKU ignorujemy wiersze BOM, w których Nr rysunku = numer złożenia (samoodwołanie),
      żeby nie powstawały fejkowe gałęzie typu A -> A.
    """
    edges: Dict[str, List[Dict[str, str]]] = {}
    labels: Dict[str, str] = {}
    types: Dict[str, str] = {}
    parent_for_num: Dict[str, str] = {}

    groups = df_full.groupby(SOURCE_COL, sort=False)

    for source_name, df_src in groups:
        if source_name is None:
            continue
        source_name = str(source_name)
        parent_num = normalize_drawing_for_match(extract_number_from_source_name(source_name))
        parent_for_num[parent_num] = source_name

        labels.setdefault(parent_num, source_name)
        types.setdefault(parent_num, classify_type(parent_num))

        for _, row in df_src.iterrows():
            child_num = normalize_drawing_for_match(row.get("Nr rysunku", ""))
            if not child_num:
                continue

            # KLUCZOWE: pomijamy samoodwołania
            if child_num == parent_num:
                continue

            child_name = str(row.get("Nazwa", "")).strip()
            qty = str(row.get("Ilość", "")).strip()

            edges.setdefault(parent_num, []).append({
                "num": child_num,
                "name": child_name,
                "qty": qty,
            })

            if child_num not in labels:
                label = child_num
                if child_name:
                    label = f"{child_num} {child_name}"
                labels[child_num] = label

            if child_num not in types:
                types[child_num] = classify_type(child_num)

    parent_nums = set(parent_for_num.keys())
    child_nums = set()
    for childs in edges.values():
        for ch in childs:
            child_nums.add(ch["num"])

    # ROOTY projektu:
    # - docelowo rootami są MODUŁY (ZZ) (bo to "produkcja/projekt"), a nie każdy BOM bez rodzica.
    # - dzięki temu BOM-y typu ...Z, które nie są nigdzie podpięte, mogą zostać OSIEROCONE (ilość całkowita = 0).
    zz_parents = [num for num in parent_nums if str(num).upper().endswith("ZZ")]

    # preferujemy ZZ, które nie są dzieckiem innych BOM-ów
    roots = [num for num in zz_parents if num not in child_nums]

    # jeśli wszystkie ZZ są gdzieś podpięte – bierzemy wszystkie ZZ jako punkty startowe
    if not roots:
        roots = sorted(zz_parents)

    # fallback (żeby nie zabić małych projektów bez ZZ):
    if not roots:
        roots = [num for num in parent_nums if num not in child_nums]
        if not roots:
            roots = sorted(parent_nums)

    return {
        "roots": roots,
        "edges": edges,
        "labels": labels,
        "types": types,
        "parent_for_num": parent_for_num,
    }


def collect_subtree_numbers(root_num: str, edges: Dict[str, List[Dict[str, str]]]) -> set:
    """
    Zbiera wszystkie numery należące do poddrzewa danego root-a (włącznie z root-em).
    
    Args:
        root_num: Numer głównego modułu (ROOT)
        edges: Słownik krawędzi hierarchii
        
    Returns:
        set: Zbiór wszystkich numerów rysunków w poddrzewie
    """
    subtree_nums = {root_num}
    visited = set()
    
    def dfs(num: str):
        if num in visited:
            return
        visited.add(num)
        
        children = edges.get(num, [])
        for ch in children:
            child_num = ch["num"]
            subtree_nums.add(child_num)
            dfs(child_num)
    
    dfs(root_num)
    return subtree_nums


def filter_to_subtree(df_full: pd.DataFrame, hierarchy: Dict[str, Any], selected_root: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Filtruje DataFrame i hierarchię do tylko wybranego poddrzewa.
    
    Args:
        df_full: Pełny DataFrame z BOM
        hierarchy: Pełna hierarchia
        selected_root: Wybrany ROOT do filtrowania
        
    Returns:
        Tuple: (przefiltrowany DataFrame, zaktualizowana hierarchia)
    """
    edges = hierarchy["edges"]
    labels = hierarchy["labels"]
    types = hierarchy["types"]
    parent_for_num = hierarchy["parent_for_num"]
    
    # Zbierz wszystkie numery z poddrzewa
    subtree_nums = collect_subtree_numbers(selected_root, edges)
    
    # Przefiltruj DataFrame - zachowaj tylko wiersze należące do poddrzewa
    # Dokumenty BOM (SOURCE_COL) mapują się na numery poprzez parent_for_num
    sources_in_subtree = set()
    for num in subtree_nums:
        if num in parent_for_num:
            sources_in_subtree.add(parent_for_num[num])
    
    # Filtruj po źródle BOM (potrzebujemy całego BOM jeśli jego numer jest w poddrzewie)
    df_filtered = df_full[df_full[SOURCE_COL].isin(sources_in_subtree)].copy()
    
    # Przefiltruj hierarchię
    filtered_edges = {}
    for parent, children in edges.items():
        if parent in subtree_nums:
            # Zachowaj tylko dzieci które są w poddrzewie
            filtered_children = [ch for ch in children if ch["num"] in subtree_nums]
            if filtered_children:
                filtered_edges[parent] = filtered_children
    
    filtered_hierarchy = {
        "roots": [selected_root],  # Tylko wybrany root
        "edges": filtered_edges,
        "labels": {k: v for k, v in labels.items() if k in subtree_nums},
        "types": {k: v for k, v in types.items() if k in subtree_nums},
        "parent_for_num": {k: v for k, v in parent_for_num.items() if k in subtree_nums},
    }
    
    return df_filtered, filtered_hierarchy


def filter_files_to_subtree(files: Dict[str, Any], active_nums: set[str]) -> Dict[str, Any]:
    """
    Filtruje strukturę plików do tylko tych, które dotyczą numerów z active_nums.
    
    Args:
        files: Struktura plików ze skanowania projektu
        active_nums: Zbiór numerów rysunków w aktywnym poddrzewie (znormalizowane)
        
    Returns:
        Dict: Przefiltrowana struktura plików
    """
    def extract_num_from_filename(filename: str) -> str:
        """
        Wyciąga numer rysunku z nazwy pliku i normalizuje go.
        Przykłady:
        - "2556-100.29Z Uchwyt 1000W" -> "2556-100.29Z"
        - "NS-100.32 Ośka napinacza" -> "NS-100.32"
        
        UWAGA: filename jest już stem (bez rozszerzenia), więc NIE używamy Path().stem ponownie
        bo to by usunęło sufiks Z/ZZ (np. "2609-100.19Z" -> "2609-100.19")
        """
        # Weź część przed pierwszą spacją (lub całość jeśli nie ma spacji)
        base = filename.split(" ", 1)[0] if " " in filename else filename
        # filename jest już stem, więc tylko normalizujemy
        return normalize_drawing_for_match(base)
    
    def file_num_in_active(file_path: Path) -> bool:
        """Sprawdza czy numer z nazwy pliku jest w active_nums"""
        filename = file_path.stem
        num_normalized = extract_num_from_filename(filename)
        return num_normalized in active_nums if num_normalized else False
    
    # Filtruj wszystkie mapy plików
    filtered_files = {}
    
    # Mapy: numer -> [ścieżki]
    for map_key in ["dxf_map", "dwf_map", "stl_map", "stp_map", "dwf_map_project", "dwf_map_library"]:
        if map_key in files:
            original_map = files[map_key]
            filtered_map = {}
            for num, paths in original_map.items():
                num_normalized = normalize_drawing_for_match(num)
                if num_normalized in active_nums:
                    filtered_map[num] = paths
            filtered_files[map_key] = filtered_map
    
    # CSV files - lista ścieżek
    if "csv_files" in files:
        filtered_files["csv_files"] = [
            p for p in files["csv_files"]
            if file_num_in_active(p)
        ]
    
    # IDW files - lista ścieżek
    if "idw_files" in files:
        filtered_files["idw_files"] = [
            p for p in files["idw_files"]
            if file_num_in_active(p)
        ]
    
    # Inne pola przekopiuj bez zmian
    for key in files:
        if key not in filtered_files:
            filtered_files[key] = files[key]
    
    return filtered_files


def traverse_hierarchy(h: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    DFS po strukturze – buduje listę węzłów (nodes) z ilościami całkowitymi
    oraz listę krawędzi (edge_rows) – ta druga na razie nieużywana w Excelu.
    """
    roots: List[str] = h["roots"]
    edges: Dict[str, List[Dict[str, str]]] = h["edges"]
    labels: Dict[str, str] = h["labels"]
    types: Dict[str, str] = h["types"]

    nodes: List[Dict[str, Any]] = []
    edge_rows: List[Dict[str, Any]] = []

    def dfs(current_num: str,
            root_num: str,
            level: int,
            parent_total_qty: float,
            path_nums: List[str],
            visited: set) -> None:
        children = edges.get(current_num, [])
        for ch in children:
            child_num = ch["num"]
            qty_str = ch["qty"]
            child_label = labels.get(child_num, child_num)

            local_qty = parse_qty(qty_str)
            if parent_total_qty is None or local_qty is None:
                total_qty = None
            else:
                total_qty = parent_total_qty * local_qty

            new_path_nums = path_nums + [child_num]

            nodes.append({
                "root_num": root_num,
                "root_label": labels.get(root_num, root_num),
                "level": level + 1,
                "num": child_num,
                "label": child_label,
                "qty_local_str": qty_str,
                "qty_total": total_qty,
                "type": types.get(child_num, ""),
                "path_nums": new_path_nums,
            })

            parent_label = labels.get(current_num, current_num)
            edge_rows.append({
                "root_num": root_num,
                "root_label": labels.get(root_num, root_num),
                "level": level + 1,
                "parent_num": current_num,
                "parent_label": parent_label,
                "child_num": child_num,
                "child_label": child_label,
                "qty": qty_str,
                "type": types.get(child_num, ""),
                "path_nums": new_path_nums,
            })

            if child_num in visited:
                continue

            visited.add(child_num)
            dfs(child_num, root_num, level + 1, total_qty, new_path_nums, visited)
            visited.remove(child_num)

    for root in roots:
        root_label = labels.get(root, root)

        nodes.append({
            "root_num": root,
            "root_label": root_label,
            "level": 0,
            "num": root,
            "label": root_label,
            "qty_local_str": "",
            "qty_total": None,
            "type": types.get(root, ""),
            "path_nums": [root],
        })

        visited = {root}
        dfs(root, root, 0, 1.0, [root], visited)

    return nodes, edge_rows


# --------------------------------
# DRZEWKO – ARKUSZE
# --------------------------------

def write_tree_ascii_sheet(ws, h: Dict[str, Any]) -> None:
    """
    Arkusz DRZEWKO ASCII – widok jak:
      2556-000.00ZZ Nalewarka 10N
      ├── 2× 2556-200.00ZZ Okapnik
      ...
    """
    roots: List[str] = h["roots"]
    edges: Dict[str, List[Dict[str, str]]] = h["edges"]
    labels: Dict[str, str] = h["labels"]

    row_ref = [1]

    for root in roots:
        visited = {root}
        root_label = labels.get(root, root)

        ws.cell(row=row_ref[0], column=1, value=root_label)
        row_ref[0] += 1

        def dfs(num: str, prefix: str, visited_local: set) -> None:
            children = edges.get(num, [])
            for idx, ch in enumerate(children):
                is_last = (idx == len(children) - 1)
                branch = "└── " if is_last else "├── "
                qty_str = ch["qty"] or "1"
                child_num = ch["num"]
                child_label = labels.get(child_num, child_num)

                line = f"{prefix}{branch}{qty_str}× {child_label}"
                ws.cell(row=row_ref[0], column=1, value=line)
                row_ref[0] += 1

                if child_num in visited_local:
                    cyc_prefix = prefix + ("    " if is_last else "│   ")
                    cyc_line = f"{cyc_prefix}[CYKL -> {child_num}]"
                    ws.cell(row=row_ref[0], column=1, value=cyc_line)
                    row_ref[0] += 1
                    continue

                visited_local.add(child_num)
                next_prefix = prefix + ("    " if is_last else "│   ")
                dfs(child_num, next_prefix, visited_local)
                visited_local.remove(child_num)

        dfs(root, "", visited)

        row_ref[0] += 1

    ws.column_dimensions[get_column_letter(1)].width = 120


def write_tree_text_sheet(ws, nodes: List[Dict[str, Any]]) -> None:
    """
    Arkusz DRZEWKO TEKST – tabelka:
    Root | Poziom | Nr rysunku | Nazwa | Ilość lokalna | Ilość całkowita | Typ | Ścieżka
    """
    headers = [
        "Root",
        "Poziom",
        "Nr rysunku",
        "Nazwa",
        "Ilość lokalna",
        "Ilość całkowita",
        "Typ",
        "Ścieżka",
    ]
    bold_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    row_idx = 2
    for node in nodes:
        root_label = node["root_label"]
        level = node["level"]
        num = node["num"]
        label = node["label"]
        qty_local_str = node["qty_local_str"]
        qty_total = node["qty_total"]
        typ = node["type"]
        path_nums = node["path_nums"]

        nazwa = label
        if label.startswith(num):
            nazwa = label[len(num):].strip()

        path_str = " > ".join(path_nums)
        qty_total_str = format_qty(qty_total)

        ws.cell(row=row_idx, column=1, value=root_label)
        ws.cell(row=row_idx, column=2, value=level)
        ws.cell(row=row_idx, column=3, value=num)
        ws.cell(row=row_idx, column=4, value=nazwa)
        ws.cell(row=row_idx, column=5, value=qty_local_str)
        ws.cell(row=row_idx, column=6, value=qty_total_str)
        ws.cell(row=row_idx, column=7, value=typ)
        ws.cell(row=row_idx, column=8, value=path_str)

        row_idx += 1

    col_widths = {
        1: 30,
        2: 8,
        3: 18,
        4: 40,
        5: 12,
        6: 14,
        7: 20,
        8: 80,
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


# --------------------------------
# ZBIORCZY – LISTA CZĘŚCI Z ILOŚCIĄ CAŁKOWITĄ
# --------------------------------

def compute_total_quantities(nodes: List[Dict[str, Any]],
                             df_full: pd.DataFrame,
                             parent_for_num: Dict[str, str]) -> List[Dict[str, Any]]:
    """
    Wylicza ilości całkowite na projekt dla numerów RMPAK (MODUŁY/STD/X/XX/Z)
    na podstawie DRZEWKA.

    ELEMENTY ZNORMALIZOWANE nie są tu liczone – do ZBIORCZY
    kopiujemy je z BOM (sekcja NORM).
    """
    meta_map: Dict[str, Dict[str, str]] = {}
    for _, row in df_full.iterrows():
        num = normalize_drawing_for_match(row.get("Nr rysunku", ""))
        if not num or num in meta_map:
            continue
        meta_map[num] = {
            "name": str(row.get("Nazwa", "")).strip(),
            "opis": str(row.get("Opis", "")).strip(),
            "material": str(row.get("Materiał", "")).strip(),
            "supplier": str(row.get("Dostawca", "")).strip(),
            "pliki_3d": str(row.get("Pliki 3D", "")).strip()
                        if "Pliki 3D" in df_full.columns else "",
            "katalog": str(row.get("Katalog", "")).strip()
                        if "Katalog" in df_full.columns else "",
        }


    # --- Uzupełnianie metadanych dla ROOT (BOM-ów), których nie ma jako wierszy w df_full ---
    # parent_for_num: {parent_num -> source_name (czyli nazwa CSV/BOM)}
    source_katalog: Dict[str, str] = {}
    if "Katalog" in df_full.columns and SOURCE_COL in df_full.columns:
        for src, g in df_full.groupby(SOURCE_COL, sort=False):
            # bierzemy pierwszą sensowną wartość katalogu z BOM-a
            vals = [str(v).strip() for v in g["Katalog"].tolist() if str(v).strip() not in ("", "nan", "None")]
            if vals:
                source_katalog[str(src)] = vals[0]

    for parent_num, source_name in parent_for_num.items():
        parent_num = str(parent_num).strip()
        if not parent_num:
            continue
        if parent_num not in meta_map:
            meta_map[parent_num] = {
                "name": str(source_name).replace(parent_num, "", 1).strip(),
                "opis": "",
                "material": "",
                "supplier": "",
                "pliki_3d": "",
                "katalog": source_katalog.get(str(source_name), ""),
            }
        else:
            # jeśli istnieje, ale katalog pusty – też uzupełnij z źródła
            if not str(meta_map[parent_num].get("katalog", "")).strip():
                meta_map[parent_num]["katalog"] = source_katalog.get(str(source_name), "")

    totals: Dict[str, Dict[str, Any]] = {}

    for node in nodes:
        num = str(node["num"]).strip()
        if not num:
            continue

        # bierzemy tylko numery RMPAK
        if not is_rmpak_number(num):
            continue

        label = node["label"]
        level = node["level"]
        typ = classify_type(num) or node.get("type", "")
        qty_total = node["qty_total"]

        # rooty = 1 szt.
        if level == 0:
            eff_total = 1.0
        else:
            eff_total = qty_total

        if eff_total is None:
            continue

        if num not in totals:
            meta = meta_map.get(num, {})
            if meta.get("name"):
                name = meta["name"]
            else:
                lbl = label or num
                name = lbl
                if lbl.startswith(num):
                    name = lbl[len(num):].strip()

            totals[num] = {
                "num": num,
                "name": name,
                "opis": meta.get("opis", ""),
                "type": typ,
                "total": 0.0,
                "material": meta.get("material", ""),
                "supplier": meta.get("supplier", ""),
                "pliki_3d": meta.get("pliki_3d", ""),
                "katalog": meta.get("katalog", ""),
                "status": "",
            }

        totals[num]["total"] += eff_total or 0.0

    return list(totals.values())


def write_zbiorczy_summary_sheet(ws,
                                 totals: List[Dict[str, Any]],
                                 df_norm: pd.DataFrame) -> None:
    """
    Arkusz ZBIORCZY:

    ELEMENTY MODUŁY (ZZ)    – z DRZEWKA (totals)
    ELEMENTY STANDARD       – z DRZEWKA (totals, w tym ZŁOŻENIE (Z))
    ELEMENTY DO CIĘCIA (X)  – z DRZEWKA (totals)
    ELEMENTY DO CIĘCIA I GIĘCIA (XX) – z DRZEWKA (totals)
    ELEMENTY ZNORMALIZOWANE – kopiowane z BOM (df_norm),
                              tak samo jak w arkuszu ELEMENTY ZNORMALIZOWANE.
    """
    modules = []
    standard = []
    cut_x = []
    cut_xx = []

    for item in totals:
        typ = item["type"]
        if typ == "MODUŁ (ZZ)":
            modules.append(item)
        elif typ in ("STANDARD", "ZŁOŻENIE (Z)"):
            standard.append(item)
        elif typ == "CIĘCIE (X)":
            cut_x.append(item)
        elif typ == "CIĘCIE+GIĘCIE (XX)":
            cut_xx.append(item)

    key_fn = lambda x: x["num"]
    modules.sort(key=key_fn)
    standard.sort(key=key_fn)
    cut_x.sort(key=key_fn)
    cut_xx.sort(key=key_fn)

    # ELEMENTY ZNORMALIZOWANE – prosto z BOM (sekcja NORM)
    # UWAGA: w ZBIORCZY chcemy ogarniać duplikaty "znormalizowanych".
    # Reguła: jeśli (Nazwa + Opis) są identyczne -> traktujemy jako jedną pozycję i sumujemy ilości.
    # To jest TYLKO prezentacja w ZBIORCZY (nie zmienia df_norm ani innych arkuszy).
    #
    # Dodatkowo (tylko w tej podtabeli):
    # - Katalog: bierzemy z BOM-u, jeśli był,
    # - jeśli pusty -> fallback z numeru BOM-u źródłowego (SOURCE), np. 2556-900.11ZZ -> 900,
    # - jeśli pozycja zsumowana z wielu BOM-ów -> wpisujemy listę katalogów: 900,950,980.
    norm: List[Dict[str, Any]] = []
    if df_norm is not None and not df_norm.empty:
        agg: Dict[Tuple[str, str], Dict[str, Any]] = {}

        for _, row in df_norm.iterrows():
            src = str(row.get(SOURCE_COL, "")).strip()
            num = str(row.get("Nr rysunku", "")).strip()
            name = str(row.get("Nazwa", "")).strip()
            opis = str(row.get("Opis", "")).strip()

            qty_str = str(row.get("Ilość całkowita", "") or row.get("Ilość", "")).strip()
            qty = parse_qty(qty_str)
            qty_val = qty if qty is not None else 0.0

            material = str(row.get("Materiał", "")).strip()
            supplier = str(row.get("Dostawca", "")).strip()
            pliki_3d = str(row.get("Pliki 3D", "")).strip() if "Pliki 3D" in df_norm.columns else ""
            katalog_raw = str(row.get("Katalog", "")).strip() if "Katalog" in df_norm.columns else ""

            # Katalog:
            # 1) z BOM-u (jeśli był),
            # 2) fallback z numeru BOM-u źródłowego (SOURCE) -> 900/950/...
            katalog_eff = katalog_raw
            if not katalog_eff:
                bn = extract_number_from_source_name(src)
                katalog_eff = katalog_from_rmpak_number(bn)

            k = (name, opis)

            if k not in agg:
                agg[k] = {
                    "num": num,
                    "name": name,
                    "opis": opis,  # nie pokazujemy w ZBIORCZY, ale używamy do klucza
                    "total": float(qty_val),
                    "type": "ZNORMALIZOWANE",
                    "material": material,
                    "supplier": supplier,
                    "pliki_3d": pliki_3d,
                    "katalog": "",  # zostanie zbudowany z katalog_set
                    "katalog_set": set([katalog_eff]) if katalog_eff else set(),
                    "status": "",
                }
            else:
                agg[k]["total"] = float(agg[k].get("total") or 0.0) + float(qty_val)

                # Jeśli numery różne, to nie ryzykujemy – zostawiamy puste.
                prev_num = str(agg[k].get("num", "")).strip()
                if prev_num and num and normalize_drawing_for_match(prev_num) != normalize_drawing_for_match(num):
                    agg[k]["num"] = ""
                elif not prev_num and num:
                    agg[k]["num"] = num

                # Dopnij brakujące pola (pierwsza niepusta wygrywa)
                if not str(agg[k].get("material", "")).strip() and material:
                    agg[k]["material"] = material
                if not str(agg[k].get("supplier", "")).strip() and supplier:
                    agg[k]["supplier"] = supplier
                if not str(agg[k].get("pliki_3d", "")).strip() and pliki_3d:
                    agg[k]["pliki_3d"] = pliki_3d

                # katalog: zbieramy wszystkie katalogi z wielu BOM-ów
                if katalog_eff:
                    if "katalog_set" not in agg[k] or not isinstance(agg[k].get("katalog_set"), set):
                        agg[k]["katalog_set"] = set()
                    agg[k]["katalog_set"].add(katalog_eff)

        # Składamy katalog jako listę: 900,950,980 (rosnąco)
        for _k in list(agg.keys()):
            ks = agg[_k].get("katalog_set")
            if isinstance(ks, set) and ks:
                try:
                    katalog_sorted = sorted(list(ks), key=lambda x: int(x) if str(x).isdigit() else str(x))
                except Exception:
                    katalog_sorted = sorted([str(x) for x in ks])
                agg[_k]["katalog"] = ",".join([str(x) for x in katalog_sorted])
            else:
                agg[_k]["katalog"] = ""

            # Status tylko dla tej podtabeli: jeśli suma = 0 -> OSIEROCONY
            try:
                _t = float(agg[_k].get("total") or 0.0)
            except Exception:
                _t = 0.0
            agg[_k]["status"] = "OSIEROCONY" if abs(_t) < 1e-9 else ""


        # stabilna kolejność: Nazwa, Opis
        for (_name, _opis) in sorted(agg.keys(), key=lambda x: (x[0], x[1])):
            norm.append(agg[(_name, _opis)])

    sections = [
        ("ELEMENTY MODUŁY (ZZ)", modules),
        ("ELEMENTY STANDARD", standard),
        ("ELEMENTY DO CIĘCIA (X)", cut_x),
        ("ELEMENTY DO CIĘCIA I GIĘCIA (XX)", cut_xx),
        ("ELEMENTY ZNORMALIZOWANE", norm),
    ]

    font_section = Font(bold=True, size=12)
    font_colheader = Font(bold=True)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    col_headers = [
        "Nr rysunku",
        "Nazwa",
        "Opis",
        "Ilość całkowita",
        "Typ",
        "Materiał",
        "Dostawca",
        "Pliki 3D",
        "Katalog",
        "Status",
    ]

    current_row = 1

    for title, rows in sections:
        # zawsze nagłówek sekcji – nawet przy pustej liście
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(col_headers),
        )
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = font_section
        cell.fill = yellow_fill
        current_row += 1

        # nagłówki kolumn
        for col_idx, header in enumerate(col_headers, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=header)
            c.font = font_colheader
        current_row += 1

        # dane – jeśli są wiersze
        for item in rows:
            num = item["num"]
            name = item["name"]
            opis = item.get("opis", "")
            total = item.get("total")
            total_str = format_qty(total)
            typ = item["type"]
            material = item["material"]
            supplier = item["supplier"]
            pliki_3d = item["pliki_3d"]
            katalog = item["katalog"]

            values = [
                num,
                name,
                opis,
                total_str,
                typ,
                material,
                supplier,
                pliki_3d,
                katalog,
                item.get("status", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 1

    col_widths = {
        1: 18,   # Nr rysunku
        2: 40,   # Nazwa
        3: 45,   # Opis
        4: 14,   # Ilość całkowita
        5: 18,   # Typ
        6: 18,   # Materiał
        7: 18,   # Dostawca
        8: 12,   # Pliki 3D
        9: 12,   # Katalog
        10: 14,  # Status
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


# --------------------------------
# EKSPORT DO EXCELA
# --------------------------------

def export_to_excel(sections: Dict[str, pd.DataFrame],
                    output_path: Path,
                    df_full: pd.DataFrame,
                    errors: List[List[str]],
                    hierarchy: Dict[str, Any],
                    nodes: List[Dict[str, Any]],
                    totals: List[Dict[str, Any]],
                    bom_total_by_source: Dict[str, float] | None = None,
                    orphan_sources: set[str] | None = None) -> None:
    log(f"Zapisuję Excel: {output_path}")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    all_sources = list(df_full[SOURCE_COL].astype(str).unique())

    # PEŁNA TABELA (BOM)
    ws_full = wb.create_sheet(SHEET_FULL)
    write_bom_with_sections(ws_full, df_full, bom_total_by_source=bom_total_by_source, orphan_sources=orphan_sources)

    # Sekcje osobne
    sheet_order = [
        (SHEET_MODULES, "MODULES"),
        (SHEET_STANDARD, "STANDARD"),
        (SHEET_X, "X"),
        (SHEET_XX, "XX"),
        (SHEET_NORM, "NORM"),
    ]

    for sheet_name, key in sheet_order:
        ws = wb.create_sheet(sheet_name)
        df_section = sections.get(key)
        # DOPINAMY ROOT-MODUŁY (ZZ) do arkusza 'ELEMENTY MODUŁY (ZZ)'
        # (rooty są widoczne w ZBIORCZY; tu chcemy je też mieć jako lista modułów)
        if key == "MODULES":
            root_rows: List[Dict[str, Any]] = []
            roots = hierarchy.get("roots", [])
            parent_for_num = hierarchy.get("parent_for_num", {})
            # source_katalog z compute_total_quantities jest lokalne, więc bierzemy katalog z df_full dla danego SOURCE
            source_katalog: Dict[str, str] = {}
            if "Katalog" in df_full.columns and SOURCE_COL in df_full.columns:
                for src, g in df_full.groupby(SOURCE_COL, sort=False):
                    vals = [str(v).strip() for v in g["Katalog"].tolist() if str(v).strip() not in ("", "nan", "None")]
                    if vals:
                        source_katalog[str(src)] = vals[0]

            for root_num in roots:
                root_num = str(root_num).strip()
                if not root_num.endswith("ZZ"):
                    continue
                source_name = parent_for_num.get(root_num, root_num)
                katalog = source_katalog.get(str(source_name), "")
                name = str(source_name).replace(root_num, "", 1).strip()
                root_rows.append({
                    SOURCE_COL: str(source_name),
                    "Poz.": "",
                    "Nr rysunku": root_num,
                    "Nazwa": name,
                    "Opis": "",
                    "Ilość": "1",
                    "Materiał": "",
                    "Dostawca": "",
                    "Pliki 3D": "",
                    "Katalog": katalog,
                    "Ilość całkowita": "1",
                    "Typ": "MODUŁ (ZZ)",
                })

            if root_rows:
                df_roots = pd.DataFrame(root_rows)
                # dopilnuj kolumn
                for col in [SOURCE_COL] + TARGET_COLUMNS:
                    if col not in df_roots.columns:
                        df_roots[col] = ""
                df_roots = df_roots[[SOURCE_COL] + TARGET_COLUMNS]
                if df_section is None or df_section.empty:
                    df_section = df_roots
                else:
                    df_section = pd.concat([df_section, df_roots], ignore_index=True)

        if df_section is None:
            df_section = pd.DataFrame(columns=[SOURCE_COL] + TARGET_COLUMNS)
        write_sectioned_sheet(ws, df_section, all_sources, bom_total_by_source=bom_total_by_source, orphan_sources=(orphan_sources or set()))

    # ZBIORCZY (sumy ilości całkowitych + znormalizowane z BOM)
    ws_combined = wb.create_sheet(SHEET_COMBINED)
    df_norm = sections.get("NORM", pd.DataFrame(columns=[SOURCE_COL] + TARGET_COLUMNS))
    write_zbiorczy_summary_sheet(ws_combined, totals, df_norm)
    # Hyperlinki w ZBIORCZY: DWF z projektu / biblioteki
    apply_dwf_hyperlinks_to_zbiorczy(ws_combined)

    # DRZEWKO
    ws_tree_ascii = wb.create_sheet(SHEET_TREE_ASCII)
    write_tree_ascii_sheet(ws_tree_ascii, hierarchy)

    ws_tree_text = wb.create_sheet(SHEET_TREE_TEXT)
    write_tree_text_sheet(ws_tree_text, nodes)

    # BŁĘDY
    ws_errors = wb.create_sheet(SHEET_ERRORS)
    write_errors_sheet(ws_errors, errors)

    # --- DOKŁADKA: oznacz numery rysunków, które mają DWF tylko w bibliotece ---
    for _ws in wb.worksheets:
        if _ws.title == SHEET_ERRORS:
            continue
        apply_library_links_to_ws(_ws)

    # Zapis pliku z obsługą błędu "plik otwarty"
    try:
        wb.save(output_path)
    except PermissionError as e:
        raise PermissionError(
            f"\n{'='*80}\n"
            f"BŁĄD ZAPISU PLIKU\n"
            f"{'='*80}\n\n"
            f"Nie można zapisać pliku:\n{output_path}\n\n"
            f"PRZYCZYNA:\n"
            f"Plik jest prawdopodobnie OTWARTY w programie Excel lub innej aplikacji.\n\n"
            f"ROZWIĄZANIE:\n"
            f"1. ZAMKNIJ plik '{output_path.name}' w Excelu\n"
            f"2. Uruchom przetwarzanie ponownie\n\n"
            f"Jeśli plik nie jest otwarty, sprawdź:\n"
            f"- Czy masz uprawnienia do zapisu w tym katalogu\n"
            f"- Czy plik nie jest zabezpieczony przed zapisem\n"
            f"- Czy dysk nie jest pełny lub zablokowany\n"
            f"{'='*80}\n"
        ) from None


# --------------------------------
# GŁÓWNA LOGIKA
# --------------------------------

def process_project(project_dir: Path, library_dir: Path = None, use_library: bool = True, selected_root: str = None, cancel_check=None) -> str:
    """
    Przetwarza projekt i generuje plik Excel.
    
    Args:
        project_dir: Ścieżka do folderu projektu
        library_dir: Ścieżka do folderu biblioteki DWF (opcjonalna)
        use_library: Czy skanować bibliotekę DWF
        selected_root: Wybrany ROOT do filtrowania (opcjonalny - jeśli None, przetwarzane są wszystkie)
        cancel_check: Funkcja sprawdzająca anulowanie operacji
    
    Returns:
        str: Nazwa wygenerowanego pliku Excel
    """
    log(f"Projekt: {project_dir}")

    def _check_cancel() -> None:
        if cancel_check and cancel_check():
            raise UserCancelled("Przetwarzanie zostało przerwane przez użytkownika.")

    _check_cancel()
    files = scan_project_files(project_dir, cancel_check=_check_cancel)
    _check_cancel()

    # --- DOKŁADKA: Biblioteka DWF (fallback) ---
    dwf_map_library: Dict[str, List[Path]] = {}
    
    # Użyj przekazanej ścieżki biblioteki lub domyślnej
    lib_path = library_dir if library_dir else LIBRARY_DWF_ROOT
    
    log(f"Biblioteka DWF - ścieżka: {lib_path}")
    log(f"Biblioteka DWF - użyj: {use_library}")
    log(f"Biblioteka DWF - istnieje: {lib_path.exists() if lib_path else False}")
    
    if use_library and lib_path and lib_path.exists():
        try:
            log(f"Skanowanie biblioteki DWF: {lib_path}")
            dwf_map_library = scan_library_dwf(lib_path)
            if dwf_map_library:
                log(f"✅ Znaleziono DWF w bibliotece: {sum(len(v) for v in dwf_map_library.values())}")
            else:
                log("⚠️ Biblioteka przeskanowana, ale brak plików DWF")
        except Exception as e:
            log(f"❌ Błąd skanowania biblioteki: {e}")
            dwf_map_library = {}
    else:
        if not use_library:
            log("ℹ️ Skanowanie biblioteki wyłączone przez użytkownika")
        elif not lib_path:
            log("⚠️ Brak ścieżki do biblioteki")
        elif not lib_path.exists():
            log(f"⚠️ Folder biblioteki nie istnieje: {lib_path}")

    files["dwf_map_project"] = files.get("dwf_map", {})
    files["dwf_map_library"] = dwf_map_library

    global DWF_MAP_PROJECT_FOR_REPORT, DWF_MAP_LIBRARY_FOR_REPORT
    DWF_MAP_PROJECT_FOR_REPORT = files.get("dwf_map_project", {}) or {}
    DWF_MAP_LIBRARY_FOR_REPORT = files.get("dwf_map_library", {}) or {}
    csv_path_by_source: Dict[str, Path] = {p.stem: p for p in files.get('csv_files', [])}
    csv_files: List[Path] = files["csv_files"]
    dxf_map: Dict[str, List[Path]] = files["dxf_map"]
    stl_map: Dict[str, List[Path]] = files["stl_map"]
    stp_map: Dict[str, List[Path]] = files["stp_map"]
    dwf_map: Dict[str, List[Path]] = files["dwf_map"]

    if not csv_files:
        log("Brak plików CSV (BOM) w projekcie - nic do zrobienia.")
        return

    all_bom_frames: List[pd.DataFrame] = []
    csv_read_errors: List[tuple] = []  # (csv_path, error_message)

    for csv_path in csv_files:
        _check_cancel()
        try:
            df = read_bom_csv(csv_path)
        except Exception as e:
            error_msg = str(e)
            csv_read_errors.append((csv_path, error_msg))
            log(f"  ❌ BŁĄD czytania CSV {csv_path.name}:")
            log(f"     {error_msg}")
            continue

        df[SOURCE_COL] = csv_path.stem

        df = assign_katalog(df, csv_path, dxf_map, stl_map, stp_map, dwf_map)
        df = update_material_from_dxf(df, dxf_map, dwf_map_project=files.get('dwf_map_project', {}), dwf_map_library=files.get('dwf_map_library', {}))
        df = assign_3d_files(df, stl_map, stp_map)

        all_bom_frames.append(df)
        _check_cancel()

    # Wyświetl podsumowanie błędów
    _check_cancel()
    if csv_read_errors:
        log("\n" + "="*70)
        log("⚠️  BŁĘDY CZYTANIA PLIKÓW CSV:")
        log("="*70)
        for csv_path, error_msg in csv_read_errors:
            log(f"\n❌ {csv_path.name}")
            log(f"   {error_msg}")
        log("\n" + "="*70)
        
        # NIEZALEŻNIE od tego czy inne CSV-y się przeczytały:
        # poinformuj użytkownika o błędach poprzez exception
        error_details = "\n".join([f"❌ {csv_path.name}\n   {error_msg}" 
                                  for csv_path, error_msg in csv_read_errors])
        
        _check_cancel()
        if not all_bom_frames:
            # Krytyczne: żaden CSV się nie przeczytał
            raise ValueError(
                f"❌ BŁĄD KRYTYCZNY: Żaden plik CSV nie został poprawnie wczytany!\n\n"
                f"Szczegóły błędów:\n{error_details}\n\n"
                f"Sprawdź:\n"
                f"• Separatory w CSV (muszą być ;)\n"
                f"• Wymagane kolumny: Poz., Nr rysunku, Nazwa, Opis, Ilość, Materiał, Dostawca\n"
                f"• Kodowanie pliku (UTF-8, cp1250, iso-8859-2)"
            )
        else:
            # Ostrzeżenie: część CSV-ów się przeczytała, ale część ma błędy
            raise ValueError(
                f"⚠️  OSTRZEŻENIE: Część plików CSV ma błędy!\n\n"
                f"Szczegóły błędów:\n{error_details}\n\n"
                f"Przetwarzanie kontynuuj z pozostałymi plikami, ale sprawdź błędy powyżej.\n\n"
                f"Przyczyny mogą to być:\n"
                f"• Separatory w CSV (muszą być ;)\n"
                f"• Wymagane kolumny: Poz., Nr rysunku, Nazwa, Opis, Ilość, Materiał, Dostawca\n"
                f"• Kodowanie pliku (UTF-8, cp1250, iso-8859-2)"
            )

    _check_cancel()
    df_full = pd.concat(all_bom_frames, ignore_index=True)

    if SOURCE_COL not in df_full.columns:
        df_full[SOURCE_COL] = ""
    for col in TARGET_COLUMNS:
        if col not in df_full.columns:
            df_full[col] = ""

    ordered_cols = [SOURCE_COL] + TARGET_COLUMNS
    df_full = df_full[ordered_cols]

    # DRZEWKO – hierarchia + ilości całkowite dla RMPAK
    _check_cancel()
    hierarchy = build_hierarchy(df_full)

    # Zapamiętaj ORYGINALNE top-level rooty PRZED filtrowaniem do poddrzewa,
    # żeby później poprawnie wybrać folder zapisu (root projektu → główny katalog).
    original_top_level_roots = {normalize_drawing_for_match(r) for r in hierarchy.get("roots", []) if r}

    # Zachowaj NIEPRZEFILTROWANE df_full (wszystkie BOM-y w projekcie) – do walidacji
    # "orphaned BOM/files". Bez tego po przefiltrowaniu do poddrzewa rodzicielskie BOM-y
    # spoza poddrzewa znikają i ich dzieci wyglądają na osierocone, mimo że są
    # podpięte gdzie indziej w projekcie. Tryb szybki dla rootu w ogóle nie filtruje
    # df_full – ta zmienna sprawia, że pełny tryb robi to samo dla orphaned.
    df_full_unfiltered = df_full

    # Zapamiętaj wybrany ROOT (dla określenia folderu zapisu później)
    selected_root_normalized = None
    if selected_root:
        selected_root_normalized = normalize_drawing_for_match(selected_root)
        log(f"📌 Filtrowanie do poddrzewa: {selected_root}")
        df_full, hierarchy = filter_to_subtree(df_full, hierarchy, selected_root_normalized)
        log(f"✅ Przefiltrowano do {len(df_full)} wierszy z poddrzewa {selected_root}")
    
    nodes, edge_rows = traverse_hierarchy(hierarchy)  # edge_rows na razie nieużywane
    parent_for_num = hierarchy.get("parent_for_num", {})
    totals = compute_total_quantities(nodes, df_full, parent_for_num)
    
    # Wygeneruj nazwę pliku wyjściowego na podstawie ROOT BOM (numer + nazwa)
    roots = hierarchy.get("roots", [])
    labels = hierarchy.get("labels", {})
    
    if roots:
        root_bom = roots[0]  # Pierwszy root (główny moduł)
        # Pobierz pełną nazwę z labels (numer + nazwa, np. "2556-000.00ZZ Nalewarka 10N")
        root_label = labels.get(root_bom, root_bom)
        # Wyczyść znaki specjalne z nazwy pliku
        safe_name = root_label.replace("/", "-").replace("\\", "-").replace(":", "-")
        output_excel_name = f"{safe_name}_OUT.xlsx"
    else:
        # Fallback - jeśli brak rootów, użyj nazwy katalogu
        output_excel_name = f"{project_dir.name}_OUT.xlsx"
    
    log(f"Nazwa pliku wyjściowego: {output_excel_name}")

    # -------------------------------------------------
    # OSIEROCONY BOM:
    # BOM (CSV) istnieje, ale nie ma żadnej ścieżki z rootów -> suma ścieżek = 0.
    # (wynika z algorytmu, nie z "ifa osierocony")
    # -------------------------------------------------
    # aktywne numery = te, które w DRZEWKU mają ilość całkowitą > 0
    active_nums = set()
    total_numeric_map: Dict[str, float] = {}
    for item in totals:
        n = normalize_drawing_for_match(item["num"])
        total_numeric_map[n] = float(item.get("total") or 0.0)
        if float(item.get("total") or 0.0) > 0.0:
            active_nums.add(n)

    # źródła (CSV/BOM) -> numer BOM
    all_sources = list(df_full[SOURCE_COL].astype(str).unique())
    orphan_sources: List[str] = []
    orphan_boms: List[Tuple[str, str]] = []  # (bom_num_clean, source_name)

    # szybki katalog z BOM-a (pierwsza sensowna wartość) – do opisu i ZBIORCZY
    source_katalog: Dict[str, str] = {}
    if "Katalog" in df_full.columns and SOURCE_COL in df_full.columns:
        for src, g in df_full.groupby(SOURCE_COL, sort=False):
            vals = [str(v).strip() for v in g["Katalog"].tolist() if str(v).strip() not in ("", "nan", "None")]
            if vals:
                source_katalog[str(src)] = vals[0]
            else:
                # fallback: z numeru BOM
                bn = extract_number_from_source_name(str(src))
                source_katalog[str(src)] = katalog_from_rmpak_number(bn)

    for src in all_sources:
        src_str = str(src)
        bom_num = extract_number_from_source_name(src_str)
        bom_num_clean = normalize_drawing_for_match(bom_num)
        if not bom_num_clean or not is_rmpak_number(bom_num_clean):
            continue

        # jeżeli BOM nie ma żadnej ścieżki z rootów -> ilość całkowita = 0
        if bom_num_clean not in active_nums and total_numeric_map.get(bom_num_clean, 0.0) == 0.0:
            # dodatkowo upewniamy się, że to nie jest root (rooty zawsze mają 1.0)
            if bom_num_clean not in hierarchy.get("roots", []):
                orphan_sources.append(src_str)
                orphan_boms.append((bom_num_clean, src_str))

    # dopisz OSIEROCONY BOM jako BŁĄD
    # (nie jest w produkcji, ale ma być widoczny w raporcie)
    # UWAGA: nie robimy "if -> 0" w DRZEWKU; tu tylko raportujemy fakt.
    # Ilości całkowite wyjdą 0 naturalnie, bo brak ścieżek.
    # -------------------------------------------------
    # Dodatkowo: OSIEROCONY BOM ma być widoczny w ZBIORCZY z ilością 0 + Status
    if orphan_boms:
        # Zamiast "doklejać" duplikaty w totals – nadpisujemy istniejące wpisy.
        # (w przeciwnym razie ZBIORCZY może pokazać wersję z ilością >0)
        totals_by_num = {normalize_drawing_for_match(t.get("num", "")): t for t in totals if t.get("num")}
        for bom_num_clean, src_str in orphan_boms:
            name = str(src_str).replace(bom_num_clean, "", 1).strip()
            typ = classify_type(bom_num_clean)
            kat = source_katalog.get(src_str, katalog_from_rmpak_number(bom_num_clean))

            key = normalize_drawing_for_match(bom_num_clean)
            if key in totals_by_num:
                totals_by_num[key]["total"] = 0.0
                totals_by_num[key]["status"] = "OSIEROCONY"
                # dopnij katalog jeśli pusty
                if not str(totals_by_num[key].get("katalog", "")).strip():
                    totals_by_num[key]["katalog"] = kat
                # typ też ujednolić (czasem mógł być policzony inaczej)
                if typ:
                    totals_by_num[key]["type"] = typ
                if name and not str(totals_by_num[key].get("name", "")).strip():
                    totals_by_num[key]["name"] = name
            else:
                totals_by_num[key] = {
                    "num": bom_num_clean,
                    "name": name,
                    "type": typ,
                    "total": 0.0,
                    "material": "",
                    "supplier": "",
                    "pliki_3d": "",
                    "katalog": kat,
                    "status": "OSIEROCONY",
                }

        totals = list(totals_by_num.values())

    # dla wierszy z OSIEROCONY BOM: ustaw "Ilość całkowita" na 0 (dla RMPAK),
    # żeby nie było pustych wartości w tabelach
    if orphan_sources:
        mask_orphan = df_full[SOURCE_COL].astype(str).isin(orphan_sources)
        mask_orphan_rmpak = mask_orphan & df_full["Nr rysunku"].astype(str).apply(is_rmpak_number)
        df_full.loc[mask_orphan_rmpak, "Ilość całkowita"] = "0"

    # mapowanie numer -> ilość całkowita (string) dla RMPAK
    total_map = {normalize_drawing_for_match(item["num"]): format_qty(item["total"]) for item in totals}

    # uzupełnij kolumnę "Ilość całkowita" w df_full (dla numerów RMPAK z DRZEWKA)
    df_full["Ilość całkowita"] = (
        df_full["Nr rysunku"]
        .astype(str)
        .map(lambda n: total_map.get(normalize_drawing_for_match(n), ""))
    )

    # Jeśli BOM jest OSIEROCONY (brak ścieżek z rootów) -> ilość całkowita = 0 (dla RMPAK)
    if "orphan_sources" in locals() and orphan_sources:
        mask_orphan = df_full[SOURCE_COL].astype(str).isin(orphan_sources)
        mask_orphan_rmpak = mask_orphan & df_full["Nr rysunku"].astype(str).apply(is_rmpak_number)
        df_full.loc[mask_orphan_rmpak, "Ilość całkowita"] = "0"

    # elementy ZNORMALIZOWANE – Ilość całkowita = Ilość lokalna * ilość BOM-u (rodzica z DRZEWKA)
    # Dzięki temu:
    # - jeśli BOM jest OSIEROCONY (brak ścieżek z rootów) -> total BOM = 0 -> znormalizowane też mają 0
    # - jeśli BOM jest w produkcji -> znormalizowane mnożą się przez ilość całkowitą BOM-u
    # Uwaga: ilość BOM-u bierzemy bezpośrednio z DRZEWKA (nodes),
    #        aby BOM-y spoza RMPAK (lub z nietypowym numerem) miały poprawną ilość.
    total_float_by_num: Dict[str, float] = {}
    for node in nodes:
        num = normalize_drawing_for_match(node.get("num", ""))
        if not num:
            continue
        level = int(node.get("level", 0) or 0)
        qty_total = node.get("qty_total")
        eff_total = 1.0 if level == 0 else qty_total
        if eff_total is None:
            continue
        total_float_by_num[num] = total_float_by_num.get(num, 0.0) + float(eff_total)

    total_bom_by_source: Dict[str, float] = {}
    for src in df_full[SOURCE_COL].astype(str).unique().tolist():
        bom_num = normalize_drawing_for_match(extract_number_from_source_name(src))
        total_bom_by_source[str(src)] = float(total_float_by_num.get(bom_num, 0.0))

    mask_norm = ~df_full["Nr rysunku"].astype(str).apply(is_rmpak_number)

    def _norm_total(row) -> str:
        qty_local = parse_qty(row.get("Ilość", ""))
        if qty_local is None:
            return ""
        src = str(row.get(SOURCE_COL, ""))
        bom_total = float(total_bom_by_source.get(src, 0.0))
        return format_qty(qty_local * bom_total)

    df_full.loc[mask_norm, "Ilość całkowita"] = df_full.loc[mask_norm].apply(_norm_total, axis=1)

    # sekcje (już z kolumną "Ilość całkowita")
    sections = split_sections(df_full)
    # błędy plikowe + nowe błędy DWF dla numerów z BOM

    # OSIEROCONY BOM – wpis w BŁĘDY
    errors_orphan: List[List[str]] = []
    if 'orphan_boms' in locals() and orphan_boms:
        orphan_sources = {src for _, src in orphan_boms}

        # przygotuj szybkie wyszukiwanie "kto używa tego numeru"
        nr_norm = df_full["Nr rysunku"].astype(str).map(normalize_drawing_for_match)
        src_col = df_full[SOURCE_COL].astype(str) if SOURCE_COL in df_full.columns else pd.Series(dtype=str)

        dwf_map: Dict[str, List[Path]] = files.get("dwf_map", {})
        dxf_map: Dict[str, List[Path]] = files.get("dxf_map", {})
        stp_map: Dict[str, List[Path]] = files.get("stp_map", {})
        stl_map: Dict[str, List[Path]] = files.get("stl_map", {})
        idw_map: Dict[str, List[Path]] = files.get("idw_map", {})

        def _related_for_key(k: str) -> List[Path]:
            out: List[Path] = []
            for mp in (dwf_map, dxf_map, stp_map, stl_map, idw_map):
                out.extend(mp.get(k, []))
            # unique
            seen=set(); uniq=[]
            for p in out:
                sp=str(p)
                if sp not in seen:
                    seen.add(sp); uniq.append(p)
            return uniq

        for bom_num_clean, src_str in orphan_boms:
            used_by = []
            try:
                if SOURCE_COL in df_full.columns:
                    used_by = sorted({s for s in src_col[nr_norm == bom_num_clean].tolist() if s and s != src_str})
            except Exception:
                used_by = []

            if not used_by:
                why = f"BOM {src_str} nie występuje w żadnym innym BOM-ie projektu (ilość całkowita = 0)."
            else:
                if all(u in orphan_sources for u in used_by):
                    why = f"BOM {src_str} jest użyty w: {', '.join(used_by)}, ale te BOM-y też nie prowadzą do rootów (ilość całkowita = 0)."
                else:
                    why = f"BOM {src_str} nie ma ścieżki do rootów projektu (ilość całkowita = 0). Używany w: {', '.join(used_by)}."

            # lista plików związanych z BOM-em
            files_list: List[Path] = []
            p_csv = csv_path_by_source.get(src_str)
            if p_csv:
                files_list.append(p_csv)
            files_list.extend(_related_for_key(bom_num_clean))

            errors_orphan.append([
                "OSIEROCONY BOM",
                (p_csv.name if p_csv else f"{src_str}.csv"),
                why,
                format_file_list(files_list),
                source_katalog.get(src_str, katalog_from_rmpak_number(bom_num_clean)),
            ])


    _check_cancel()
    # Błędy plikowe liczymy tylko dla elementów z ilością całkowitą > 0 (odszumianie),
    # czyli tych, które mają ścieżki z rootów w DRZEWKU.
    # mapowanie numer -> Nazwa z BOM (do pełnych nazw w arkuszu BŁĘDY)
    name_by_num: Dict[str, str] = {}
    try:
        if df_full is not None and hasattr(df_full, "iterrows") and ("Nr rysunku" in df_full.columns) and ("Nazwa" in df_full.columns):
            for _i, _r in df_full.iterrows():
                _nr = str(_r.get("Nr rysunku", "") or "")
                _key = normalize_drawing_for_match(_nr)
                if not _key:
                    continue
                _nm = str(_r.get("Nazwa", "") or "").strip()
                if _nm and (_key not in name_by_num):
                    name_by_num[_key] = _nm
    except Exception:
        name_by_num = {}

    _check_cancel()
    errors_files = find_file_errors(files, active_nums=active_nums, name_by_num=name_by_num)
    _check_cancel()
    errors_filename_mismatch = find_filename_mismatch_errors(files, name_by_num=name_by_num, df_full=df_full)
    _check_cancel()
    errors_file_number_mismatch = find_file_number_mismatch_errors(files, df_full=df_full, name_by_num=name_by_num)
    _check_cancel()
    errors_suffix = find_suffix_consistency_errors(files)  # ten błąd zostawiamy zawsze
    _check_cancel()
    errors_csv_dup_num = find_csv_duplicate_drawing_number_errors(files)
    _check_cancel()
    errors_missing = find_bom_missing_files_errors(df_full, files, active_nums=active_nums)
    _check_cancel()
    errors_dup_in_bom = find_bom_duplicate_in_single_bom_errors(df_full)
    _check_cancel()
    errors_name_mismatch = find_bom_name_mismatch_errors(df_full, csv_path_by_source)
    _check_cancel()
    errors_spacing = find_bom_spacing_errors(files)
    _check_cancel()
    errors_missing_csv = find_missing_csv_for_assemblies_errors(files, files.get("csv_files", []), name_by_num=name_by_num)
    _check_cancel()
    errors_missing_dxf = find_bom_missing_dxf_errors(
        df_full, dxf_map,
        dwf_map_project=files.get("dwf_map_project", {}),
        dwf_map_library=files.get("dwf_map_library", {}),
    )
    
    # ZAMIATANIE KOŃCOWE: pliki nieprzypisane do żadnego BOM
    # Zbierz numery które już zostały zgłoszone w błędach o wyższym priorytecie
    _check_cancel()
    already_reported_numbers = set()
    for error_list in [errors_files, errors_suffix, errors_missing, errors_dup_in_bom, errors_name_mismatch, errors_missing_csv, errors_missing_dxf, errors_file_number_mismatch, errors_orphan, errors_csv_dup_num]:
        for error_row in error_list:
            # error_row format: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
            # Wyekstraktuj numer z "Nazwa pliku" (index 1)
            if len(error_row) > 1 and error_row[1]:
                # Nazwa pliku może być: "2556-100.40XX.dwf" lub sam numer "2556-100.40XX"
                name = str(error_row[1])
                # Spróbuj wyekstraktować numer (pierwszy token przed spacją lub cała nazwa)
                num_candidate = name.split(" ", 1)[0] if " " in name else name
                # Usuń rozszerzenie jeśli jest
                num_candidate = Path(num_candidate).stem if "." in num_candidate else num_candidate
                num_key = normalize_drawing_for_match(num_candidate)
                if num_key and is_rmpak_number(num_key):
                    already_reported_numbers.add(num_key)
    
    # Wyklucz również root BOM (matki) – nie powinny trafić do "pliki śmieci"
    try:
        root_nums = {normalize_drawing_for_match(r) for r in hierarchy.get("roots", []) if r}
    except Exception:
        root_nums = set()
    already_reported_numbers.update(root_nums)

    # Jeśli wybrano poddrzewo (selected_root), pomijamy w walidacji orphaned WSZYSTKIE
    # numery plików spoza tego poddrzewa. df_full i hierarchy są już przefiltrowane,
    # ale `files` nie – bez tego CSV/IDW/DWF z innych modułów wyglądały jak osierocone.
    # (Tryb szybki rozwiązuje to przez filter_files_to_subtree(files, active_nums).)
    if selected_root_normalized:
        subtree_nums: set = set(active_nums) if active_nums else set()
        try:
            if df_full is not None and not df_full.empty and "Nr rysunku" in df_full.columns:
                for _nr in df_full["Nr rysunku"].astype(str).tolist():
                    _k = normalize_drawing_for_match(_nr)
                    if _k:
                        subtree_nums.add(_k)
        except Exception:
            pass
        try:
            for _src in df_full[SOURCE_COL].astype(str).unique():
                _k = normalize_drawing_for_match(extract_number_from_source_name(str(_src)))
                if _k:
                    subtree_nums.add(_k)
        except Exception:
            pass
        subtree_nums.add(selected_root_normalized)

        out_of_subtree: set = set()
        for _csv in files.get("csv_files", []) or []:
            _k = normalize_drawing_for_match(extract_drawing_number_from_filename(_csv))
            if _k and _k not in subtree_nums:
                out_of_subtree.add(_k)
        for _map_key in ("idw_map", "dwf_map", "dxf_map", "stp_map", "stl_map"):
            for _k in (files.get(_map_key, {}) or {}).keys():
                if _k and _k not in subtree_nums:
                    out_of_subtree.add(_k)

        if out_of_subtree:
            log(f"📌 Pomijam w walidacji orphaned {len(out_of_subtree)} numerów spoza wybranego poddrzewa")
            already_reported_numbers.update(out_of_subtree)

    # Wywołaj funkcję z wykluczeniem numerów już zgłoszonych
    # UWAGA: orphaned używa df_full_unfiltered (cały projekt), żeby BOM-y rodziców
    # spoza wybranego poddrzewa nadal liczyły się jako "rodzice" – analogicznie do
    # trybu szybkiego dla głównego rootu (który nie filtruje df_full).
    _check_cancel()
    errors_orphaned_bom = find_orphaned_bom_errors(files, df_full_unfiltered, exclude_numbers=already_reported_numbers, name_by_num=name_by_num)
    _check_cancel()
    errors_orphaned_files = find_orphaned_files_errors(files, df_full_unfiltered, exclude_numbers=already_reported_numbers, name_by_num=name_by_num)

    errors = errors_spacing + errors_files + errors_filename_mismatch + errors_file_number_mismatch + errors_suffix + errors_csv_dup_num + errors_missing + errors_dup_in_bom + errors_name_mismatch + errors_missing_csv + errors_missing_dxf + errors_orphan + errors_orphaned_bom + errors_orphaned_files

    # Określ folder zapisu
    # - Jeśli wybrano TOP-LEVEL ROOT projektu (lub wszystkie moduły) → główny folder projektu
    # - Jeśli wybrano podmoduł → folder gdzie leży jego CSV
    output_dir = project_dir  # domyślnie główny folder

    if selected_root_normalized and selected_root_normalized not in original_top_level_roots:
        # Znajdź plik CSV wybranego modułu (podmoduł)
        source_name = parent_for_num.get(selected_root_normalized)
        if source_name and source_name in csv_path_by_source:
            csv_path = csv_path_by_source[source_name]
            output_dir = csv_path.parent
            log(f"📁 Folder zapisu (z CSV wybranego podmodułu): {output_dir}")
        else:
            log(f"⚠️ Nie znaleziono CSV dla modułu {selected_root_normalized}, zapis w głównym folderze")
    else:
        if selected_root_normalized:
            log(f"📁 Folder zapisu (wybrany moduł jest ROOT-em projektu → główny folder): {output_dir}")
        else:
            log(f"📁 Folder zapisu (główny folder projektu): {output_dir}")
    
    # zapis Excela
    output_path = output_dir / output_excel_name
    log(f"Zapisuję: {output_path}")
    _check_cancel()
    export_to_excel(
        sections,
        output_path,
        df_full,
        errors,
        hierarchy,
        nodes,
        totals,
        bom_total_by_source=total_bom_by_source,
        orphan_sources=(orphan_sources if 'orphan_sources' in locals() else set()),
    )

    log(f"Gotowe. Plik: {output_path}")
    return output_excel_name


def process_errors_only_DEPRECATED(project_dir: Path, library_dir: Path = None, use_library: bool = True, cancel_check=None) -> List[List[str]]:
    """
    DEPRECATED - Ta funkcja jest niekompletna. Użyj wersji poniżej.
    """
    raise NotImplementedError("Ta wersja funkcji jest niekompletna. Użyj pełnej wersji poniżej.")


def get_available_roots(project_dir: Path, cancel_check=None) -> Tuple[List[Tuple[str, str]], int]:
    """
    Szybkie skanowanie projektu aby pobrać listę dostępnych modułów (BOMów).
    
    Args:
        project_dir: Ścieżka do folderu projektu
        cancel_check: Funkcja sprawdzająca anulowanie
        
    Returns:
        Tuple[Lista krotek (numer_modułu, etykieta_modułu), liczba_rootów].
        Na początku listy są ROOTy (najwyższe moduły), potem pozostałe alfabetycznie.
    """
    def _check_cancel():
        if cancel_check and cancel_check():
            raise UserCancelled("Operacja przerwana")
    
    try:
        _check_cancel()
        files = scan_project_files(project_dir, cancel_check=_check_cancel)
        csv_files = files.get("csv_files", [])
        
        if not csv_files:
            return [], 0
        
        # Wczytaj minimalnie potrzebne dane z CSV
        all_bom_frames = []
        for csv_path in csv_files:
            _check_cancel()
            try:
                df = read_bom_csv(csv_path)
                df[SOURCE_COL] = csv_path.stem
                all_bom_frames.append(df)
            except Exception:
                continue  # Pomiń błędne pliki
        
        if not all_bom_frames:
            return [], 0
        
        _check_cancel()
        df_full = pd.concat(all_bom_frames, ignore_index=True)
        
        # Zbuduj hierarchię aby mieć etykiety i roots
        _check_cancel()
        hierarchy = build_hierarchy(df_full)
        
        labels = hierarchy.get("labels", {})
        parent_for_num = hierarchy.get("parent_for_num", {})
        roots = hierarchy.get("roots", [])
        
        # Rozdziel moduły na ROOTy i pozostałe
        root_modules = []
        other_modules = []
        
        for num, source_name in parent_for_num.items():
            label = labels.get(num, source_name)
            if num in roots:
                root_modules.append((num, label))
            else:
                other_modules.append((num, label))
        
        # Sortuj każdą grupę alfabetycznie
        root_modules.sort(key=lambda x: x[1])
        other_modules.sort(key=lambda x: x[1])
        
        # Połącz: najpierw ROOTy, potem reszta
        module_list = root_modules + other_modules
        roots_count = len(root_modules)
        
        return module_list, roots_count
        
    except Exception as e:
        log(f"Błąd pobierania modułów: {e}")
        return [], 0


def process_errors_only(project_dir: Path, library_dir: Path = None, use_library: bool = True, selected_root: str = None, cancel_check=None) -> List[List[str]]:
    """
    Tryb szybki - tylko walidacja błędów bez generowania pełnego Excel.
    Pomija budowę drzewka, totali, merge BOM - tylko sprawdza błędy.
    
    Args:
        project_dir: Ścieżka do folderu projektu
        library_dir: Ścieżka do folderu biblioteki DWF (opcjonalna)
        use_library: Czy skanować bibliotekę DWF
        selected_root: Wybrany moduł do przetworzenia (None = wszystkie)
        cancel_check: Funkcja sprawdzająca czy użytkownik anulował
    
    Returns:
        List[List[str]]: Lista błędów w formacie [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
    """
    log(f"[TRYB SZYBKI] Projekt: {project_dir}")
    
    if selected_root:
        log(f"[TRYB SZYBKI] Wybrane poddrzewo: {selected_root}")

    def _check_cancel() -> None:
        if cancel_check and cancel_check():
            raise UserCancelled("Walidacja została przerwana przez użytkownika.")

    _check_cancel()
    files = scan_project_files(project_dir, cancel_check=_check_cancel)
    _check_cancel()

    # Biblioteka DWF
    dwf_map_library: Dict[str, List[Path]] = {}
    lib_path = library_dir if library_dir else LIBRARY_DWF_ROOT
    
    if use_library and lib_path and lib_path.exists():
        try:
            log(f"[TRYB SZYBKI] Skanowanie biblioteki DWF: {lib_path}")
            dwf_map_library = scan_library_dwf(lib_path)
            if dwf_map_library:
                log(f"✅ Znaleziono DWF w bibliotece: {sum(len(v) for v in dwf_map_library.values())}")
        except Exception as e:
            log(f"❌ Błąd skanowania biblioteki: {e}")
            dwf_map_library = {}

    files["dwf_map_project"] = files.get("dwf_map", {})
    files["dwf_map_library"] = dwf_map_library

    csv_path_by_source: Dict[str, Path] = {p.stem: p for p in files.get('csv_files', [])}
    csv_files: List[Path] = files["csv_files"]
    dxf_map: Dict[str, List[Path]] = files["dxf_map"]
    stl_map: Dict[str, List[Path]] = files["stl_map"]
    stp_map: Dict[str, List[Path]] = files["stp_map"]
    dwf_map: Dict[str, List[Path]] = files["dwf_map"]

    if not csv_files:
        log("[TRYB SZYBKI] Brak plików CSV (BOM) w projekcie")
        return []

    # Wczytaj CSV
    all_bom_frames: List[pd.DataFrame] = []
    csv_read_errors: List[tuple] = []

    for csv_path in csv_files:
        _check_cancel()
        try:
            df = read_bom_csv(csv_path)
        except Exception as e:
            csv_read_errors.append((csv_path, str(e)))
            log(f"  ❌ BŁĄD czytania CSV {csv_path.name}: {e}")
            continue

        df[SOURCE_COL] = csv_path.stem
        df = assign_katalog(df, csv_path, dxf_map, stl_map, stp_map, dwf_map)
        df = update_material_from_dxf(df, dxf_map, dwf_map_project=files.get('dwf_map_project', {}), dwf_map_library=files.get('dwf_map_library', {}))
        df = assign_3d_files(df, stl_map, stp_map)
        all_bom_frames.append(df)

    if not all_bom_frames:
        log("[TRYB SZYBKI] Żaden CSV nie został wczytany poprawnie")
        return []

    _check_cancel()
    df_full = pd.concat(all_bom_frames, ignore_index=True)

    if SOURCE_COL not in df_full.columns:
        df_full[SOURCE_COL] = ""

    # Mapowanie: numer -> nazwa (dla pełnych nazw w błędach)
    name_by_num: Dict[str, str] = {}
    if "Nr rysunku" in df_full.columns and "Nazwa" in df_full.columns:
        for _i, _r in df_full.iterrows():
            _nr = str(_r.get("Nr rysunku", "") or "")
            _key = normalize_drawing_for_match(_nr)
            if not _key:
                continue
            _nm = str(_r.get("Nazwa", "") or "").strip()
            if _nm and (_key not in name_by_num):
                name_by_num[_key] = _nm

    # BUDOWA HIERARCHII I ACTIVE_NUMS (dla filtrowania błędów do wybranego poddrzewa)
    active_nums: set[str] | None = None
    is_root_module = False
    
    if selected_root:
        # Jeśli wybrano konkretny moduł, budujemy hierarchię i active_nums
        log("[TRYB SZYBKI] Budowa hierarchii dla filtrowania do wybranego poddrzewa...")
        _check_cancel()
        hierarchy = build_hierarchy(df_full)
        
        # Normalizuj selected_root
        selected_root_normalized = normalize_drawing_for_match(selected_root)
        
        # Sprawdź czy to główny ROOT (najwyższy poziom)
        roots = hierarchy.get("roots", [])
        is_root_module = (selected_root_normalized in roots)
        
        if is_root_module:
            log(f"[TRYB SZYBKI] Wybrany moduł '{selected_root}' jest głównym ROOT-em")
            log("[TRYB SZYBKI] Sprawdzanie WSZYSTKICH błędów w katalogu (pełny BOM + wszystkie pliki)")
            
            # Dla ROOT-a: NIE filtrujemy df_full ani files, NIE ustawiamy active_nums
            # active_nums pozostaje None aby sprawdzić WSZYSTKIE błędy (włącznie z orphaned files)
            log("[TRYB SZYBKI] active_nums = None (sprawdzanie wszystkich plików, włącznie z orphaned)")
            
        else:
            log(f"[TRYB SZYBKI] Wybrany moduł '{selected_root}' jest podmodułem")
            log("[TRYB SZYBKI] Filtrowanie BOM i plików do wybranego poddrzewa...")
            
            # Filtruj do poddrzewa (BOM)
            df_full, hierarchy = filter_to_subtree(df_full, hierarchy, selected_root_normalized)
            log(f"[TRYB SZYBKI] Przefiltrowano do {len(df_full)} wierszy z poddrzewa {selected_root}")
            
            # Zaktualizuj name_by_num dla przefiltrowanego df_full
            name_by_num = {}
            if "Nr rysunku" in df_full.columns and "Nazwa" in df_full.columns:
                for _i, _r in df_full.iterrows():
                    _nr = str(_r.get("Nr rysunku", "") or "")
                    _key = normalize_drawing_for_match(_nr)
                    if not _key:
                        continue
                    _nm = str(_r.get("Nazwa", "") or "").strip()
                    if _nm and (_key not in name_by_num):
                        name_by_num[_key] = _nm
            
            # Zbuduj totals aby mieć active_nums
            nodes, _ = traverse_hierarchy(hierarchy)
            parent_for_num = hierarchy.get("parent_for_num", {})
            totals = compute_total_quantities(nodes, df_full, parent_for_num)
            
            # Aktywne numery = te, które mają ilość całkowitą > 0
            active_nums = set()
            for item in totals:
                n = normalize_drawing_for_match(item["num"])
                if float(item.get("total") or 0.0) > 0.0:
                    active_nums.add(n)
            
            log(f"[TRYB SZYBKI] Znaleziono {len(active_nums)} aktywnych numerów w poddrzewie")
            
            # FILTRUJ PLIKI do tylko tych z poddrzewa
            files = filter_files_to_subtree(files, active_nums)
            
            # Zaktualizuj zmienne używane później
            csv_files = files.get("csv_files", [])
            dxf_map = files.get("dxf_map", {})
            stl_map = files.get("stl_map", {})
            stp_map = files.get("stp_map", {})
            dwf_map = files.get("dwf_map", {})
            csv_path_by_source = {p.stem: p for p in csv_files}
            
            log(f"[TRYB SZYBKI] Po filtrowaniu: {len(csv_files)} plików CSV, "
                f"{sum(len(v) for v in dxf_map.values())} DXF, "
                f"{sum(len(v) for v in dwf_map.values())} DWF")

    # WALIDACJA BŁĘDÓW
    log("[TRYB SZYBKI] Walidacja błędów...")
    
    _check_cancel()
    errors_files = find_file_errors(files, active_nums=active_nums, name_by_num=name_by_num)
    _check_cancel()
    errors_filename_mismatch = find_filename_mismatch_errors(files, name_by_num=name_by_num, df_full=df_full)
    _check_cancel()
    errors_file_number_mismatch = find_file_number_mismatch_errors(files, df_full=df_full, name_by_num=name_by_num)
    _check_cancel()
    errors_suffix = find_suffix_consistency_errors(files)
    _check_cancel()
    errors_csv_dup_num = find_csv_duplicate_drawing_number_errors(files)
    _check_cancel()
    errors_missing = find_bom_missing_files_errors(df_full, files, active_nums=active_nums)
    _check_cancel()
    errors_dup_in_bom = find_bom_duplicate_in_single_bom_errors(df_full)
    _check_cancel()
    errors_name_mismatch = find_bom_name_mismatch_errors(df_full, csv_path_by_source)
    _check_cancel()
    errors_spacing = find_bom_spacing_errors(files)
    _check_cancel()
    errors_missing_csv = find_missing_csv_for_assemblies_errors(files, files.get("csv_files", []), name_by_num=name_by_num)
    _check_cancel()
    errors_missing_dxf = find_bom_missing_dxf_errors(
        df_full, dxf_map,
        dwf_map_project=files.get("dwf_map_project", {}),
        dwf_map_library=files.get("dwf_map_library", {}),
    )
    
    # Zbierz już zgłoszone numery
    _check_cancel()
    already_reported_numbers = set()
    for error_list in [errors_files, errors_suffix, errors_missing, errors_dup_in_bom, errors_name_mismatch, errors_missing_csv, errors_missing_dxf, errors_file_number_mismatch, errors_csv_dup_num]:
        for error_row in error_list:
            if len(error_row) > 1 and error_row[1]:
                name = str(error_row[1])
                num_candidate = name.split(" ", 1)[0] if " " in name else name
                num_candidate = Path(num_candidate).stem if "." in num_candidate else num_candidate
                num_key = normalize_drawing_for_match(num_candidate)
                if num_key and is_rmpak_number(num_key):
                    already_reported_numbers.add(num_key)
    
    # Dla podmodułu: dodaj SAM wybrany moduł do exclude (bo po filtrowaniu zawsze wygląda na osierocony)
    # Ale NADAL sprawdzaj orphaned WEWNĄTRZ modułu (jego elementy składowe)
    if selected_root and not is_root_module:
        selected_root_normalized = normalize_drawing_for_match(selected_root)
        already_reported_numbers.add(selected_root_normalized)
        log(f"[TRYB SZYBKI] Pomijam sam wybrany moduł '{selected_root}' w walidacji orphaned")

    # Walidacja orphaned (pomijając już zgłoszone + wybrany moduł dla podmodułów)
    log("[TRYB SZYBKI] Sprawdzanie plików osieroconych (orphaned)...")
    _check_cancel()
    errors_orphaned_bom = find_orphaned_bom_errors(files, df_full, exclude_numbers=already_reported_numbers, name_by_num=name_by_num)
    _check_cancel()
    errors_orphaned_files = find_orphaned_files_errors(files, df_full, exclude_numbers=already_reported_numbers, name_by_num=name_by_num)

    # Połącz wszystkie błędy (orphaned files na końcu)
    all_errors = errors_spacing + errors_files + errors_filename_mismatch + errors_file_number_mismatch + errors_suffix + errors_csv_dup_num + errors_missing + errors_dup_in_bom + errors_name_mismatch + errors_missing_csv + errors_missing_dxf + errors_orphaned_bom + errors_orphaned_files

    log(f"[TRYB SZYBKI] Znaleziono {len(all_errors)} błędów")
    return all_errors


def _pick_project_dir_gui(initial: Path | None = None) -> Path | None:
    """Windowsowe okno wyboru katalogu projektu."""
    root = tk.Tk()
    root.withdraw()
    # na wierzch
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass

    initdir = str(initial) if initial and initial.exists() else None
    selected = filedialog.askdirectory(
        title="Wybierz katalog projektu (LOGISTYKA_AUTOMAT)",
        initialdir=initdir,
        mustexist=True,
    )
    root.destroy()

    if not selected:
        return None
    return Path(selected).resolve()


def main():
    # Uruchom GUI
    app = LogistykaGUI()
    app.mainloop()


class LogistykaGUI(tk.Tk):
    """Graficzny interfejs dla RM_IMPORT"""
    
    def __init__(self):
        super().__init__()
        
        self.title("📊 RM_IMPORT_V16_UTF")
        self.geometry("1000x900")  # Zwiększona wysokość aby wszystkie elementy (w tym status bar) były w pełni widoczne
        self.resizable(True, True)
        
        # Zmienne
        self.project_dir = None
        self.scanned_files = None
        self.is_processing = False
        self.is_scanning = False
        self.cancel_requested = False
        self.reset_pending = False
        self.last_path = None
        self.library_dir = Path(r"B:\\")  # Domyślna ścieżka biblioteki
        self.use_library = tk.BooleanVar(value=True)  # Czy skanować bibliotekę
        
        # Wczytaj ostatnią ścieżkę
        self._load_last_path()
        
        # Ustaw callback logów
        log.gui_callback = self.append_log
        
        self._create_widgets()
    
    def _create_widgets(self):
        """Buduje GUI"""
        
        # === HEADER ===
        header = ttk.Frame(self, padding=(10, 10))
        header.pack(fill="x")
        
        title_lbl = ttk.Label(
            header,
            text="📊 RM_IMPORT_V16_UTF",
            font=("Segoe UI", 14, "bold")
        )
        title_lbl.pack(anchor="w")
        
        subtitle_lbl = ttk.Label(
            header,
            text="Przetwarza pliki CSV z BOM i generuje zbiorczy Excel z analizą",
            foreground="gray"
        )
        subtitle_lbl.pack(anchor="w", pady=(2, 0))
        
        ttk.Separator(self, orient="horizontal").pack(fill="x", pady=10)
        
        # === WYBÓR FOLDERU ===
        folder_frame = ttk.LabelFrame(self, text=" 📁 Folder projektu ", padding=(10, 10))
        folder_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.folder_var = tk.StringVar(value="Nie wybrano folderu...")
        ttk.Label(folder_frame, textvariable=self.folder_var, foreground="blue").pack(side="left", fill="x", expand=True)
        ttk.Button(folder_frame, text="📁 Wybierz folder...", command=self.choose_folder).pack(side="right", padx=(10, 0))
        ttk.Button(folder_frame, text="🔍 Skanuj", command=self.scan_folder, state="disabled").pack(side="right")
        
        # === WYBÓR FOLDERU BIBLIOTEKI DWF ===
        library_frame = ttk.LabelFrame(self, text=" 📚 Biblioteka DWF (opcjonalnie) ", padding=(10, 10))
        library_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.library_var = tk.StringVar(value=str(self.library_dir))
        
        # Checkbox do włączania/wyłączania skanowania biblioteki
        self.lib_checkbox = ttk.Checkbutton(
            library_frame,
            text="Skanuj bibliotekę",
            variable=self.use_library
        )
        self.lib_checkbox.pack(side="left", padx=(0, 10))
        
        ttk.Label(library_frame, textvariable=self.library_var, foreground="gray").pack(side="left", fill="x", expand=True)
        ttk.Button(library_frame, text="📂 Zmień folder...", command=self.choose_library).pack(side="right")
        
        # === ZNALEZIONE PLIKI ===
        files_frame = ttk.LabelFrame(self, text=" 📋 Znalezione pliki ", padding=(10, 10))
        files_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Statystyki
        stats_frame = ttk.Frame(files_frame)
        stats_frame.pack(fill="x", pady=(0, 5))
        
        self.stats_var = tk.StringVar(value="Wybierz folder i naciśnij 'Skanuj' aby wyświetlić statystyki plików")
        ttk.Label(stats_frame, textvariable=self.stats_var).pack(anchor="w")
        
        # Lista plików CSV - interaktywna z przyciskami
        list_frame = ttk.Frame(files_frame)
        list_frame.pack(fill="both", expand=True)
        
        ttk.Label(list_frame, text="Pliki CSV (BOM):").pack(anchor="w", pady=(0, 3))
        
        # Canvas ze scrollbarem dla listy CSV
        canvas_frame = ttk.Frame(list_frame, relief="sunken", borderwidth=1)
        canvas_frame.pack(fill="both", expand=True)
        
        self.csv_canvas = tk.Canvas(canvas_frame, bg="white", highlightthickness=0)
        csv_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.csv_canvas.yview)
        self.csv_scroll_frame = ttk.Frame(self.csv_canvas)
        
        # Binding do aktualizacji scrollregion
        self.csv_scroll_frame.bind(
            "<Configure>",
            lambda e: self.csv_canvas.configure(scrollregion=self.csv_canvas.bbox("all"))
        )
        
        # Binding do scrollowania myszką - zwiększona czułość
        def on_mousewheel(event):
            # Większa wartość = szybsze scrollowanie (3 jednostki zamiast 1)
            self.csv_canvas.yview_scroll(int(-1*(event.delta/120)) * 3, "units")
            return "break"  # Zatrzymaj propagację eventu
        
        # Funkcja do rekurencyjnego bindowania scroll do wszystkich dzieci
        def bind_tree(widget):
            widget.bind("<MouseWheel>", on_mousewheel)
            for child in widget.winfo_children():
                bind_tree(child)
        
        # Bind do canvas i wszystkich jego dzieci
        self.csv_canvas.bind("<MouseWheel>", on_mousewheel)
        bind_tree(self.csv_scroll_frame)
        
        # Przechowaj funkcję bind_tree aby móc jej użyć po dodaniu nowych elementów
        self.csv_bind_tree = bind_tree
        
        # Binding do rozciągania scroll_frame na pełną szerokość Canvas
        def on_canvas_configure(event):
            self.csv_canvas.itemconfig(self.csv_window_id, width=event.width)
        
        self.csv_canvas.bind("<Configure>", on_canvas_configure)
        
        self.csv_window_id = self.csv_canvas.create_window((0, 0), window=self.csv_scroll_frame, anchor="nw")
        self.csv_canvas.configure(yscrollcommand=csv_scrollbar.set)
        
        self.csv_canvas.pack(side="left", fill="both", expand=True)
        csv_scrollbar.pack(side="right", fill="y")
        
        # Lista do przechowywania ścieżek CSV
        self.csv_file_paths = []
        
        # === LOGI ===
        log_frame = ttk.LabelFrame(self, text=" 📝 Log przetwarzania ", padding=(10, 10))
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, state="disabled", wrap="word")
        self.log_text.pack(fill="both", expand=True)
        
        # === AKCJE ===
        action_frame = ttk.Frame(self, padding=(10, 10))
        action_frame.pack(fill="x")
        
        # Checkbox trybu szybkiego
        self.quick_mode = tk.BooleanVar(value=False)
        self.quick_check = ttk.Checkbutton(
            action_frame,
            text="⚡ Tryb szybki - tylko błędy (bez generowania Excel)",
            variable=self.quick_mode
        )
        self.quick_check.pack(side="left", padx=(0, 20))
        
        self.btn_generate = ttk.Button(
            action_frame,
            text="▶️ Generuj Excel (nazwa z ROOT BOM)",
            command=self.generate_excel,
            state="disabled"
        )
        self.btn_generate.pack(side="left", padx=(0, 10))

        self.btn_stop = ttk.Button(
            action_frame,
            text="⏹️ STOP",
            command=self.stop_processing,
            state="disabled"
        )
        self.btn_stop.pack(side="left", padx=(0, 10))

        self.btn_reset = ttk.Button(
            action_frame,
            text="🔄 RESET",
            command=self.reset_app,
            state="normal"
        )
        self.btn_reset.pack(side="left", padx=(0, 10))
        
        self.progress = ttk.Progressbar(action_frame, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        ttk.Button(action_frame, text="❌ Zamknij", command=self.quit).pack(side="right")
        
        self.btn_open_folder = ttk.Button(
            action_frame,
            text="📁 Otwórz folder",
            command=self.open_project_folder,
            state="disabled"
        )
        self.btn_open_folder.pack(side="right", padx=(0, 10))
        
        # === STATUS BAR ===
        self.status_var = tk.StringVar(value="Gotowy. Wybierz folder projektu.")
        status_bar = ttk.Label(
            self,
            textvariable=self.status_var,
            relief="sunken",
            padding=(5, 5)  # Zwiększony padding dla lepszej widoczności
        )
        status_bar.pack(fill="x", side="bottom", pady=(5, 0))
    
    def append_log(self, msg: str):
        """Dodaj wiadomość do logu"""
        # Bezpiecznie z wątków: przekieruj do głównego wątku GUI
        try:
            if threading.current_thread() is not threading.main_thread():
                self.after(0, lambda: self.append_log(msg))
                return
        except Exception:
            pass
        self.log_text.config(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")
        self.update_idletasks()
    
    def choose_folder(self):
        """Wybierz folder projektu"""
        initial_dir = str(self.last_path) if self.last_path and self.last_path.exists() else str(Path.cwd())
        
        folder = filedialog.askdirectory(
            title="Wybierz folder projektu z plikami CSV",
            initialdir=initial_dir
        )
        
        if folder:
            self.project_dir = Path(folder)
            self.folder_var.set(str(self.project_dir))
            self.status_var.set(f"Wybrany folder: {self.project_dir.name}")
            
            # Zapisz ostatnią ścieżkę
            self._save_last_path(self.project_dir)
            
            # Aktywuj przycisk skanowania
            for widget in self.winfo_children():
                if isinstance(widget, ttk.LabelFrame) and "Folder" in widget.cget("text"):
                    for btn in widget.winfo_children():
                        if isinstance(btn, ttk.Button) and "Skanuj" in btn.cget("text"):
                            btn.config(state="normal")
    
    def choose_library(self):
        """Wybierz folder biblioteki DWF"""
        initial_dir = str(self.library_dir) if self.library_dir.exists() else str(Path.cwd())
        
        folder = filedialog.askdirectory(
            title="Wybierz folder biblioteki DWF",
            initialdir=initial_dir
        )
        
        if folder:
            self.library_dir = Path(folder)
            self.library_var.set(str(self.library_dir))
            self.status_var.set(f"Biblioteka DWF: {self.library_dir.name}")
    
    def scan_folder(self):
        """Skanuj folder w poszukiwaniu plików"""
        if not self.project_dir or not self.project_dir.exists():
            messagebox.showwarning("Błąd", "Folder nie istnieje!")
            return

        if self.is_scanning:
            return

        self.is_scanning = True
        self.cancel_requested = False
        self.btn_stop.config(state="normal")
        self.progress.start(10)
        self.status_var.set("Skanowanie plików...")
        self.update_idletasks()

        thread = threading.Thread(target=self._run_scan)
        thread.daemon = True
        thread.start()

    def _run_scan(self):
        """Skanowanie w osobnym wątku z możliwością STOP"""
        try:
            self.append_log("\n" + "="*60)
            self.append_log(f"Skanowanie folderu: {self.project_dir}")
            if self.use_library.get() and self.library_dir.exists():
                self.append_log(f"Biblioteka DWF: {self.library_dir}")
            self.append_log("="*60)

            scanned = scan_project_files(self.project_dir, cancel_check=lambda: self.cancel_requested)
            self.after(0, lambda: self._on_scan_success(scanned))

        except UserCancelled as e:
            msg = str(e)
            self.append_log(f"\n⏹️ {msg}")
            self.after(0, lambda m=msg: self._on_scan_cancel(m))

        except Exception as e:
            msg = str(e)
            self.append_log(f"❌ Błąd skanowania: {msg}")
            self.after(0, lambda m=msg: self._on_scan_error(m))

    def _on_scan_success(self, scanned_files: Dict[str, object]):
        """Callback po udanym skanowaniu"""
        self.scanned_files = scanned_files
        self.is_scanning = False
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.progress.stop()

        # Wyświetl statystyki
        csv_count = len(self.scanned_files.get("csv_files", []))
        idw_count = len(self.scanned_files.get("idw_files", []))
        dxf_count = sum(len(v) for v in self.scanned_files.get("dxf_map", {}).values())
        stl_count = sum(len(v) for v in self.scanned_files.get("stl_map", {}).values())
        stp_count = sum(len(v) for v in self.scanned_files.get("stp_map", {}).values())
        dwf_count = sum(len(v) for v in self.scanned_files.get("dwf_map", {}).values())

        stats_text = (
            f"CSV: {csv_count}  |  IDW: {idw_count}  |  "
            f"DXF: {dxf_count}  |  STL: {stl_count}  |  "
            f"STP: {stp_count}  |  DWF: {dwf_count}"
        )
        self.stats_var.set(stats_text)

        # Wypełnij listę CSV z przyciskami
        for widget in self.csv_scroll_frame.winfo_children():
            widget.destroy()

        self.csv_file_paths = self.scanned_files.get("csv_files", [])

        for idx, csv_file in enumerate(self.csv_file_paths):
            rel_path = csv_file.relative_to(self.project_dir) if csv_file.is_relative_to(self.project_dir) else csv_file

            row_frame = ttk.Frame(self.csv_scroll_frame)
            row_frame.pack(fill="x", pady=2)
            row_frame.columnconfigure(0, weight=1)
            row_frame.columnconfigure(1, weight=0)

            file_label = tk.Entry(row_frame, relief="flat", readonlybackground="white")
            file_label.insert(0, str(rel_path))
            file_label.config(state="readonly")
            file_label.grid(row=0, column=0, sticky="ew", padx=(5, 5))

            open_btn = ttk.Button(
                row_frame,
                text="📂",
                width=3,
                command=lambda f=csv_file: self.open_file_location(f)
            )
            open_btn.grid(row=0, column=1, sticky="e")
        
        # Rebind scroll do wszystkich nowo utworzonych elementów
        if hasattr(self, 'csv_bind_tree'):
            self.csv_bind_tree(self.csv_scroll_frame)

        if csv_count > 0:
            self.btn_generate.config(state="normal")
            self.btn_open_folder.config(state="normal")
            self.status_var.set(f"✅ Znaleziono {csv_count} plików CSV. Gotowy do generowania.")
        else:
            self.btn_generate.config(state="disabled")
            self.status_var.set("⚠️ Brak plików CSV w tym folderze!")
            messagebox.showwarning("Brak CSV", "W wybranym folderze nie znaleziono plików CSV z BOM!")

    def _on_scan_cancel(self, msg: str):
        """Callback po przerwanym skanowaniu"""
        self.is_scanning = False
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.progress.stop()
        self.status_var.set("⏹️ Skanowanie zatrzymane")

        messagebox.showinfo(
            "Skanowanie zatrzymane",
            f"{msg}\n\nMożesz uruchomić skanowanie ponownie lub użyć RESET."
        )

        if self.reset_pending:
            self._do_reset()

    def _on_scan_error(self, error_msg: str):
        """Callback po błędzie skanowania"""
        self.is_scanning = False
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.progress.stop()
        self.status_var.set("❌ Błąd skanowania")
        messagebox.showerror("Błąd", f"Nie można przeskanować folderu:\n{error_msg}")
    
    def generate_excel(self):
        """Generuj plik Excel"""
        if not self.project_dir or not self.scanned_files:
            messagebox.showwarning("Błąd", "Najpierw zeskanuj folder!")
            return
        
        if self.is_processing:
            messagebox.showinfo("Info", "Przetwarzanie w toku...")
            return
        
        # Pobierz dostępne moduły z projektu (ROOTy pierwsze, potem podmoduły)
        try:
            available_modules, roots_count = get_available_roots(self.project_dir, cancel_check=lambda: self.cancel_requested)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie można pobrać listy modułów:\n{str(e)}")
            return
        
        if not available_modules:
            messagebox.showwarning("Błąd", "Nie znaleziono żadnych modułów w projekcie!")
            return
        
        # Pokaż dialog wyboru ROOT
        selected_root = self._show_root_selection_dialog(available_modules, roots_count)
        
        if selected_root is None:
            # Użytkownik anulował wybór
            return
        
        # Zresetuj nazwę wygenerowanego pliku z poprzedniej sesji
        self.generated_filename = None
        
        # Zapisz wybrany ROOT dla _run_processing
        self.selected_root = selected_root
        
        # Wyłącz przyciski i rozpocznij przetwarzanie
        self.btn_generate.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.is_processing = True
        self.cancel_requested = False
        self.progress.start(10)
        self.status_var.set("⏳ Przetwarzanie...")
        
        # Wyczyść logi
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")
        
        # Uruchom w osobnym wątku
        thread = threading.Thread(target=self._run_processing)
        thread.daemon = True
        thread.start()
    
    def _show_root_selection_dialog(self, available_modules: List[Tuple[str, str]], roots_count: int) -> str:
        """
        Pokazuje dialog wyboru modułu do przetworzenia.
        
        Args:
            available_modules: Lista krotek (numer, etykieta) - ROOTy pierwsze, potem podmoduły
            roots_count: Liczba głównych modułów (ROOTów) na początku listy
            
        Returns:
            Wybrany numer modułu lub pusty string jeśli wszystkie, lub None jeśli anulowano
        """
        dialog = tk.Toplevel(self)
        dialog.title("Wybór modułu do przetworzenia")
        dialog.geometry("700x500")
        dialog.resizable(True, True)
        dialog.transient(self)
        dialog.grab_set()
        
        # Wycentruj dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        result = [None]  # Lista żeby móc modyfikować w funkcjach wewnętrznych
        
        # Treść
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill="both", expand=True)
        
        ttk.Label(
            frame,
            text="Wybierz moduł do przetworzenia:",
            font=("Arial", 11, "bold")
        ).pack(pady=(0, 10))
        
        ttk.Label(
            frame,
            text="Możesz wygenerować Excel dla całego projektu lub tylko dla wybranego poddrzewa.\n⭐ - moduł główny (ROOT)",
            wraplength=650
        ).pack(pady=(0, 15))
        
        # Lista ROOT-ów
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        
        # Listbox z opcjami
        listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=("Courier", 10),
            height=15
        )
        listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Dodaj opcję "Wszystkie moduły"
        listbox.insert("end", "▶ WSZYSTKIE MODUŁY (cały projekt)")
        listbox.itemconfig(0, bg="#e3f2fd")
        
        # Dodaj separację
        listbox.insert("end", "─" * 70)
        listbox.itemconfig(1, fg="#999", selectbackground="#999")
        
        # Dodaj moduły z oznakowaniem
        separator_index = None
        for idx, (num, label) in enumerate(available_modules):
            if idx < roots_count:
                # ROOTY - główne moduły
                listbox.insert("end", f"⭐ {label}")
            else:
                # Pierwszy podmoduł - dodaj separator
                if separator_index is None:
                    separator_index = listbox.size()
                    listbox.insert("end", "─" * 70)
                    listbox.itemconfig(separator_index, fg="#999", selectbackground="#999")
                # Podmoduły
                listbox.insert("end", f"   {label}")
        
        # Zaznacz domyślnie pierwszy moduł (indeks 2, bo 0=wszystkie, 1=separator)
        if len(available_modules) > 0:
            listbox.selection_set(2)
            listbox.see(2)
        
        # Podwójne kliknięcie = zatwierdzenie
        def on_double_click(event):
            confirm()
        
        listbox.bind("<Double-Button-1>", on_double_click)
        
        # Przyciski
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(10, 0))
        
        def confirm():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Brak wyboru", "Wybierz moduł z listy!", parent=dialog)
                return
            
            idx = selection[0]
            
            if idx == 0:
                # Wszystkie moduły
                result[0] = ""  # Pusty string = wszystkie
            elif idx == 1:
                # Pierwszy separator - nie można wybrać
                messagebox.showwarning("Błędny wybór", "Wybierz moduł z listy!", parent=dialog)
                return
            elif separator_index is not None and idx == separator_index:
                # Drugi separator (między ROOTami a podmodułami) - nie można wybrać
                messagebox.showwarning("Błędny wybór", "Wybierz moduł z listy!", parent=dialog)
                return
            else:
                # Konkretny moduł
                # idx 2 = pierwszy moduł (available_modules[0])
                # idx 3 = drugi moduł (available_modules[1])
                # ...
                # Jeśli jest separator_index, to wszystko po nim jest przesunięte o 1
                
                if separator_index is not None and idx > separator_index:
                    # Po separatorze - korekta o 3 (wszystkie + 2 separatory)
                    module_idx = idx - 3
                else:
                    # Przed separatorem lub brak separatora - korekta o 2 (wszystkie + separator)
                    module_idx = idx - 2
                
                if 0 <= module_idx < len(available_modules):
                    result[0] = available_modules[module_idx][0]  # numer modułu
                else:
                    messagebox.showerror("Błąd", "Nieprawidłowy wybór!", parent=dialog)
                    return
            
            dialog.destroy()
        
        def cancel():
            result[0] = None
            dialog.destroy()
        
        ttk.Button(btn_frame, text="✅ Generuj Excel", command=confirm, style="Accent.TButton").pack(side="right", padx=5)
        ttk.Button(btn_frame, text="❌ Anuluj", command=cancel).pack(side="right", padx=5)
        
        # Oczekuj na zamknięcie okna
        self.wait_window(dialog)
        
        return result[0]

    def stop_processing(self):
        """Zatrzymaj przetwarzanie"""
        if not self.is_processing and not self.is_scanning:
            return
        self.cancel_requested = True
        if self.is_scanning:
            self.status_var.set("⏹️ Zatrzymywanie skanowania...")
        else:
            self.status_var.set("⏹️ Zatrzymywanie przetwarzania...")
        self.btn_stop.config(state="disabled")

    def reset_app(self):
        """Resetuje stan aplikacji"""
        if self.is_processing or self.is_scanning:
            # jeśli coś trwa, najpierw zatrzymaj, a reset wykonaj po anulowaniu
            self.cancel_requested = True
            self.reset_pending = True
            self.status_var.set("⏹️ Zatrzymywanie przed resetem...")
            return

        self._do_reset()

    def _do_reset(self):
        """Właściwy reset (wywoływany gdy nic nie trwa)."""
        self.reset_pending = False
        self.cancel_requested = False
        self.project_dir = None
        self.scanned_files = None
        self.folder_var.set("Nie wybrano folderu...")
        self.stats_var.set("Wybierz folder i naciśnij 'Skanuj' aby wyświetlić statystyki plików")
        self.status_var.set("Gotowy. Wybierz folder projektu.")

        # wyczyść listę CSV
        for widget in self.csv_scroll_frame.winfo_children():
            widget.destroy()
        self.csv_file_paths = []

        # wyczyść logi
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

        # przyciski
        self.btn_generate.config(state="disabled")
        self.btn_stop.config(state="disabled")
        self.btn_open_folder.config(state="disabled")
    
    def _run_processing(self):
        """Uruchamia przetwarzanie w osobnym wątku"""
        try:
            self.append_log("\n" + "="*60)
            
            # Sprawdź tryb
            if self.quick_mode.get():
                # TRYB SZYBKI - tylko błędy
                self.append_log("START SZYBKIEJ WALIDACJI (TYLKO BŁĘDY)")
                self.append_log("="*60 + "\n")
                
                # Pobierz wybrany ROOT (jeśli wybrano konkretny)
                selected_root = getattr(self, 'selected_root', None)
                if selected_root == "":
                    # Pusty string oznacza "wszystkie"
                    selected_root = None
                
                if selected_root:
                    self.append_log(f"📌 Sprawdzanie błędów dla wybranego modułu: {selected_root}")
                    self.append_log("")
                else:
                    self.append_log("📌 Sprawdzanie błędów dla wszystkich modułów w projekcie")
                    self.append_log("")
                
                errors = process_errors_only(
                    self.project_dir,
                    library_dir=self.library_dir,
                    use_library=self.use_library.get(),
                    selected_root=selected_root,
                    cancel_check=lambda: self.cancel_requested
                )
                
                self.append_log("\n" + "="*60)
                self.append_log(f"✅ WALIDACJA ZAKOŃCZONA - znaleziono {len(errors)} błędów")
                self.append_log("="*60)
                
                # Przekaż błędy do GUI
                self.after(0, lambda: self._show_errors_window(errors))
                
            else:
                # TRYB NORMALNY - pełny Excel
                self.append_log("START PRZETWARZANIA")
                self.append_log("="*60 + "\n")
                
                # Pobierz wybrany ROOT (jeśli wybrano konkretny)
                selected_root = getattr(self, 'selected_root', None)
                if selected_root == "":
                    # Pusty string oznacza "wszystkie"
                    selected_root = None
                
                if selected_root:
                    self.append_log(f"📌 Przetwarzanie wybranego modułu: {selected_root}")
                    self.append_log("")
                else:
                    self.append_log("📌 Przetwarzanie wszystkich modułów w projekcie")
                    self.append_log("")
                
                output_filename = process_project(
                    self.project_dir,
                    library_dir=self.library_dir,
                    use_library=self.use_library.get(),
                    selected_root=selected_root,
                    cancel_check=lambda: self.cancel_requested
                )
                self.generated_filename = output_filename
                
                self.append_log("\n" + "="*60)
                self.append_log("✅ ZAKOŃCZONO POMYŚLNIE")
                self.append_log("="*60)
                
                # Wróć do głównego wątku GUI
                self.after(0, self._on_success)
            
        except UserCancelled as e:
            msg = str(e)
            self.append_log(f"\n⏹️ {msg}")
            self.after(0, lambda m=msg: self._on_cancel(m))

        except PermissionError as e:
            # Specjalna obsługa błędu Permission denied
            error_msg = str(e)
            self.append_log(f"\n❌ BŁĄD: {error_msg}")
            
            # Wróć do głównego wątku GUI
            self.after(0, lambda m=error_msg: self._on_permission_error(m))
            
        except Exception as e:
            msg = str(e)
            self.append_log(f"\n❌ BŁĄD: {msg}")
            import traceback
            self.append_log(traceback.format_exc())
            
            # Wróć do głównego wątku GUI
            self.after(0, lambda m=msg: self._on_error(m))
    
    def _on_success(self):
        """Callback po udanym przetworzeniu"""
        self.progress.stop()
        self.is_processing = False
        self.btn_generate.config(state="normal")
        self.btn_open_folder.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        
        filename = getattr(self, 'generated_filename', 'LOGISTYKA_OUT.xlsx')
        self.status_var.set(f"✅ Gotowe! Wygenerowano {filename}")
    
    def _on_error(self, error_msg: str):
        """Callback po błędzie"""
        self.progress.stop()
        self.is_processing = False
        self.btn_generate.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.status_var.set("❌ Błąd przetwarzania")
        
        messagebox.showerror(
            "Błąd przetwarzania",
            f"Wystąpił błąd podczas generowania:\n\n{error_msg}\n\n"
            f"Sprawdź logi powyżej."
        )
    
    def _on_permission_error(self, error_msg: str):
        """Callback po błędzie uprawnień (plik otwarty)"""
        self.progress.stop()
        self.is_processing = False
        self.btn_generate.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.status_var.set("❌ Plik jest otwarty - zamknij go i spróbuj ponownie")
        
        # Wyodrębnij nazwę pliku z komunikatu błędu
        import re
        file_match = re.search(r"'([^']+\.xlsx)'", error_msg)
        file_name = file_match.group(1) if file_match else "plik wyjściowy"
        
        messagebox.showerror(
            "Plik jest otwarty",
            f"⚠️ PLIK JEST OTWARTY W EXCELU ⚠️\n\n"
            f"Zamknij plik:\n{file_name}\n\n"
            f"...i uruchom przetwarzanie ponownie.\n\n"
            f"Jeśli plik nie jest otwarty, sprawdź uprawnienia do zapisu."
        )

    def _on_cancel(self, msg: str):
        """Callback po anulowaniu przetwarzania"""
        self.progress.stop()
        self.is_processing = False
        self.btn_generate.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        self.status_var.set("⏹️ Przetwarzanie zatrzymane")

        messagebox.showinfo(
            "Przetwarzanie zatrzymane",
            f"{msg}\n\nMożesz uruchomić przetwarzanie ponownie lub użyć RESET."
        )

        if self.reset_pending:
            self._do_reset()
    
    def _show_errors_window(self, errors: List[List[str]]):
        """Wyświetla okno z listą błędów (tryb szybki)"""
        # Zakończ przetwarzanie
        self.progress.stop()
        self.is_processing = False
        self.btn_generate.config(state="normal")
        self.btn_open_folder.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.cancel_requested = False
        
        if not errors:
            self.status_var.set("✅ Brak błędów - projekt jest poprawny!")
            messagebox.showinfo(
                "Brak błędów",
                "🎉 Gratulacje!\n\n"
                "Nie znaleziono żadnych błędów w projekcie.\n"
                "Wszystkie pliki są poprawne i spójne."
            )
            return
        
        self.status_var.set(f"⚠️ Znaleziono {len(errors)} błędów")
        
        # Utwórz okno z błędami
        error_win = tk.Toplevel(self)
        error_win.title(f"Znalezione błędy ({len(errors)})")
        error_win.geometry("1200x800")
        error_win.transient(self)
        
        # Nagłówek
        header_frame = ttk.Frame(error_win, padding=(10, 10))
        header_frame.pack(fill="x")
        
        ttk.Label(
            header_frame,
            text=f"⚠️ Znaleziono {len(errors)} błędów w projekcie",
            font=("Segoe UI", 12, "bold"),
            foreground="red"
        ).pack(anchor="w")
        
        ttk.Label(
            header_frame,
            text=f"Projekt: {self.project_dir.name}",
            foreground="gray"
        ).pack(anchor="w", pady=(5, 0))
        
        ttk.Separator(error_win, orient="horizontal").pack(fill="x", pady=10)
        
        # Grupuj błędy według typu
        errors_by_type: Dict[str, List[List[str]]] = {}
        for err in errors:
            err_type = err[0] if err else "NIEZNANY"
            errors_by_type.setdefault(err_type, []).append(err)
        
        # FILTRY - Checkboxy dla kategorii błędów
        filter_frame = ttk.LabelFrame(error_win, text=" 🔍 Filtry kategorii błędów ", padding=(10, 10))
        filter_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # Słownik do przechowywania stanów checkboxów
        category_vars: Dict[str, tk.BooleanVar] = {}
        
        # Custom sortowanie - "PLIKI – nieprzypisane do żadnego BOM" zawsze na końcu
        def sort_key_filter(err_type):
            if "nieprzypisane do żadnego BOM" in err_type:
                return ("ZZZ", err_type)
            return ("", err_type)
        
        # Tworzenie checkboxów dla każdej kategorii
        sorted_categories = sorted(errors_by_type.keys(), key=sort_key_filter)
        
        # Podziel na kolumny dla lepszej czytelności (max 3 kolumny)
        num_cols = min(3, len(sorted_categories))
        col_frames = []
        for col_idx in range(num_cols):
            col_frame = ttk.Frame(filter_frame)
            col_frame.pack(side="left", fill="both", expand=True, padx=5)
            col_frames.append(col_frame)
        
        # Rozmieść kategorie po kolumnach
        for idx, err_type in enumerate(sorted_categories):
            col_idx = idx % num_cols
            var = tk.BooleanVar(value=True)  # Domyślnie wszystkie zaznaczone
            category_vars[err_type] = var
            
            count = len(errors_by_type[err_type])
            cb = ttk.Checkbutton(
                col_frames[col_idx],
                text=f"{err_type} ({count})",
                variable=var,
                command=lambda: update_display()
            )
            cb.pack(anchor="w", pady=2)
        
        # Przyciski szybkiego wyboru
        quick_select_frame = ttk.Frame(filter_frame)
        quick_select_frame.pack(fill="x", pady=(10, 0))
        
        def select_all():
            for var in category_vars.values():
                var.set(True)
            update_display()
        
        def deselect_all():
            for var in category_vars.values():
                var.set(False)
            update_display()
        
        ttk.Button(quick_select_frame, text="✓ Zaznacz wszystko", command=select_all).pack(side="left", padx=(0, 5))
        ttk.Button(quick_select_frame, text="✗ Odznacz wszystko", command=deselect_all).pack(side="left")
        
        ttk.Separator(error_win, orient="horizontal").pack(fill="x", pady=10)
        
        # Treść błędów
        text_frame = ttk.Frame(error_win, padding=(10, 0, 10, 10))
        text_frame.pack(fill="both", expand=True)
        
        text_widget = scrolledtext.ScrolledText(
            text_frame,
            wrap="word",
            font=("Consolas", 11)
        )
        text_widget.pack(fill="both", expand=True)
        
        # Tagi do kolorowania (większe czcionki, ciemnoniebieskie kolory zamiast szarych)
        text_widget.tag_config("header", font=("Segoe UI", 13, "bold"), foreground="#d32f2f")
        text_widget.tag_config("separator", foreground="#1565c0")
        text_widget.tag_config("number", foreground="blue", font=("Consolas", 11, "bold"))
        text_widget.tag_config("filename", font=("Consolas", 11, "bold"), foreground="#1976d2")
        text_widget.tag_config("katalog", foreground="#0d47a1", font=("Consolas", 10))
        text_widget.tag_config("files", foreground="#0d47a1", font=("Consolas", 10))
        
        # Funkcja do aktualizacji wyświetlanych błędów
        def update_display():
            """Odśwież wyświetlanie błędów na podstawie zaznaczonych kategorii"""
            text_widget.config(state="normal")
            text_widget.delete("1.0", "end")
            
            # Custom sortowanie - "PLIKI – nieprzypisane do żadnego BOM" zawsze na końcu
            def sort_key(item):
                err_type = item[0]
                if "nieprzypisane do żadnego BOM" in err_type:
                    return ("ZZZ", err_type)  # Przesuń na koniec
                return ("", err_type)  # Normalne sortowanie alfabetyczne
            
            # Filtruj kategorie według zaznaczonych checkboxów
            filtered_errors_by_type = {
                err_type: err_list 
                for err_type, err_list in errors_by_type.items() 
                if category_vars[err_type].get()
            }
            
            # Policz całkowitą liczbę wyświetlanych błędów
            total_displayed = sum(len(err_list) for err_list in filtered_errors_by_type.values())
            
            if not filtered_errors_by_type:
                text_widget.insert("end", "Brak błędów do wyświetlenia.\n\n", "header")
                text_widget.insert("end", "Wszystkie kategorie są odfiltrowane. Zaznacz przynajmniej jedną kategorię aby zobaczyć błędy.")
            else:
                # Wypisz błędy pogrupowane
                for idx, (err_type, err_list) in enumerate(sorted(filtered_errors_by_type.items(), key=sort_key)):
                    if idx > 0:
                        text_widget.insert("end", "\n" + "="*100 + "\n\n")
                    
                    text_widget.insert("end", f"▶ {err_type} ({len(err_list)})\n", "header")
                    text_widget.insert("end", "="*100 + "\n\n", "separator")
                    
                    for err_idx, err in enumerate(err_list, 1):
                        # Format: [Typ, Nazwa pliku, Opis, Lista plików, Katalog]
                        nazwa_pliku = err[1] if len(err) > 1 else ""
                        opis = err[2] if len(err) > 2 else ""
                        pliki = err[3] if len(err) > 3 else ""
                        katalog = err[4] if len(err) > 4 else ""
                        
                        text_widget.insert("end", f"{err_idx}. ", "number")
                        text_widget.insert("end", f"{nazwa_pliku}\n", "filename")
                        
                        if katalog:
                            text_widget.insert("end", f"   Katalog: {katalog}\n", "katalog")
                        
                        if opis:
                            # Formatuj opis z wcięciem
                            for line in opis.split("\n"):
                                text_widget.insert("end", f"   {line}\n")
                        
                        if pliki:
                            text_widget.insert("end", "   Pliki:\n", "files")
                            for line in pliki.split("\n"):
                                if line.strip():
                                    text_widget.insert("end", f"   {line}\n", "files")
                        
                        text_widget.insert("end", "\n")
            
            text_widget.config(state="disabled")
            
            # Aktualizuj tytuł okna
            error_win.title(f"Znalezione błędy ({total_displayed} z {len(errors)} wyświetlanych)")
        
        # Pierwsze wyświetlenie
        update_display()
        
        # Przyciski
        btn_frame = ttk.Frame(error_win, padding=(10, 10))
        btn_frame.pack(fill="x")
        
        def save_to_file():
            """Zapisz błędy do pliku tekstowego (tylko zaznaczone kategorie)"""
            file_path = filedialog.asksaveasfilename(
                title="Zapisz błędy do pliku",
                defaultextension=".txt",
                filetypes=[("Pliki tekstowe", "*.txt"), ("Wszystkie pliki", "*.*")],
                initialfile=f"ERRORS_{self.project_dir.name}.txt"
            )
            
            if not file_path:
                return
            
            try:
                # Filtruj kategorie według zaznaczonych checkboxów
                filtered_errors_by_type = {
                    err_type: err_list 
                    for err_type, err_list in errors_by_type.items() 
                    if category_vars[err_type].get()
                }
                
                total_displayed = sum(len(err_list) for err_list in filtered_errors_by_type.values())
                
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(f"RAPORT BŁĘDÓW - {self.project_dir.name}\n")
                    f.write(f"Data: {Path(file_path).stat().st_mtime}\n")
                    f.write(f"Liczba błędów: {total_displayed} (z {len(errors)} całkowitych)\n")
                    f.write("="*100 + "\n\n")
                    
                    # Custom sortowanie - "PLIKI – nieprzypisane do żadnego BOM" zawsze na końcu
                    def sort_key_file(item):
                        err_type = item[0]
                        if "nieprzypisane do żadnego BOM" in err_type:
                            return ("ZZZ", err_type)
                        return ("", err_type)
                    
                    if not filtered_errors_by_type:
                        f.write("Brak zaznaczonych kategorii błędów do zapisania.\n")
                    else:
                        for idx, (err_type, err_list) in enumerate(sorted(filtered_errors_by_type.items(), key=sort_key_file)):
                            if idx > 0:
                                f.write("\n" + "="*100 + "\n\n")
                            
                            f.write(f"▶ {err_type} ({len(err_list)})\n")
                            f.write("="*100 + "\n\n")
                            
                            for err_idx, err in enumerate(err_list, 1):
                                nazwa_pliku = err[1] if len(err) > 1 else ""
                                opis = err[2] if len(err) > 2 else ""
                                pliki = err[3] if len(err) > 3 else ""
                                katalog = err[4] if len(err) > 4 else ""
                                
                                f.write(f"{err_idx}. {nazwa_pliku}\n")
                                if katalog:
                                    f.write(f"   Katalog: {katalog}\n")
                                if opis:
                                    for line in opis.split("\n"):
                                        f.write(f"   {line}\n")
                                if pliki:
                                    f.write("   Pliki:\n")
                                    for line in pliki.split("\n"):
                                        if line.strip():
                                            f.write(f"   {line}\n")
                                f.write("\n")
                
                messagebox.showinfo("Zapisano", f"Błędy zapisane do:\n{file_path}\n\nZapisano {total_displayed} błędów.")
            except Exception as e:
                messagebox.showerror("Błąd", f"Nie można zapisać pliku:\n{e}")
        
        ttk.Button(btn_frame, text="💾 Zapisz do pliku TXT", command=save_to_file).pack(side="left", padx=(0, 10))
        ttk.Button(btn_frame, text="✖ Zamknij", command=error_win.destroy).pack(side="right")
        
        # Fokus na okno
        error_win.focus_set()
        
        # Wyśrodkuj okno względem rodzica
        error_win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - error_win.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - error_win.winfo_height()) // 2
        error_win.geometry(f"+{max(0, x)}+{max(0, y)}")
        
    def open_project_folder(self):
        """Otwiera folder projektu w eksploratorze plików i podświetla wygenerowany XLSX (jeśli istnieje)"""
        if not self.project_dir or not self.project_dir.exists():
            messagebox.showwarning("Błąd", "Folder projektu nie istnieje!")
            return
        
        try:
            import subprocess
            import platform
            
            # Spróbuj znaleźć wygenerowany plik XLSX
            generated_file = None
            
            # 1. Spróbuj użyć zapisanej nazwy pliku z ostatniego przetwarzania
            if hasattr(self, 'generated_filename') and self.generated_filename:
                generated_file = self.project_dir / self.generated_filename
                if not generated_file.exists():
                    generated_file = None
            
            # 2. Jeśli nie znaleziono, spróbuj znaleźć najnowszy plik XLSX w katalogu projektu
            if not generated_file:
                try:
                    xlsx_files = list(self.project_dir.glob("*.xlsx"))
                    if xlsx_files:
                        # Posortuj po dacie modyfikacji (najnowszy pierwszy)
                        xlsx_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                        generated_file = xlsx_files[0]
                except Exception:
                    generated_file = None
            
            if platform.system() == "Windows":
                if generated_file:
                    # Otwórz z zaznaczonym plikiem
                    subprocess.run(['explorer', '/select,', str(generated_file)], check=False)
                else:
                    # Otwórz sam folder
                    subprocess.run(['explorer', str(self.project_dir)], check=False)
            elif platform.system() == "Darwin":  # macOS
                if generated_file:
                    subprocess.run(['open', '-R', str(generated_file)], check=False)
                else:
                    subprocess.run(['open', str(self.project_dir)], check=False)
            else:  # Linux
                # Linux - większość FM nie wspiera selekcji, otwórz folder
                subprocess.run(['xdg-open', str(self.project_dir)], check=False)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie można otworzyć folderu:\n{e}")
    
    def open_file_location(self, file_path: Path):
        """Otwiera folder zawierający plik i zaznacza go"""
        if not file_path.exists():
            messagebox.showwarning("Błąd", f"Plik nie istnieje:\n{file_path}")
            return
        
        try:
            import subprocess
            import platform
            
            if platform.system() == "Windows":
                # Windows - zaznacza plik w Explorerze
                subprocess.run(['explorer', '/select,', str(file_path)], check=False)
            elif platform.system() == "Darwin":  # macOS
                # macOS - zaznacza plik w Finderze
                subprocess.run(['open', '-R', str(file_path)], check=False)
            else:  # Linux
                # Linux - otwiera folder (większość FM nie wspiera selekcji)
                subprocess.run(['xdg-open', str(file_path.parent)], check=False)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie można otworzyć lokalizacji:\n{e}")
    
    def _load_last_path(self):
        """Wczytuje ostatnio używaną ścieżkę z pliku konfiguracyjnego"""
        try:
            if GUI_CONFIG_FILE.exists():
                with open(GUI_CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    last_path_str = config.get('last_project_path')
                    if last_path_str:
                        self.last_path = Path(last_path_str)
        except Exception:
            pass  # Ignoruj błędy ładowania
    
    def _save_last_path(self, path: Path):
        """Zapisuje ostatnio używaną ścieżkę do pliku konfiguracyjnego"""
        try:
            config = {'last_project_path': str(path)}
            with open(GUI_CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
            self.last_path = path
        except Exception:
            pass  # Ignoruj błędy zapisu


if __name__ == "__main__":
    main()