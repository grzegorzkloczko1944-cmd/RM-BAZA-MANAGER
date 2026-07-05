r"""
============================================================================
Parser_RM_BAZA - wykrywanie anomalii cenowych elementów handlowych
============================================================================
Samodzielny, READ-ONLY skrypt. Nie modyfikuje RM_BAZA_v15*.py ani schematu SQL.

Co robi:
1. Łączy się z master.sqlite (read-only) i wszystkimi bazami projektowymi
   (read-only, przez DatabaseManager.open_project_remote - tryb mode=ro).
2. Zbiera pozycje klasy STANDARD / ZNORMALIZOWANE (elementy handlowe) ze
   wszystkich projektów.
3. Grupuje je fuzzy-matchingiem po znormalizowanej nazwie (bez AI) - np.
   różne zapisy "łożysko 6004 INOX" trafiają do jednej grupy, a warianty
   różniące się słowem kluczowym (INOX / 2RS / ...) NIE są ze sobą mieszane.
4. W obrębie każdej grupy liczy statystyki cenowe (min/mediana/max) i
   wypisuje grupy, gdzie cena skrajna odstaje od mediany grupy powyżej
   zadanego progu.
5. Wypisuje wynik na konsolę i eksportuje do XLSX.

Uruchomienie:
    python Parser_RM_BAZA.py
    python Parser_RM_BAZA.py --threshold 1.4 --min-group-size 3
    python Parser_RM_BAZA.py --out C:\raporty\anomalie.xlsx
============================================================================
"""

import argparse
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from statistics import median

from database_manager import DatabaseManager

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    openpyxl = None

# Ścieżki bazy RM_BAZA
MASTER_PATH = r"C:\RMPAK_CLIENT\RM_MANAGER\RM_BAZA\master.sqlite"
PROJECTS_DIR = r"C:\RMPAK_CLIENT\RM_MANAGER\RM_BAZA\projects"
LOCAL_DIR = r"C:/RMPAK_CLIENT/_parser_rm_baza_tmp"

# Klasy uznawane za "element handlowy"
COMMERCIAL_CLASSES = {"STANDARD", "ZNORMALIZOWANE"}

# Słowa-warianty, które NIE mogą się mieszać w jednej grupie mimo podobieństwa
# tekstowego (np. "łożysko 6004" i "łożysko 6004 INOX" to różne pozycje cenowo)
VARIANT_KEYWORDS = ["inox", "nierdzewn", "2rs", "2z", "zz", "ocynk", "czarn"]

DEFAULT_THRESHOLD = 1.4  # cena_max / mediana grupy powyżej tego = anomalia
DEFAULT_MIN_GROUP_SIZE = 2  # od ilu wystąpień w grupie liczymy statystykę


@dataclass
class ItemRow:
    project_id: int
    project_name: str
    item_id: int
    name: str
    drawing_no: str
    price_pln: float  # cena JEDNOSTKOWA (za 1 szt.) - tak jest w RM_BAZA (SUM(price_pln * qty) = wartość)
    qty: float
    supplier_id: int
    supplier_name: str
    class_effective: str

    @property
    def unit_price(self):
        return self.price_pln


@dataclass
class Group:
    key: str
    variant_tag: str
    items: list = field(default_factory=list)

    @property
    def prices(self):
        return [it.unit_price for it in self.items if it.unit_price is not None and it.unit_price > 0]


def normalize_name(name: str) -> str:
    """Znormalizuj nazwę do porównań: lowercase, usuń znaki specjalne, ujednolić spacje."""
    if not name:
        return ""
    text = name.lower()
    text = text.replace("ł", "l").replace("ą", "a").replace("ę", "e").replace("ó", "o") \
        .replace("ś", "s").replace("ć", "c").replace("ż", "z").replace("ź", "z").replace("ń", "n")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


# Wzorzec oznaczeń katalogowych typu FESTO/SMC: litery-liczba-liczba(-litery/liczby...)
# np. "dfm 20 40 p a gf", "lbn 20 25", "dsnu 25 100 pps a" (po normalize_name, myślniki
# już zamienione na spacje) - liczby po prefiksie to realne wymiary (średnica, skok)
# różnicujące cenę, NIE wolno ich usuwać jako "wariant" ani rozmywać fuzzy matchingiem:
# "dfm 20 40" i "dfm 20 20" to różne, nieporównywalne produkty.
# Kotwiczone na POCZĄTKU nazwy (ew. po jednym numerycznym prefiksie zamówienia/indeksu),
# żeby nie łapać przypadkowego fragmentu w środku dłuższego opisu słownego
# (np. "Obejma ... (zakres 36-39)" nie jest kodem katalogowym).
CATALOG_CODE_RE = re.compile(r"^(?:\d+\s)?([a-z]{2,}\d*(?:\s\d+){2,})")


def extract_catalog_code(normalized: str) -> str:
    """Zwróć kod katalogowy (litery + WSZYSTKIE kolejne liczby) z POCZĄTKU nazwy,
    pomijając ewentualny numeryczny prefiks zamówienia/indeksu
    (np. "170843 dfm 20 40 p a gf" -> "dfm 20 40").

    Bierze wszystkie kolejne liczby, nie tylko pierwsze dwie - kody wieloczłonowe jak
    "GN 113.4-6-20" (po normalize_name: "gn 113 4 6 20") mają 4 liczby, a obcięcie do
    dwóch ("gn 113 4") zlepiłoby różne warianty (6-20 vs 12-35) w jedną grupę.

    Pusty string jeśli nazwa nie pasuje do wzorca kodu katalogowego.
    """
    match = CATALOG_CODE_RE.match(normalized)
    if not match:
        return ""
    return match.group(1)


def is_catalog_code(normalized: str) -> bool:
    """Rozpoznaj, czy znormalizowana nazwa zawiera kod katalogowy z wymiarami (litery-liczba-liczba)."""
    return bool(extract_catalog_code(normalized))


def extract_variant_tag(normalized: str) -> str:
    """Wyodrębnij tag wariantu (np. 'inox') z znormalizowanej nazwy.

    "inox" i "nierdzewn" to synonimy tego samego wariantu materiałowego,
    dlatego oba mapują na wspólny tag "inox" zamiast tworzyć osobne grupy.
    """
    synonyms = {"nierdzewn": "inox"}
    found = set()
    for kw in VARIANT_KEYWORDS:
        if kw in normalized:
            found.add(synonyms.get(kw, kw))
    return "+".join(sorted(found)) if found else "(standard)"


def extract_base_tokens(normalized: str) -> str:
    """Usuń tokeny-warianty z nazwy, zostaw 'bazowy' identyfikator do grupowania."""
    tokens = [t for t in normalized.split(" ") if not any(kw in t for kw in VARIANT_KEYWORDS)]
    return " ".join(tokens)


def token_similarity(a: str, b: str) -> float:
    """Podobieństwo dwóch znormalizowanych nazw na podstawie wspólnych tokenów (Jaccard)."""
    ta = {t for t in a.split(" ") if t}
    tb = {t for t in b.split(" ") if t}
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def group_items(items: list, similarity_threshold: float = 0.6) -> list:
    """Grupuj pozycje po podobieństwie bazowej nazwy + identycznym tagu wariantu.

    Dwa tryby dopasowania:
    - Kod katalogowy (np. "DFM-20-40-P-A-GF") -> klucz to CAŁY znormalizowany kod,
      dopasowanie dokładne (exact match). Liczby to realne wymiary (średnica, skok) -
      różne wymiary = różny, nieporównywalny produkt, więc żadne fuzzy/usuwanie tokenów.
    - Nazwa opisowa (np. "Łożysko 6004 INOX") -> dotychczasowy fuzzy matching po
      tokenach z wariantem (INOX/2RS/...) wyodrębnionym osobno.

    Prosty jednoprzebiegowy clustering: każda nowa pozycja dołącza do pierwszej
    pasującej grupy, w przeciwnym razie zakłada nową grupę.
    """
    groups: list = []
    for it in items:
        norm = normalize_name(it.name)

        catalog_code = extract_catalog_code(norm)
        if catalog_code:
            base = catalog_code
            variant = "(kod katalogowy)"
            match_exact = True
        else:
            base = extract_base_tokens(norm)
            variant = extract_variant_tag(norm)
            match_exact = False

        target = None
        for g in groups:
            if g.variant_tag != variant:
                continue
            if match_exact:
                if base == g.key:
                    target = g
                    break
            elif token_similarity(base, g.key) >= similarity_threshold:
                target = g
                break

        if target is None:
            target = Group(key=base, variant_tag=variant)
            groups.append(target)
        target.items.append(it)

    return groups


def load_commercial_items(db: DatabaseManager) -> list:
    """Pobierz pozycje klasy handlowej ze wszystkich projektów (read-only)."""
    all_items = []
    projects = db.get_projects()
    suppliers = {sid: sname for sid, sname in db.get_suppliers()}

    for project_id, project_name, active, project_type in projects:
        if project_type == "WAREHOUSE":
            continue

        try:
            db.open_project_remote(project_id, project_type)
        except FileNotFoundError:
            print(f"  ⚠️  Pomijam projekt {project_name} ({project_id}): brak pliku bazy")
            continue
        except Exception as e:
            print(f"  ⚠️  Pomijam projekt {project_name} ({project_id}): {e}")
            continue

        try:
            rows = db.get_project_items(project_id, show_hidden=False)
        except Exception as e:
            print(f"  ⚠️  Błąd odczytu pozycji projektu {project_name} ({project_id}): {e}")
            continue

        for r in rows:
            class_eff = r["class_effective"]
            if class_eff not in COMMERCIAL_CLASSES:
                continue
            price = r["price_pln"]
            if price is None:
                continue
            # qty potrzebne tylko do kolumny "Wartość razem" (price_pln * qty) w raporcie -
            # price_pln to już cena jednostkowa, nie dzielimy przez qty
            qty = r["order_qty"] if r["order_qty"] not in (None, 0) else r["qty_bom"]
            if qty in (None, 0):
                qty = 1
            all_items.append(ItemRow(
                project_id=project_id,
                project_name=project_name,
                item_id=r["id"],
                name=r["name"] or "",
                drawing_no=r["drawing_no"] or "",
                price_pln=float(price),
                qty=float(qty),
                supplier_id=r["supplier_id"],
                supplier_name=suppliers.get(r["supplier_id"], "(brak)"),
                class_effective=class_eff,
            ))

    return all_items


def find_anomalies(groups: list, threshold: float, min_group_size: int) -> list:
    """Zwróć grupy, w których cena maksymalna przekracza medianę * threshold."""
    anomalies = []
    for g in groups:
        prices = g.prices
        if len(prices) < min_group_size:
            continue
        med = median(prices)
        if med <= 0:
            continue
        max_price = max(prices)
        ratio = max_price / med
        if ratio >= threshold:
            anomalies.append((g, med, max_price, ratio))
    anomalies.sort(key=lambda x: x[3], reverse=True)
    return anomalies


def print_report(anomalies: list):
    if not anomalies:
        print("\nBrak wykrytych anomalii cenowych przy zadanym progu.")
        return

    print(f"\n{'='*100}")
    print(f"WYKRYTE ANOMALIE CENOWE - {len(anomalies)} grup(y)")
    print(f"{'='*100}")

    for g, med, max_price, ratio in anomalies:
        print(f"\nGrupa: \"{g.key}\" [wariant: {g.variant_tag}]  ({len(g.items)} pozycji)")
        print(f"  Mediana ceny jedn.: {med:.2f} PLN   Cena jedn. max: {max_price:.2f} PLN   Odchylenie: {ratio:.2f}x")
        for it in sorted(g.items, key=lambda x: x.unit_price or 0, reverse=True):
            up = it.unit_price
            if up is None:
                continue
            flag = "  <-- SPRAWDŹ" if up == max_price and ratio >= 1.0 else ""
            print(f"    - {up:>8.2f} PLN/szt (qty={it.qty:g}, razem={up * it.qty:.2f} PLN) | "
                  f"projekt={it.project_name:<25} dostawca={it.supplier_name:<20} "
                  f"rysunek={it.drawing_no} nazwa=\"{it.name}\"{flag}")


def export_xlsx(anomalies: list, out_path: str):
    if openpyxl is None:
        print("\n⚠️  Brak biblioteki openpyxl - pomijam eksport XLSX (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalie cenowe"

    headers = ["Grupa", "Wariant", "Mediana ceny jedn. [PLN]", "Cena jedn. [PLN]", "Odchylenie",
               "Ilość", "Wartość razem [PLN]", "Projekt", "Dostawca", "Nr rysunku", "Nazwa pozycji", "Item ID"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for g, med, max_price, ratio in anomalies:
        for it in sorted(g.items, key=lambda x: x.unit_price or 0, reverse=True):
            up = it.unit_price
            if up is None:
                continue
            row = [
                g.key, g.variant_tag, round(med, 2), round(up, 2),
                round(up / med, 2) if med else "",
                it.qty, round(up * it.qty, 2),
                it.project_name, it.supplier_name, it.drawing_no, it.name, it.item_id,
            ]
            ws.append(row)
            if up == max_price:
                for cell in ws[ws.max_row]:
                    cell.fill = highlight

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 50)

    wb.save(out_path)
    print(f"\n✅ Zapisano raport XLSX: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Parser_RM_BAZA - anomalie cenowe elementów handlowych")
    parser.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD,
                         help=f"Próg odchylenia cena_max/mediana (domyślnie {DEFAULT_THRESHOLD})")
    parser.add_argument("--min-group-size", type=int, default=DEFAULT_MIN_GROUP_SIZE,
                         help=f"Minimalna liczba pozycji w grupie do analizy (domyślnie {DEFAULT_MIN_GROUP_SIZE})")
    parser.add_argument("--out", type=str, default=None,
                         help="Ścieżka pliku XLSX wyjściowego (domyślnie Parser_RM_BAZA_<data>.xlsx)")
    args = parser.parse_args()

    out_path = args.out or f"Parser_RM_BAZA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print("Parser_RM_BAZA - start (READ-ONLY)")
    print(f"Master: {MASTER_PATH}")
    print(f"Projects dir: {PROJECTS_DIR}")

    db = DatabaseManager(master_path=MASTER_PATH, projects_dir=PROJECTS_DIR, local_dir=LOCAL_DIR)

    if not db.connect_master():
        print("❌ Nie udało się połączyć z master.sqlite")
        sys.exit(1)

    print("\nWczytywanie pozycji handlowych (STANDARD / ZNORMALIZOWANE) ze wszystkich projektów...")
    items = load_commercial_items(db)
    print(f"Wczytano {len(items)} pozycji handlowych z cenami.")

    print("\nGrupowanie po podobieństwie nazwy...")
    groups = group_items(items)
    print(f"Utworzono {len(groups)} grup.")

    anomalies = find_anomalies(groups, threshold=args.threshold, min_group_size=args.min_group_size)
    print_report(anomalies)
    export_xlsx(anomalies, out_path)

    db.close_all()


if __name__ == "__main__":
    main()
